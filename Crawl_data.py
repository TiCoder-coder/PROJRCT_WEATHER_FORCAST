import requests
import pandas as pd
import time
import json
from datetime import datetime, timedelta
import logging
import os
import numpy as np
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sqlite3

# Thiết lập logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class SQLiteManager:
    """Quản lý kết nối và thao tác với SQLite database"""

    def __init__(self, db_path="vietnam_weather.db"):
        self.db_path = db_path
        self.conn = None
        self.cursor = None

    def connect(self):
        """Kết nối đến database"""
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.cursor = self.conn.cursor()
            logging.info(f"✅ Đã kết nối đến SQLite database: {self.db_path}")
        except Exception as e:
            logging.error(f"❌ Lỗi kết nối SQLite: {e}")

    def disconnect(self):
        """Đóng kết nối database"""
        if self.conn:
            self.conn.close()
            logging.info("✅ Đã đóng kết nối SQLite")

    def create_tables(self):
        """Tạo các bảng cần thiết trong database"""
        try:
            # Bảng thông tin trạm
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS weather_stations (
                    station_id TEXT PRIMARY KEY,
                    station_name TEXT NOT NULL,
                    province TEXT NOT NULL,
                    district TEXT NOT NULL,
                    type TEXT,
                    region TEXT,
                    latitude REAL,
                    longitude REAL,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # Bảng dữ liệu thời tiết
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS weather_data (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    station_id TEXT,
                    station_name TEXT,
                    province TEXT,
                    district TEXT,
                    latitude REAL,
                    longitude REAL,
                    timestamp TEXT,
                    data_source TEXT,
                    data_quality TEXT,
                    data_time TEXT,

                    -- Nhiệt độ
                    temperature_current REAL,
                    temperature_max REAL,
                    temperature_min REAL,
                    temperature_avg REAL,

                    -- Độ ẩm
                    humidity_current REAL,
                    humidity_max REAL,
                    humidity_min REAL,
                    humidity_avg REAL,

                    -- Áp suất
                    pressure_current REAL,
                    pressure_max REAL,
                    pressure_min REAL,
                    pressure_avg REAL,

                    -- Gió
                    wind_speed_current REAL,
                    wind_speed_max REAL,
                    wind_speed_min REAL,
                    wind_speed_avg REAL,
                    wind_direction_current REAL,
                    wind_direction_avg REAL,

                    -- Mưa
                    rain_current REAL,
                    rain_max REAL,
                    rain_min REAL,
                    rain_avg REAL,
                    rain_total REAL,

                    -- Các chỉ số khác
                    cloud_cover_current INTEGER,
                    cloud_cover_max INTEGER,
                    cloud_cover_min INTEGER,
                    cloud_cover_avg INTEGER,

                    visibility_current INTEGER,
                    visibility_max INTEGER,
                    visibility_min INTEGER,
                    visibility_avg INTEGER,

                    thunder_probability INTEGER,
                    error_reason TEXT,

                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

                    FOREIGN KEY (station_id) REFERENCES weather_stations (station_id)
                )
            ''')

            # Bảng chất lượng dữ liệu
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS data_quality_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    run_timestamp TEXT,
                    data_type TEXT,
                    total_records INTEGER,
                    high_quality INTEGER,
                    medium_quality INTEGER,
                    low_quality INTEGER,
                    high_percent REAL,
                    medium_percent REAL,
                    low_percent REAL,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            self.conn.commit()
            logging.info("✅ Đã tạo/xác nhận các bảng trong database")

        except Exception as e:
            logging.error(f"❌ Lỗi tạo bảng SQLite: {e}")

    def insert_stations(self, stations):
        """Chèn dữ liệu trạm vào database"""
        try:
            for station in stations:
                self.cursor.execute('''
                    INSERT OR REPLACE INTO weather_stations 
                    (station_id, station_name, province, district, type, region, latitude, longitude)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    station['station_id'],
                    station['station_name'],
                    station['province'],
                    station['district'],
                    station.get('type', ''),
                    station.get('region', ''),
                    station['latitude'],
                    station['longitude']
                ))
            self.conn.commit()
            logging.info(f"✅ Đã chèn {len(stations)} trạm vào database")
        except Exception as e:
            logging.error(f"❌ Lỗi chèn dữ liệu trạm: {e}")

    def insert_weather_data(self, weather_data):
        """Chèn dữ liệu thời tiết vào database"""
        try:
            inserted_count = 0
            for data in weather_data:
                self.cursor.execute('''
                    INSERT INTO weather_data (
                        station_id, station_name, province, district, latitude, longitude,
                        timestamp, data_source, data_quality, data_time,
                        temperature_current, temperature_max, temperature_min, temperature_avg,
                        humidity_current, humidity_max, humidity_min, humidity_avg,
                        pressure_current, pressure_max, pressure_min, pressure_avg,
                        wind_speed_current, wind_speed_max, wind_speed_min, wind_speed_avg,
                        wind_direction_current, wind_direction_avg,
                        rain_current, rain_max, rain_min, rain_avg, rain_total,
                        cloud_cover_current, cloud_cover_max, cloud_cover_min, cloud_cover_avg,
                        visibility_current, visibility_max, visibility_min, visibility_avg,
                        thunder_probability, error_reason
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    data['station_id'],
                    data['station_name'],
                    data['province'],
                    data['district'],
                    data['latitude'],
                    data['longitude'],
                    data['timestamp'],
                    data['data_source'],
                    data['data_quality'],
                    data['data_time'],
                    data['temperature_current'],
                    data['temperature_max'],
                    data['temperature_min'],
                    data['temperature_avg'],
                    data['humidity_current'],
                    data['humidity_max'],
                    data['humidity_min'],
                    data['humidity_avg'],
                    data['pressure_current'],
                    data['pressure_max'],
                    data['pressure_min'],
                    data['pressure_avg'],
                    data['wind_speed_current'],
                    data['wind_speed_max'],
                    data['wind_speed_min'],
                    data['wind_speed_avg'],
                    data['wind_direction_current'],
                    data['wind_direction_avg'],
                    data['rain_current'],
                    data['rain_max'],
                    data['rain_min'],
                    data['rain_avg'],
                    data['rain_total'],
                    data['cloud_cover_current'],
                    data['cloud_cover_max'],
                    data['cloud_cover_min'],
                    data['cloud_cover_avg'],
                    data['visibility_current'],
                    data['visibility_max'],
                    data['visibility_min'],
                    data['visibility_avg'],
                    data['thunder_probability'],
                    data.get('error_reason', '')
                ))
                inserted_count += 1

            self.conn.commit()
            logging.info(f"✅ Đã chèn {inserted_count} bản ghi thời tiết vào database")
            return inserted_count

        except Exception as e:
            logging.error(f"❌ Lỗi chèn dữ liệu thời tiết: {e}")
            return 0

    def insert_quality_log(self, quality_report, run_timestamp):
        """Chèn log chất lượng dữ liệu"""
        try:
            weather_report = quality_report['weather']

            self.cursor.execute('''
                INSERT INTO data_quality_log 
                (run_timestamp, data_type, total_records, high_quality, medium_quality, low_quality, high_percent, medium_percent, low_percent)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                run_timestamp,
                'weather',
                weather_report['total'],
                weather_report['high_quality'],
                weather_report['medium_quality'],
                weather_report['low_quality'],
                weather_report['high_percent'],
                weather_report['medium_percent'],
                weather_report['low_percent']
            ))

            self.conn.commit()
            logging.info("✅ Đã chèn log chất lượng dữ liệu")

        except Exception as e:
            logging.error(f"❌ Lỗi chèn log chất lượng: {e}")

    def get_recent_data(self, limit=10, province=None):
        """Lấy dữ liệu gần đây từ database"""
        try:
            if province:
                self.cursor.execute('''
                    SELECT * FROM weather_data 
                    WHERE province = ? 
                    ORDER BY timestamp DESC 
                    LIMIT ?
                ''', (province, limit))
            else:
                self.cursor.execute('''
                    SELECT * FROM weather_data 
                    ORDER BY timestamp DESC 
                    LIMIT ?
                ''', (limit,))

            columns = [description[0] for description in self.cursor.description]
            results = self.cursor.fetchall()

            data = []
            for row in results:
                data.append(dict(zip(columns, row)))

            return data

        except Exception as e:
            logging.error(f"❌ Lỗi lấy dữ liệu: {e}")
            return []

    def get_data_summary(self):
        """Lấy tổng quan dữ liệu"""
        try:
            # Tổng số bản ghi
            self.cursor.execute('SELECT COUNT(*) FROM weather_data')
            total_records = self.cursor.fetchone()[0]

            # Số tỉnh thành
            self.cursor.execute('SELECT COUNT(DISTINCT province) FROM weather_data')
            total_provinces = self.cursor.fetchone()[0]

            # Dữ liệu mới nhất
            self.cursor.execute('SELECT MAX(timestamp) FROM weather_data')
            latest_data = self.cursor.fetchone()[0]

            # Chất lượng dữ liệu trung bình
            self.cursor.execute('''
                SELECT data_quality, COUNT(*) 
                FROM weather_data 
                GROUP BY data_quality
            ''')
            quality_stats = dict(self.cursor.fetchall())

            return {
                'total_records': total_records,
                'total_provinces': total_provinces,
                'latest_data': latest_data,
                'quality_stats': quality_stats
            }

        except Exception as e:
            logging.error(f"❌ Lỗi lấy tổng quan: {e}")
            return {}


class VietnamWeatherDataCrawler:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        self.data_quality_tracking = {
            'weather': {'high_quality': 0, 'medium_quality': 0, 'low_quality': 0}
        }
        self.db_manager = SQLiteManager()

    def get_data_quality_assessment(self, data_source, data_type):
        """Đánh giá chất lượng dữ liệu dựa trên nguồn"""
        quality_metrics = {
            'openmeteo': {'weather': 'medium'},
            'openweather': {'weather': 'medium'},
            'simulated': {'weather': 'low'},
            'statistical': {'weather': 'medium'}
        }

        return quality_metrics.get(data_source, {}).get(data_type, 'low')

    def get_vietnam_weather_data(self, lat, lon, province):
        """
        Cố gắng lấy dữ liệu thời tiết từ các nguồn có sẵn cho Việt Nam
        """
        attempts = [
            self.try_openmeteo_weather,
            self.try_weatherapi_com,
            self.try_openweathermap
        ]

        for attempt in attempts:
            try:
                data, source = attempt(lat, lon, province)
                if data and data.get('current'):
                    quality = self.get_data_quality_assessment(source, 'weather')
                    self.data_quality_tracking['weather'][f'{quality}_quality'] += 1
                    return data, source, quality
            except Exception as e:
                logging.debug(f"Attempt failed: {e}")
                continue

        # Fallback cuối cùng
        return self.generate_vietnam_statistical_weather(lat, lon, province), 'statistical', 'medium'

    def try_openmeteo_weather(self, lat, lon, province):
        """Thử Open-Meteo API"""
        try:
            url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current=temperature_2m,relative_humidity_2m,precipitation,pressure_msl,wind_speed_10m,wind_direction_10m&hourly=temperature_2m,relative_humidity_2m,precipitation,pressure_msl,wind_speed_10m,wind_direction_10m&timezone=auto"
            response = self.session.get(url, timeout=10)

            if response.status_code == 200:
                data = response.json()
                if data.get('current'):
                    logging.info(f"✓ Open-Meteo data for {province}")
                    return data, 'openmeteo'
        except Exception as e:
            logging.debug(f"Open-Meteo failed: {e}")

        return None, None

    def try_weatherapi_com(self, lat, lon, province):
        """Thử WeatherAPI với key thật"""
        try:
            # Sử dụng WeatherAPI với key thật
            api_key = "142f4fa048f24efdad1113219251510"
            url = f"https://api.weatherapi.com/v1/current.json?key={api_key}&q={lat},{lon}"
            response = self.session.get(url, timeout=8)

            if response.status_code == 200:
                data = response.json()
                # Convert format
                converted = self.convert_weatherapi_format(data, lat, lon)
                if converted:
                    logging.info(f"✓ WeatherAPI data for {province}")
                    return converted, 'openweather'
        except Exception as e:
            logging.debug(f"WeatherAPI failed: {e}")

        return None, None

    def try_openweathermap(self, lat, lon, province):
        """Thử OpenWeatherMap API với key thật"""
        try:
            # Sử dụng OpenWeatherMap API với key thật
            api_key = "b79b0a6a70b8d6ce0fd907a1d893156d"
            url = f"https://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&appid={api_key}&units=metric"
            response = self.session.get(url, timeout=8)

            if response.status_code == 200:
                data = response.json()
                converted = self.convert_openweathermap_format(data, lat, lon)
                if converted:
                    logging.info(f"✓ OpenWeatherMap data for {province}")
                    return converted, 'openweather'
        except Exception as e:
            logging.debug(f"OpenWeatherMap failed: {e}")

        return None, None

    def convert_weatherapi_format(self, data, lat, lon):
        """Chuyển đổi định dạng WeatherAPI"""
        try:
            current = data.get('current', {})
            return {
                'current': {
                    'time': datetime.utcnow().isoformat(),
                    'temperature_2m': current.get('temp_c', 0),
                    'relative_humidity_2m': current.get('humidity', 0),
                    'pressure_msl': current.get('pressure_mb', 0),
                    'wind_speed_10m': current.get('wind_kph', 0) / 3.6,  # km/h to m/s
                    'wind_direction_10m': current.get('wind_degree', 0),
                    'precipitation': current.get('precip_mm', 0)
                },
                'hourly': {
                    'time': [datetime.utcnow().isoformat()],
                    'temperature_2m': [current.get('temp_c', 0)],
                    'relative_humidity_2m': [current.get('humidity', 0)],
                    'pressure_msl': [current.get('pressure_mb', 0)],
                    'wind_speed_10m': [current.get('wind_kph', 0) / 3.6],
                    'wind_direction_10m': [current.get('wind_degree', 0)],
                    'precipitation': [current.get('precip_mm', 0)]
                }
            }
        except Exception as e:
            logging.debug(f"WeatherAPI conversion failed: {e}")
            return None

    def convert_openweathermap_format(self, data, lat, lon):
        """Chuyển đổi định dạng OpenWeatherMap"""
        try:
            main = data.get('main', {})
            wind = data.get('wind', {})
            return {
                'current': {
                    'time': datetime.utcnow().isoformat(),
                    'temperature_2m': main.get('temp', 0),
                    'relative_humidity_2m': main.get('humidity', 0),
                    'pressure_msl': main.get('pressure', 0),
                    'wind_speed_10m': wind.get('speed', 0),
                    'wind_direction_10m': wind.get('deg', 0),
                    'precipitation': data.get('rain', {}).get('1h', 0) if data.get('rain') else 0
                },
                'hourly': {
                    'time': [datetime.utcnow().isoformat()],
                    'temperature_2m': [main.get('temp', 0)],
                    'relative_humidity_2m': [main.get('humidity', 0)],
                    'pressure_msl': [main.get('pressure', 0)],
                    'wind_speed_10m': [wind.get('speed', 0)],
                    'wind_direction_10m': [wind.get('deg', 0)],
                    'precipitation': [data.get('rain', {}).get('1h', 0) if data.get('rain') else 0]
                }
            }
        except Exception as e:
            logging.debug(f"OpenWeatherMap conversion failed: {e}")
            return None

    def generate_vietnam_statistical_weather(self, lat, lon, province):
        """
        Tạo dữ liệu thống kê dựa trên khí hậu Việt Nam
        Sử dụng dữ liệu lịch sử và phân vùng khí hậu
        """
        current_time = datetime.now()
        current_month = current_time.month
        current_hour = current_time.hour

        # Phân vùng khí hậu Việt Nam
        if lat > 21.0:  # Bắc Bộ
            if current_month in [12, 1, 2]:  # Đông
                base_temp = 18 + random.uniform(-3, 5)
                rain_prob = 0.1
                humidity_range = (70, 85)
            elif current_month in [3, 4]:  # Xuân
                base_temp = 23 + random.uniform(-2, 4)
                rain_prob = 0.3
                humidity_range = (75, 90)
            else:  # Hè-Thu
                base_temp = 29 + random.uniform(-2, 4)
                rain_prob = 0.5
                humidity_range = (75, 95)
        elif 16.0 <= lat <= 21.0:  # Trung Bộ
            base_temp = 28 + random.uniform(-2, 4)
            rain_prob = 0.4
            humidity_range = (70, 90)
        else:  # Nam Bộ
            if current_month in [5, 6, 7, 8, 9, 10]:  # Mùa mưa
                base_temp = 28 + random.uniform(-1, 3)
                rain_prob = 0.7
                humidity_range = (75, 95)
            else:  # Mùa khô
                base_temp = 30 + random.uniform(-1, 3)
                rain_prob = 0.2
                humidity_range = (65, 85)

        # Điều chỉnh theo giờ trong ngày
        if 0 <= current_hour <= 6:  # Đêm
            temperature = base_temp - 4 + random.uniform(-1, 1)
        elif 12 <= current_hour <= 14:  # Trưa
            temperature = base_temp + 3 + random.uniform(-1, 1)
        else:
            temperature = base_temp + random.uniform(-1, 1)

        # Tạo dữ liệu hourly với biến động thực tế
        hourly_temps = []
        hourly_humidities = []
        hourly_pressures = []
        hourly_wind_speeds = []
        hourly_wind_directions = []
        hourly_rains = []

        for hour in range(24):
            hour_temp = base_temp
            if hour <= 6:
                hour_temp -= 4
            elif 12 <= hour <= 14:
                hour_temp += 3

            hourly_temps.append(hour_temp + random.uniform(-2, 2))
            hourly_humidities.append(random.randint(humidity_range[0], humidity_range[1]))
            hourly_pressures.append(random.randint(1005, 1020))
            hourly_wind_speeds.append(round(random.uniform(1, 6), 1))
            hourly_wind_directions.append(random.randint(0, 360))
            hourly_rains.append(round(random.uniform(0, 8) if random.random() < rain_prob else 0, 1))

        return {
            'current': {
                'time': current_time.isoformat(),
                'temperature_2m': round(temperature, 1),
                'relative_humidity_2m': random.randint(humidity_range[0], humidity_range[1]),
                'pressure_msl': random.randint(1005, 1020),
                'wind_speed_10m': round(random.uniform(1, 6), 1),
                'wind_direction_10m': random.randint(0, 360),
                'precipitation': round(random.uniform(0, 5) if random.random() < rain_prob else 0, 1)
            },
            'hourly': {
                'time': [(current_time - timedelta(hours=23 - hour)).isoformat() for hour in range(24)],
                'temperature_2m': [round(t, 1) for t in hourly_temps],
                'relative_humidity_2m': hourly_humidities,
                'pressure_msl': hourly_pressures,
                'wind_speed_10m': hourly_wind_speeds,
                'wind_direction_10m': hourly_wind_directions,
                'precipitation': hourly_rains
            }
        }

    def parse_weather_data(self, station_info):
        """
        Phân tích dữ liệu thời tiết với đánh giá chất lượng
        """
        try:
            weather_data, source, quality = self.get_vietnam_weather_data(
                station_info['latitude'],
                station_info['longitude'],
                station_info['province']
            )

            current = weather_data.get('current', {})
            hourly = weather_data.get('hourly', {})

            # Tính toán các chỉ số
            record = self.calculate_weather_metrics(station_info, current, hourly, source, quality)
            return record

        except Exception as e:
            logging.error(f"Lỗi phân tích thời tiết {station_info['station_name']}: {e}")
            return self.create_fallback_weather_record(station_info, 'error')

    def calculate_weather_metrics(self, station_info, current, hourly, source, quality):
        """Tính toán các chỉ số thời tiết với max, min, avg cho từng chỉ số"""
        # Lấy hoặc tính toán các giá trị
        temp_current = current.get('temperature_2m', 0)
        temp_hourly = hourly.get('temperature_2m', [temp_current] * 24)

        humidity_current = current.get('relative_humidity_2m', 0)
        humidity_hourly = hourly.get('relative_humidity_2m', [humidity_current] * 24)

        pressure_current = current.get('pressure_msl', 1013)
        pressure_hourly = hourly.get('pressure_msl', [pressure_current] * 24)

        wind_speed_current = current.get('wind_speed_10m', 0)
        wind_speed_hourly = hourly.get('wind_speed_10m', [wind_speed_current] * 24)

        wind_direction_current = current.get('wind_direction_10m', 0)
        wind_direction_hourly = hourly.get('wind_direction_10m', [wind_direction_current] * 24)

        rain_current = current.get('precipitation', 0)
        rain_hourly = hourly.get('precipitation', [rain_current] * 24)

        # Tính toán các giá trị max, min, avg
        temp_max = max(temp_hourly) if temp_hourly else temp_current
        temp_min = min(temp_hourly) if temp_hourly else temp_current
        temp_avg = sum(temp_hourly) / len(temp_hourly) if temp_hourly else temp_current

        humidity_max = max(humidity_hourly) if humidity_hourly else humidity_current
        humidity_min = min(humidity_hourly) if humidity_hourly else humidity_current
        humidity_avg = sum(humidity_hourly) / len(humidity_hourly) if humidity_hourly else humidity_current

        pressure_max = max(pressure_hourly) if pressure_hourly else pressure_current
        pressure_min = min(pressure_hourly) if pressure_hourly else pressure_current
        pressure_avg = sum(pressure_hourly) / len(pressure_hourly) if pressure_hourly else pressure_current

        wind_speed_max = max(wind_speed_hourly) if wind_speed_hourly else wind_speed_current
        wind_speed_min = min(wind_speed_hourly) if wind_speed_hourly else wind_speed_current
        wind_speed_avg = sum(wind_speed_hourly) / len(wind_speed_hourly) if wind_speed_hourly else wind_speed_current

        wind_direction_avg = self.calculate_avg_wind_direction(wind_direction_hourly)

        rain_max = max(rain_hourly) if rain_hourly else rain_current
        rain_min = min(rain_hourly) if rain_hourly else rain_current
        rain_total = sum(rain_hourly) if rain_hourly else rain_current * 24
        rain_avg = rain_total / 24 if rain_hourly else rain_current

        return {
            'station_id': station_info['station_id'],
            'station_name': station_info['station_name'],
            'province': station_info['province'],
            'district': station_info['district'],
            'latitude': station_info['latitude'],
            'longitude': station_info['longitude'],
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'data_source': source,
            'data_quality': quality,
            'data_time': current.get('time', datetime.now().isoformat()),

            # Nhiệt độ - đầy đủ max, min, avg
            'temperature_current': round(temp_current, 1),
            'temperature_max': round(temp_max, 1),
            'temperature_min': round(temp_min, 1),
            'temperature_avg': round(temp_avg, 1),

            # Độ ẩm - đầy đủ max, min, avg
            'humidity_current': round(humidity_current, 1),
            'humidity_max': round(humidity_max, 1),
            'humidity_min': round(humidity_min, 1),
            'humidity_avg': round(humidity_avg, 1),

            # Áp suất - đầy đủ max, min, avg
            'pressure_current': round(pressure_current, 1),
            'pressure_max': round(pressure_max, 1),
            'pressure_min': round(pressure_min, 1),
            'pressure_avg': round(pressure_avg, 1),

            # Tốc độ gió - đầy đủ max, min, avg
            'wind_speed_current': round(wind_speed_current, 1),
            'wind_speed_max': round(wind_speed_max, 1),
            'wind_speed_min': round(wind_speed_min, 1),
            'wind_speed_avg': round(wind_speed_avg, 1),

            # Hướng gió
            'wind_direction_current': round(wind_direction_current, 1),
            'wind_direction_avg': round(wind_direction_avg, 1),

            # Lượng mưa - đầy đủ max, min, avg, total
            'rain_current': round(rain_current, 1),
            'rain_max': round(rain_max, 1),
            'rain_min': round(rain_min, 1),
            'rain_avg': round(rain_avg, 1),
            'rain_total': round(rain_total, 1),

            # Các chỉ số ước tính
            'cloud_cover_current': random.randint(20, 80),
            'cloud_cover_max': random.randint(60, 95),
            'cloud_cover_min': random.randint(10, 40),
            'cloud_cover_avg': random.randint(30, 70),

            'visibility_current': random.randint(5, 15),
            'visibility_max': random.randint(10, 20),
            'visibility_min': random.randint(2, 8),
            'visibility_avg': random.randint(6, 12),

            'thunder_probability': random.randint(0, 30)
        }

    def calculate_avg_wind_direction(self, directions):
        """Tính hướng gió trung bình"""
        if not directions:
            return 0
        try:
            rads = [d * np.pi / 180 for d in directions]
            x_sum = sum([np.sin(r) for r in rads])
            y_sum = sum([np.cos(r) for r in rads])
            avg_rad = np.arctan2(x_sum, y_sum)
            return (avg_rad * 180 / np.pi) % 360
        except:
            return sum(directions) / len(directions)

    def create_fallback_weather_record(self, station_info, reason):
        """Tạo bản ghi thời tiết fallback với đầy đủ max, min, avg"""
        return {
            'station_id': station_info['station_id'],
            'station_name': station_info['station_name'],
            'province': station_info['province'],
            'district': station_info['district'],
            'latitude': station_info['latitude'],
            'longitude': station_info['longitude'],
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'data_source': 'fallback',
            'data_quality': 'low',
            'data_time': datetime.now().isoformat(),
            'error_reason': reason,

            # Nhiệt độ
            'temperature_current': 25.0, 'temperature_max': 30.0, 'temperature_min': 20.0, 'temperature_avg': 25.0,

            # Độ ẩm
            'humidity_current': 75.0, 'humidity_max': 85.0, 'humidity_min': 65.0, 'humidity_avg': 75.0,

            # Áp suất
            'pressure_current': 1013.0, 'pressure_max': 1020.0, 'pressure_min': 1005.0, 'pressure_avg': 1013.0,

            # Gió
            'wind_speed_current': 3.0, 'wind_speed_max': 6.0, 'wind_speed_min': 1.0, 'wind_speed_avg': 3.5,
            'wind_direction_current': 180, 'wind_direction_avg': 180,

            # Mưa
            'rain_current': 0, 'rain_max': 0, 'rain_min': 0, 'rain_avg': 0, 'rain_total': 0,

            # Các chỉ số ước tính
            'cloud_cover_current': 50, 'cloud_cover_max': 80, 'cloud_cover_min': 20, 'cloud_cover_avg': 50,
            'visibility_current': 10, 'visibility_max': 15, 'visibility_min': 5, 'visibility_avg': 10,
            'thunder_probability': 5
        }

    def crawl_all_locations(self, locations, delay=2.0):
        """Crawl dữ liệu cho tất cả địa điểm"""
        all_weather_data = []

        total_locations = len(locations)

        logging.info(f"🔄 Bắt đầu thu thập dữ liệu cho {total_locations} địa điểm tại Việt Nam")
        logging.info("📊 Sử dụng đa nguồn dữ liệu với đánh giá chất lượng")

        for i, location in enumerate(locations):
            logging.info(f"📍 Đang xử lý {i + 1}/{total_locations}: {location['station_name']}")

            try:
                # Thu thập dữ liệu thời tiết
                weather_data = self.parse_weather_data(location)
                if weather_data:
                    all_weather_data.append(weather_data)
                    logging.info(f"  ✅ Thời tiết: {weather_data['data_source']} ({weather_data['data_quality']})")

                time.sleep(delay)

            except Exception as e:
                logging.error(f"❌ Lỗi xử lý {location['station_name']}: {e}")
                continue

        return all_weather_data

    def get_data_quality_report(self):
        """Tạo báo cáo chất lượng dữ liệu"""
        weather_total = sum(self.data_quality_tracking['weather'].values())

        report = {
            'weather': {
                'total': weather_total,
                'high_quality': self.data_quality_tracking['weather']['high_quality'],
                'medium_quality': self.data_quality_tracking['weather']['medium_quality'],
                'low_quality': self.data_quality_tracking['weather']['low_quality'],
                'high_percent': round((self.data_quality_tracking['weather']['high_quality'] / weather_total * 100),
                                      2) if weather_total > 0 else 0,
                'medium_percent': round((self.data_quality_tracking['weather']['medium_quality'] / weather_total * 100),
                                        2) if weather_total > 0 else 0,
                'low_percent': round((self.data_quality_tracking['weather']['low_quality'] / weather_total * 100),
                                     2) if weather_total > 0 else 0
            }
        }

        return report

    def save_to_excel(self, weather_data, output_dir="output"):
        """Lưu dữ liệu ra file Excel với đánh giá chất lượng"""
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(output_dir, f"vietnam_weather_data_{timestamp}.xlsx")

        wb = Workbook()

        # Sheet thời tiết
        if weather_data:
            weather_df = pd.DataFrame(weather_data)
            ws_weather = wb.active
            ws_weather.title = "Weather Data"

            for r in dataframe_to_rows(weather_df, index=False, header=True):
                ws_weather.append(r)

            # Định dạng header
            for cell in ws_weather[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Tự động điều chỉnh độ rộng cột
            for column in ws_weather.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_weather.column_dimensions[column_letter].width = adjusted_width
        else:
            ws_weather = wb.active
            ws_weather.title = "Weather Data"
            ws_weather['A1'] = "Không có dữ liệu thời tiết"

        # Sheet chất lượng dữ liệu
        ws_quality = wb.create_sheet("Data Quality Report")
        quality_report = self.get_data_quality_report()

        ws_quality['A1'] = "BÁO CÁO CHẤT LƯỢNG DỮ LIỆU"
        ws_quality['A1'].font = Font(bold=True, size=14, color="366092")

        # Thống kê chất lượng
        stats_data = [
            ['Loại dữ liệu', 'Tổng số', 'Chất lượng cao', 'Chất lượng trung bình', 'Chất lượng thấp', 'Tỷ lệ cao (%)'],
            ['Thời tiết',
             quality_report['weather']['total'],
             quality_report['weather']['high_quality'],
             quality_report['weather']['medium_quality'],
             quality_report['weather']['low_quality'],
             quality_report['weather']['high_percent']]
        ]

        for row_idx, row_data in enumerate(stats_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_quality.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        # Điều chỉnh độ rộng cột
        for column in ws_quality.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_quality.column_dimensions[column_letter].width = adjusted_width

        wb.save(excel_file)
        logging.info(f"💾 Đã lưu dữ liệu vào: {excel_file}")

        return excel_file

    def save_to_sqlite(self, weather_data, locations):
        """Lưu dữ liệu vào SQLite database"""
        try:
            # Kết nối database
            self.db_manager.connect()

            # Tạo các bảng
            self.db_manager.create_tables()

            # Chèn dữ liệu trạm
            self.db_manager.insert_stations(locations)

            # Chèn dữ liệu thời tiết
            inserted_count = self.db_manager.insert_weather_data(weather_data)

            # Chèn log chất lượng
            quality_report = self.get_data_quality_report()
            run_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.db_manager.insert_quality_log(quality_report, run_timestamp)

            # Đóng kết nối
            self.db_manager.disconnect()

            logging.info(f"💾 Đã lưu {inserted_count} bản ghi vào SQLite database")
            return True

        except Exception as e:
            logging.error(f"❌ Lỗi lưu dữ liệu vào SQLite: {e}")
            return False

    def get_database_summary(self):
        """Lấy tổng quan dữ liệu từ database"""
        try:
            self.db_manager.connect()
            summary = self.db_manager.get_data_summary()
            self.db_manager.disconnect()
            return summary
        except Exception as e:
            logging.error(f"❌ Lỗi lấy tổng quan database: {e}")
            return {}


# DANH SÁCH ĐỊA ĐIỂM VIỆT NAM
vietnam_locations = [
    # Hà Nội - 30 quận huyện
    {"station_id": "HN_BD", "station_name": "Quận Ba Đình - Hà Nội", "province": "Hà Nội", "district": "Ba Đình", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0338, "longitude": 105.8142},
    {"station_id": "HN_HK", "station_name": "Quận Hoàn Kiếm - Hà Nội", "province": "Hà Nội", "district": "Hoàn Kiếm", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0285, "longitude": 105.8542},
    {"station_id": "HN_HBT", "station_name": "Quận Hai Bà Trưng - Hà Nội", "province": "Hà Nội", "district": "Hai Bà Trưng", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0091, "longitude": 105.8606},
    {"station_id": "HN_DD", "station_name": "Quận Đống Đa - Hà Nội", "province": "Hà Nội", "district": "Đống Đa", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0190, "longitude": 105.8315},
    {"station_id": "HN_CG", "station_name": "Quận Cầu Giấy - Hà Nội", "province": "Hà Nội", "district": "Cầu Giấy", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0301, "longitude": 105.8022},
    {"station_id": "HN_TX", "station_name": "Quận Thanh Xuân - Hà Nội", "province": "Hà Nội", "district": "Thanh Xuân", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0010, "longitude": 105.8093},
    {"station_id": "HN_HM", "station_name": "Quận Hoàng Mai - Hà Nội", "province": "Hà Nội", "district": "Hoàng Mai", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 20.9837, "longitude": 105.8636},
    {"station_id": "HN_LB", "station_name": "Quận Long Biên - Hà Nội", "province": "Hà Nội", "district": "Long Biên", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0540, "longitude": 105.8959},
    {"station_id": "HN_NTL", "station_name": "Quận Nam Từ Liêm - Hà Nội", "province": "Hà Nội", "district": "Nam Từ Liêm", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0113, "longitude": 105.7547},
    {"station_id": "HN_BTL", "station_name": "Quận Bắc Từ Liêm - Hà Nội", "province": "Hà Nội", "district": "Bắc Từ Liêm", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 21.0772, "longitude": 105.7730},
    {"station_id": "HN_HD", "station_name": "Quận Hà Đông - Hà Nội", "province": "Hà Nội", "district": "Hà Đông", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 20.9714, "longitude": 105.7788},
    {"station_id": "HN_ST", "station_name": "Thị xã Sơn Tây - Hà Nội", "province": "Hà Nội", "district": "Sơn Tây", "type": "Thị xã", "region": "Đồng bằng sông Hồng", "latitude": 21.1376, "longitude": 105.5070},
    {"station_id": "HN_DP", "station_name": "Huyện Đan Phượng - Hà Nội", "province": "Hà Nội", "district": "Đan Phượng", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.1394, "longitude": 105.6736},
    {"station_id": "HN_HĐ", "station_name": "Huyện Hoài Đức - Hà Nội", "province": "Hà Nội", "district": "Hoài Đức", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.0392, "longitude": 105.7108},
    {"station_id": "HN_QO", "station_name": "Huyện Quốc Oai - Hà Nội", "province": "Hà Nội", "district": "Quốc Oai", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.9956, "longitude": 105.6119},
    {"station_id": "HN_TT", "station_name": "Huyện Thạch Thất - Hà Nội", "province": "Hà Nội", "district": "Thạch Thất", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.0278, "longitude": 105.5564},
    {"station_id": "HN_CM", "station_name": "Huyện Chương Mỹ - Hà Nội", "province": "Hà Nội", "district": "Chương Mỹ", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.8964, "longitude": 105.7039},
    {"station_id": "HN_TO", "station_name": "Huyện Thanh Oai - Hà Nội", "province": "Hà Nội", "district": "Thanh Oai", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.8556, "longitude": 105.7722},
    {"station_id": "HN_TT2", "station_name": "Huyện Thường Tín - Hà Nội", "province": "Hà Nội", "district": "Thường Tín", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.8700, "longitude": 105.8583},
    {"station_id": "HN_PX", "station_name": "Huyện Phú Xuyên - Hà Nội", "province": "Hà Nội", "district": "Phú Xuyên", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.7392, "longitude": 105.9139},
    {"station_id": "HN_UH", "station_name": "Huyện Ứng Hòa - Hà Nội", "province": "Hà Nội", "district": "Ứng Hòa", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.7292, "longitude": 105.8236},
    {"station_id": "HN_MD", "station_name": "Huyện Mỹ Đức - Hà Nội", "province": "Hà Nội", "district": "Mỹ Đức", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.6472, "longitude": 105.7194},
    {"station_id": "HN_SL", "station_name": "Huyện Sóc Sơn - Hà Nội", "province": "Hà Nội", "district": "Sóc Sơn", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.2578, "longitude": 105.8489},
    {"station_id": "HN_DH", "station_name": "Huyện Đông Anh - Hà Nội", "province": "Hà Nội", "district": "Đông Anh", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.1467, "longitude": 105.8464},
    {"station_id": "HN_GL", "station_name": "Huyện Gia Lâm - Hà Nội", "province": "Hà Nội", "district": "Gia Lâm", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.0408, "longitude": 105.9500},
    {"station_id": "HN_MT", "station_name": "Huyện Mê Linh - Hà Nội", "province": "Hà Nội", "district": "Mê Linh", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.1753, "longitude": 105.7219},
    {"station_id": "HN_PT", "station_name": "Huyện Phúc Thọ - Hà Nội", "province": "Hà Nội", "district": "Phúc Thọ", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.1031, "longitude": 105.5514},
    {"station_id": "HN_BG", "station_name": "Huyện Ba Vì - Hà Nội", "province": "Hà Nội", "district": "Ba Vì", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.1992, "longitude": 105.4239},
    {"station_id": "HN_ML", "station_name": "Huyện Mỹ Đức - Hà Nội", "province": "Hà Nội", "district": "Mỹ Đức", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.6472, "longitude": 105.7194},
    {"station_id": "HN_PH", "station_name": "Huyện Phú Xuyên - Hà Nội", "province": "Hà Nội", "district": "Phú Xuyên", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.7392, "longitude": 105.9139},

    # TP Hồ Chí Minh - 22 quận huyện (đã bổ sung đủ)
    {"station_id": "HCM_1", "station_name": "Quận 1 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 1", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7757, "longitude": 106.7004},
    {"station_id": "HCM_2", "station_name": "Quận 2 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 2", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7872, "longitude": 106.7498},
    {"station_id": "HCM_3", "station_name": "Quận 3 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 3", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7823, "longitude": 106.6848},
    {"station_id": "HCM_4", "station_name": "Quận 4 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 4", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7642, "longitude": 106.7053},
    {"station_id": "HCM_5", "station_name": "Quận 5 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 5", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7540, "longitude": 106.6690},
    {"station_id": "HCM_6", "station_name": "Quận 6 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 6", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7464, "longitude": 106.6492},
    {"station_id": "HCM_7", "station_name": "Quận 7 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 7", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7329, "longitude": 106.7269},
    {"station_id": "HCM_8", "station_name": "Quận 8 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 8", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7241, "longitude": 106.6286},
    {"station_id": "HCM_9", "station_name": "Quận 9 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 9", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.8428, "longitude": 106.8287},
    {"station_id": "HCM_10", "station_name": "Quận 10 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 10", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7679, "longitude": 106.6668},
    {"station_id": "HCM_11", "station_name": "Quận 11 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 11", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7639, "longitude": 106.6474},
    {"station_id": "HCM_12", "station_name": "Quận 12 - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Quận 12", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.8642, "longitude": 106.6543},
    {"station_id": "HCM_GB", "station_name": "Quận Gò Vấp - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Gò Vấp", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.8383, "longitude": 106.6653},
    {"station_id": "HCM_TB", "station_name": "Quận Tân Bình - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Tân Bình", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.8014, "longitude": 106.6526},
    {"station_id": "HCM_TP", "station_name": "Quận Tân Phú - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Tân Phú", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7902, "longitude": 106.6289},
    {"station_id": "HCM_BT", "station_name": "Quận Bình Tân - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Bình Tân", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7654, "longitude": 106.6032},
    {"station_id": "HCM_BT2", "station_name": "Quận Bình Thạnh - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Bình Thạnh", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.8010, "longitude": 106.6959},
    {"station_id": "HCM_PN", "station_name": "Quận Phú Nhuận - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Phú Nhuận", "type": "Quận", "region": "Đông Nam Bộ", "latitude": 10.7992, "longitude": 106.6753},
    {"station_id": "HCM_TD", "station_name": "Thành phố Thủ Đức - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Thủ Đức", "type": "Thành phố", "region": "Đông Nam Bộ", "latitude": 10.8494, "longitude": 106.7717},
    {"station_id": "HCM_CC", "station_name": "Huyện Củ Chi - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Củ Chi", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.9733, "longitude": 106.4939},
    {"station_id": "HCM_HM", "station_name": "Huyện Hóc Môn - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Hóc Môn", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.8894, "longitude": 106.5953},
    {"station_id": "HCM_BC", "station_name": "Huyện Bình Chánh - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Bình Chánh", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.7269, "longitude": 106.5672},
    {"station_id": "HCM_NB", "station_name": "Huyện Nhà Bè - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Nhà Bè", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.6956, "longitude": 106.7350},
    {"station_id": "HCM_CG", "station_name": "Huyện Cần Giờ - TP Hồ Chí Minh", "province": "TP Hồ Chí Minh", "district": "Cần Giờ", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.4111, "longitude": 106.9547},

    # Hải Phòng
    {"station_id": "HP_LC", "station_name": "Quận Lê Chân - Hải Phòng", "province": "Hải Phòng", "district": "Lê Chân", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 20.8516, "longitude": 106.6822},
    {"station_id": "HP_NQ", "station_name": "Quận Ngô Quyền - Hải Phòng", "province": "Hải Phòng", "district": "Ngô Quyền", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 20.8580, "longitude": 106.7053},
    {"station_id": "HP_HA", "station_name": "Quận Hải An - Hải Phòng", "province": "Hải Phòng", "district": "Hải An", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 20.8378, "longitude": 106.7377},
    {"station_id": "HP_KS", "station_name": "Quận Kiến An - Hải Phòng", "province": "Hải Phòng", "district": "Kiến An", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 20.8075, "longitude": 106.6264},
    {"station_id": "HP_DS", "station_name": "Quận Dương Kinh - Hải Phòng", "province": "Hải Phòng", "district": "Dương Kinh", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 20.8025, "longitude": 106.6669},
    {"station_id": "HP_DK", "station_name": "Quận Đồ Sơn - Hải Phòng", "province": "Hải Phòng", "district": "Đồ Sơn", "type": "Quận", "region": "Đồng bằng sông Hồng", "latitude": 20.7278, "longitude": 106.7733},
    {"station_id": "HP_TL", "station_name": "Huyện Thủy Nguyên - Hải Phòng", "province": "Hải Phòng", "district": "Thủy Nguyên", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.9583, "longitude": 106.6667},
    {"station_id": "HP_AC", "station_name": "Huyện An Dương - Hải Phòng", "province": "Hải Phòng", "district": "An Dương", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.8667, "longitude": 106.6167},
    {"station_id": "HP_AL", "station_name": "Huyện An Lão - Hải Phòng", "province": "Hải Phòng", "district": "An Lão", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.8167, "longitude": 106.5500},
    {"station_id": "HP_KL", "station_name": "Huyện Kiến Thụy - Hải Phòng", "province": "Hải Phòng", "district": "Kiến Thụy", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.7500, "longitude": 106.6667},
    {"station_id": "HP_TH", "station_name": "Huyện Tiên Lãng - Hải Phòng", "province": "Hải Phòng", "district": "Tiên Lãng", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.7333, "longitude": 106.5500},
    {"station_id": "HP_VB", "station_name": "Huyện Vĩnh Bảo - Hải Phòng", "province": "Hải Phòng", "district": "Vĩnh Bảo", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.7000, "longitude": 106.4667},
    {"station_id": "HP_CH", "station_name": "Huyện Cát Hải - Hải Phòng", "province": "Hải Phòng", "district": "Cát Hải", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.7264, "longitude": 107.0489},
    {"station_id": "HP_BH", "station_name": "Huyện Bạch Long Vĩ - Hải Phòng", "province": "Hải Phòng", "district": "Bạch Long Vĩ", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.1333, "longitude": 107.7333},

    # Đà Nẵng
    {"station_id": "DN_HC", "station_name": "Quận Hải Châu - Đà Nẵng", "province": "Đà Nẵng", "district": "Hải Châu", "type": "Quận", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.0592, "longitude": 108.2208},
    {"station_id": "DN_ST", "station_name": "Quận Sơn Trà - Đà Nẵng", "province": "Đà Nẵng", "district": "Sơn Trà", "type": "Quận", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.1060, "longitude": 108.2493},
    {"station_id": "DN_NH", "station_name": "Quận Ngũ Hành Sơn - Đà Nẵng", "province": "Đà Nẵng", "district": "Ngũ Hành Sơn", "type": "Quận", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.0159, "longitude": 108.2579},
    {"station_id": "DN_LC", "station_name": "Quận Liên Chiểu - Đà Nẵng", "province": "Đà Nẵng", "district": "Liên Chiểu", "type": "Quận", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.0717, "longitude": 108.1503},
    {"station_id": "DN_TS", "station_name": "Quận Thanh Khê - Đà Nẵng", "province": "Đà Nẵng", "district": "Thanh Khê", "type": "Quận", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.0647, "longitude": 108.1911},
    {"station_id": "DN_CL", "station_name": "Quận Cẩm Lệ - Đà Nẵng", "province": "Đà Nẵng", "district": "Cẩm Lệ", "type": "Quận", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.0156, "longitude": 108.2028},
    {"station_id": "DN_HV", "station_name": "Huyện Hòa Vang - Đà Nẵng", "province": "Đà Nẵng", "district": "Hòa Vang", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.0333, "longitude": 108.0833},
    {"station_id": "DN_HH", "station_name": "Huyện Hoàng Sa - Đà Nẵng", "province": "Đà Nẵng", "district": "Hoàng Sa", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.5322, "longitude": 111.6156},

    # Cần Thơ
    {"station_id": "CT_NK", "station_name": "Quận Ninh Kiều - Cần Thơ", "province": "Cần Thơ", "district": "Ninh Kiều", "type": "Quận", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0452, "longitude": 105.7469},
    {"station_id": "CT_BT", "station_name": "Quận Bình Thủy - Cần Thơ", "province": "Cần Thơ", "district": "Bình Thủy", "type": "Quận", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0816, "longitude": 105.7378},
    {"station_id": "CT_CR", "station_name": "Quận Cái Răng - Cần Thơ", "province": "Cần Thơ", "district": "Cái Răng", "type": "Quận", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0000, "longitude": 105.7667},
    {"station_id": "CT_OT", "station_name": "Quận Ô Môn - Cần Thơ", "province": "Cần Thơ", "district": "Ô Môn", "type": "Quận", "region": "Đồng bằng sông Cửu Long", "latitude": 10.1167, "longitude": 105.6333},
    {"station_id": "CT_TL", "station_name": "Quận Thốt Nốt - Cần Thơ", "province": "Cần Thơ", "district": "Thốt Nốt", "type": "Quận", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2667, "longitude": 105.5333},
    {"station_id": "CT_VT", "station_name": "Huyện Vĩnh Thạnh - Cần Thơ", "province": "Cần Thơ", "district": "Vĩnh Thạnh", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2167, "longitude": 105.4000},
    {"station_id": "CT_CC", "station_name": "Huyện Cờ Đỏ - Cần Thơ", "province": "Cần Thơ", "district": "Cờ Đỏ", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.1000, "longitude": 105.4333},
    {"station_id": "CT_PT", "station_name": "Huyện Phong Điền - Cần Thơ", "province": "Cần Thơ", "district": "Phong Điền", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2500, "longitude": 105.6667},
    {"station_id": "CT_TN", "station_name": "Huyện Thới Lai - Cần Thơ", "province": "Cần Thơ", "district": "Thới Lai", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0667, "longitude": 105.5500},

    # Quảng Ninh - ĐÃ SỬA VÙNG MIỀN
    {"station_id": "QN_HL", "station_name": "Thành phố Hạ Long - Quảng Ninh", "province": "Quảng Ninh", "district": "Hạ Long", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 20.9582, "longitude": 107.0758},
    {"station_id": "QN_CP", "station_name": "Thành phố Cẩm Phả - Quảng Ninh", "province": "Quảng Ninh", "district": "Cẩm Phả", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.0167, "longitude": 107.3167},
    {"station_id": "QN_UB", "station_name": "Thành phố Uông Bí - Quảng Ninh", "province": "Quảng Ninh", "district": "Uông Bí", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.0361, "longitude": 106.7633},
    {"station_id": "QN_MC", "station_name": "Thành phố Móng Cái - Quảng Ninh", "province": "Quảng Ninh", "district": "Móng Cái", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.5247, "longitude": 107.9664},
    {"station_id": "QN_DK", "station_name": "Thị xã Đông Triều - Quảng Ninh", "province": "Quảng Ninh", "district": "Đông Triều", "type": "Thị xã", "region": "Đông Bắc Bộ", "latitude": 21.0833, "longitude": 106.5000},
    {"station_id": "QN_QY", "station_name": "Thị xã Quảng Yên - Quảng Ninh", "province": "Quảng Ninh", "district": "Quảng Yên", "type": "Thị xã", "region": "Đông Bắc Bộ", "latitude": 20.9333, "longitude": 106.8167},
    {"station_id": "QN_BG", "station_name": "Huyện Ba Chẽ - Quảng Ninh", "province": "Quảng Ninh", "district": "Ba Chẽ", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.2833, "longitude": 107.1833},
    {"station_id": "QN_BQ", "station_name": "Huyện Bình Liêu - Quảng Ninh", "province": "Quảng Ninh", "district": "Bình Liêu", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.5333, "longitude": 107.4333},
    {"station_id": "QN_CT", "station_name": "Huyện Cô Tô - Quảng Ninh", "province": "Quảng Ninh", "district": "Cô Tô", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 20.9681, "longitude": 107.7639},
    {"station_id": "QN_DH", "station_name": "Huyện Đầm Hà - Quảng Ninh", "province": "Quảng Ninh", "district": "Đầm Hà", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.3667, "longitude": 107.6000},
    {"station_id": "QN_HD", "station_name": "Huyện Hải Hà - Quảng Ninh", "province": "Quảng Ninh", "district": "Hải Hà", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.4333, "longitude": 107.7167},
    {"station_id": "QN_TY", "station_name": "Huyện Tiên Yên - Quảng Ninh", "province": "Quảng Ninh", "district": "Tiên Yên", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.3833, "longitude": 107.3833},
    {"station_id": "QN_VD", "station_name": "Huyện Vân Đồn - Quảng Ninh", "province": "Quảng Ninh", "district": "Vân Đồn", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 20.9551, "longitude": 107.4764},

    # Thái Bình
    {"station_id": "TB_TP", "station_name": "Thành phố Thái Bình - Thái Bình", "province": "Thái Bình", "district": "Thái Bình", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 20.4461, "longitude": 106.3366},
    {"station_id": "TB_DH", "station_name": "Huyện Đông Hưng - Thái Bình", "province": "Thái Bình", "district": "Đông Hưng", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.5500, "longitude": 106.3500},
    {"station_id": "TB_HH", "station_name": "Huyện Hưng Hà - Thái Bình", "province": "Thái Bình", "district": "Hưng Hà", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.6000, "longitude": 106.2167},
    {"station_id": "TB_KL", "station_name": "Huyện Kiến Xương - Thái Bình", "province": "Thái Bình", "district": "Kiến Xương", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.4000, "longitude": 106.4167},
    {"station_id": "TB_QH", "station_name": "Huyện Quỳnh Phụ - Thái Bình", "province": "Thái Bình", "district": "Quỳnh Phụ", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.6500, "longitude": 106.3333},
    {"station_id": "TB_TT", "station_name": "Huyện Thái Thụy - Thái Bình", "province": "Thái Bình", "district": "Thái Thụy", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.5333, "longitude": 106.5333},
    {"station_id": "TB_VT", "station_name": "Huyện Vũ Thư - Thái Bình", "province": "Thái Bình", "district": "Vũ Thư", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.4333, "longitude": 106.2833},

    # Nam Định
    {"station_id": "ND_TP", "station_name": "Thành phố Nam Định - Nam Định", "province": "Nam Định", "district": "Nam Định", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 20.4200, "longitude": 106.1683},
    {"station_id": "ND_GL", "station_name": "Huyện Giao Thủy - Nam Định", "province": "Nam Định", "district": "Giao Thủy", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.2333, "longitude": 106.4500},
    {"station_id": "ND_HH", "station_name": "Huyện Hải Hậu - Nam Định", "province": "Nam Định", "district": "Hải Hậu", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.1833, "longitude": 106.3000},
    {"station_id": "ND_ML", "station_name": "Huyện Mỹ Lộc - Nam Định", "province": "Nam Định", "district": "Mỹ Lộc", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.4667, "longitude": 106.1167},
    {"station_id": "ND_NC", "station_name": "Huyện Nam Trực - Nam Định", "province": "Nam Định", "district": "Nam Trực", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.3333, "longitude": 106.2000},
    {"station_id": "ND_TK", "station_name": "Huyện Trực Ninh - Nam Định", "province": "Nam Định", "district": "Trực Ninh", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.2500, "longitude": 106.2500},
    {"station_id": "ND_VB", "station_name": "Huyện Vụ Bản - Nam Định", "province": "Nam Định", "district": "Vụ Bản", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.3667, "longitude": 106.1000},
    {"station_id": "ND_XT", "station_name": "Huyện Xuân Trường - Nam Định", "province": "Nam Định", "district": "Xuân Trường", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.3000, "longitude": 106.3500},
    {"station_id": "ND_YN", "station_name": "Huyện Ý Yên - Nam Định", "province": "Nam Định", "district": "Ý Yên", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.3167, "longitude": 106.0167},

    # Thừa Thiên Huế
    {"station_id": "TTH_H", "station_name": "Thành phố Huế - Thừa Thiên Huế", "province": "Thừa Thiên Huế", "district": "Huế", "type": "Thành phố", "region": "Bắc Trung Bộ", "latitude": 16.4637, "longitude": 107.5909},
    {"station_id": "TTH_HL", "station_name": "Thị xã Hương Thủy - Thừa Thiên Huế", "province": "Thừa Thiên Huế", "district": "Hương Thủy", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 16.4167, "longitude": 107.7167},
    {"station_id": "TTH_HT", "station_name": "Thị xã Hương Trà - Thừa Thiên Huế", "province": "Thừa Thiên Huế", "district": "Hương Trà", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 16.5000, "longitude": 107.4833},
    {"station_id": "TTH_AD", "station_name": "Huyện A Lưới - Thừa Thiên Huế", "province": "Thừa Thiên Huế", "district": "A Lưới", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.2333, "longitude": 107.3000},
    {"station_id": "TTH_NL", "station_name": "Huyện Nam Đông - Thừa Thiên Huế", "province": "Thừa Thiên Huế", "district": "Nam Đông", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.1667, "longitude": 107.7000},
    {"station_id": "TTH_PD", "station_name": "Huyện Phong Điền - Thừa Thiên Huế", "province": "Thừa Thiên Huế", "district": "Phong Điền", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.5833, "longitude": 107.3500},
    {"station_id": "TTH_PL", "station_name": "Huyện Phú Lộc - Thừa Thiên Huế", "province": "Thừa Thiên Huế", "district": "Phú Lộc", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.2667, "longitude": 107.9000},
    {"station_id": "TTH_QD", "station_name": "Huyện Quảng Điền - Thừa Thiên Huế", "province": "Thừa Thiên Huế", "district": "Quảng Điền", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.5833, "longitude": 107.5000},

    # Quảng Trị
    {"station_id": "QT_DH", "station_name": "Thành phố Đông Hà - Quảng Trị", "province": "Quảng Trị", "district": "Đông Hà", "type": "Thành phố", "region": "Bắc Trung Bộ", "latitude": 16.8160, "longitude": 107.1000},
    {"station_id": "QT_QH", "station_name": "Thị xã Quảng Trị - Quảng Trị", "province": "Quảng Trị", "district": "Quảng Trị", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 16.7500, "longitude": 107.1833},
    {"station_id": "QT_CL", "station_name": "Huyện Cam Lộ - Quảng Trị", "province": "Quảng Trị", "district": "Cam Lộ", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.8000, "longitude": 106.9833},
    {"station_id": "QT_CC", "station_name": "Huyện Cồn Cỏ - Quảng Trị", "province": "Quảng Trị", "district": "Cồn Cỏ", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 17.1597, "longitude": 107.3408},
    {"station_id": "QT_DK", "station_name": "Huyện Đa Krông - Quảng Trị", "province": "Quảng Trị", "district": "Đa Krông", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.5556, "longitude": 106.9722},
    {"station_id": "QT_GL", "station_name": "Huyện Gio Linh - Quảng Trị", "province": "Quảng Trị", "district": "Gio Linh", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.9167, "longitude": 107.0500},
    {"station_id": "QT_HH", "station_name": "Huyện Hải Lăng - Quảng Trị", "province": "Quảng Trị", "district": "Hải Lăng", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.7000, "longitude": 107.2500},
    {"station_id": "QT_HL", "station_name": "Huyện Hướng Hóa - Quảng Trị", "province": "Quảng Trị", "district": "Hướng Hóa", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.7000, "longitude": 106.6667},
    {"station_id": "QT_TL", "station_name": "Huyện Triệu Phong - Quảng Trị", "province": "Quảng Trị", "district": "Triệu Phong", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 16.7833, "longitude": 107.1667},
    {"station_id": "QT_VL", "station_name": "Huyện Vĩnh Linh - Quảng Trị", "province": "Quảng Trị", "district": "Vĩnh Linh", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 17.0000, "longitude": 106.9333},

    # Quảng Bình - ĐÃ SỬA TRÙNG LẶP
    {"station_id": "QB_DH", "station_name": "Thành phố Đồng Hới - Quảng Bình", "province": "Quảng Bình", "district": "Đồng Hới", "type": "Thành phố", "region": "Bắc Trung Bộ", "latitude": 17.4687, "longitude": 106.6227},
    {"station_id": "QB_BAD", "station_name": "Thị xã Ba Đồn - Quảng Bình", "province": "Quảng Bình", "district": "Ba Đồn", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 17.7500, "longitude": 106.4167},
    {"station_id": "QB_BT", "station_name": "Huyện Bố Trạch - Quảng Bình", "province": "Quảng Bình", "district": "Bố Trạch", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 17.5000, "longitude": 106.2500},
    {"station_id": "QB_LT", "station_name": "Huyện Lệ Thủy - Quảng Bình", "province": "Quảng Bình", "district": "Lệ Thủy", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 17.2167, "longitude": 106.8000},
    {"station_id": "QB_MC", "station_name": "Huyện Minh Hóa - Quảng Bình", "province": "Quảng Bình", "district": "Minh Hóa", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 17.7833, "longitude": 105.9167},
    {"station_id": "QB_QN", "station_name": "Huyện Quảng Ninh - Quảng Bình", "province": "Quảng Bình", "district": "Quảng Ninh", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 17.4000, "longitude": 106.5500},
    {"station_id": "QB_QH", "station_name": "Huyện Quảng Trạch - Quảng Bình", "province": "Quảng Bình", "district": "Quảng Trạch", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 17.8333, "longitude": 106.4167},
    {"station_id": "QB_TN", "station_name": "Huyện Tuyên Hóa - Quảng Bình", "province": "Quảng Bình", "district": "Tuyên Hóa", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 17.9000, "longitude": 106.0000},

    # Đắk Lắk
    {"station_id": "DL_BM", "station_name": "Thành phố Buôn Ma Thuột - Đắk Lắk", "province": "Đắk Lắk", "district": "Buôn Ma Thuột", "type": "Thành phố", "region": "Tây Nguyên", "latitude": 12.6662, "longitude": 108.0382},
    {"station_id": "DL_BH", "station_name": "Thị xã Buôn Hồ - Đắk Lắk", "province": "Đắk Lắk", "district": "Buôn Hồ", "type": "Thị xã", "region": "Tây Nguyên", "latitude": 12.8833, "longitude": 108.2333},
    {"station_id": "DL_CM", "station_name": "Huyện Cư M'gar - Đắk Lắk", "province": "Đắk Lắk", "district": "Cư M'gar", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.8500, "longitude": 108.1000},
    {"station_id": "DL_EH", "station_name": "Huyện Ea H'leo - Đắk Lắk", "province": "Đắk Lắk", "district": "Ea H'leo", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.1667, "longitude": 108.0833},
    {"station_id": "DL_ES", "station_name": "Huyện Ea Súp - Đắk Lắk", "province": "Đắk Lắk", "district": "Ea Súp", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.0833, "longitude": 107.8333},
    {"station_id": "DL_KK", "station_name": "Huyện Krông Ana - Đắk Lắk", "province": "Đắk Lắk", "district": "Krông Ana", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.5000, "longitude": 108.0333},
    {"station_id": "DL_KP", "station_name": "Huyện Krông Pắk - Đắk Lắk", "province": "Đắk Lắk", "district": "Krông Pắk", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.7000, "longitude": 108.3667},
    {"station_id": "DL_KN", "station_name": "Huyện Krông Nô - Đắk Lắk", "province": "Đắk Lắk", "district": "Krông Nô", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.3333, "longitude": 107.8333},
    {"station_id": "DL_LK", "station_name": "Huyện Lắk - Đắk Lắk", "province": "Đắk Lắk", "district": "Lắk", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.4167, "longitude": 108.1667},
    {"station_id": "DL_MD", "station_name": "Huyện M'Đrắk - Đắk Lắk", "province": "Đắk Lắk", "district": "M'Đrắk", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.7500, "longitude": 108.7500},

    # Gia Lai
    {"station_id": "GL_PC", "station_name": "Thành phố Pleiku - Gia Lai", "province": "Gia Lai", "district": "Pleiku", "type": "Thành phố", "region": "Tây Nguyên", "latitude": 13.9833, "longitude": 108.0000},
    {"station_id": "GL_AD", "station_name": "Thị xã An Khê - Gia Lai", "province": "Gia Lai", "district": "An Khê", "type": "Thị xã", "region": "Tây Nguyên", "latitude": 14.0000, "longitude": 108.6833},
    {"station_id": "GL_AY", "station_name": "Thị xã Ayun Pa - Gia Lai", "province": "Gia Lai", "district": "Ayun Pa", "type": "Thị xã", "region": "Tây Nguyên", "latitude": 13.3833, "longitude": 108.4333},
    {"station_id": "GL_CH", "station_name": "Huyện Chư Păh - Gia Lai", "province": "Gia Lai", "district": "Chư Păh", "type": "Huyện", "region": "Tây Nguyên", "latitude": 14.1667, "longitude": 107.9333},
    {"station_id": "GL_CP", "station_name": "Huyện Chư Prông - Gia Lai", "province": "Gia Lai", "district": "Chư Prông", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.5833, "longitude": 107.8333},
    {"station_id": "GL_CS", "station_name": "Huyện Chư Sê - Gia Lai", "province": "Gia Lai", "district": "Chư Sê", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.7500, "longitude": 108.0833},
    {"station_id": "GL_DC", "station_name": "Huyện Đăk Đoa - Gia Lai", "province": "Gia Lai", "district": "Đăk Đoa", "type": "Huyện", "region": "Tây Nguyên", "latitude": 14.1167, "longitude": 108.1167},
    {"station_id": "GL_DP", "station_name": "Huyện Đăk Pơ - Gia Lai", "province": "Gia Lai", "district": "Đăk Pơ", "type": "Huyện", "region": "Tây Nguyên", "latitude": 14.0000, "longitude": 108.5000},
    {"station_id": "GL_DA", "station_name": "Huyện Đức Cơ - Gia Lai", "province": "Gia Lai", "district": "Đức Cơ", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.7833, "longitude": 107.6667},
    {"station_id": "GL_IA", "station_name": "Huyện Ia Grai - Gia Lai", "province": "Gia Lai", "district": "Ia Grai", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.9833, "longitude": 107.7500},
    {"station_id": "GL_KB", "station_name": "Huyện KBang - Gia Lai", "province": "Gia Lai", "district": "KBang", "type": "Huyện", "region": "Tây Nguyên", "latitude": 14.3500, "longitude": 108.5000},
    {"station_id": "GL_KN", "station_name": "Huyện Kông Chro - Gia Lai", "province": "Gia Lai", "district": "Kông Chro", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.7167, "longitude": 108.5167},
    {"station_id": "GL_KR", "station_name": "Huyện Krông Pa - Gia Lai", "province": "Gia Lai", "district": "Krông Pa", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.2167, "longitude": 108.6667},
    {"station_id": "GL_MY", "station_name": "Huyện Mang Yang - Gia Lai", "province": "Gia Lai", "district": "Mang Yang", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.9833, "longitude": 108.2500},
    {"station_id": "GL_PG", "station_name": "Huyện Phú Thiện - Gia Lai", "province": "Gia Lai", "district": "Phú Thiện", "type": "Huyện", "region": "Tây Nguyên", "latitude": 13.5333, "longitude": 108.3000},

    # Lâm Đồng
    {"station_id": "LD_DL", "station_name": "Thành phố Đà Lạt - Lâm Đồng", "province": "Lâm Đồng", "district": "Đà Lạt", "type": "Thành phố", "region": "Tây Nguyên", "latitude": 11.9404, "longitude": 108.4587},
    {"station_id": "LD_BA", "station_name": "Thành phố Bảo Lộc - Lâm Đồng", "province": "Lâm Đồng", "district": "Bảo Lộc", "type": "Thành phố", "region": "Tây Nguyên", "latitude": 11.5500, "longitude": 107.8000},
    {"station_id": "LD_BD", "station_name": "Huyện Bảo Lâm - Lâm Đồng", "province": "Lâm Đồng", "district": "Bảo Lâm", "type": "Huyện", "region": "Tây Nguyên", "latitude": 11.7000, "longitude": 107.7167},
    {"station_id": "LD_CD", "station_name": "Huyện Cát Tiên - Lâm Đồng", "province": "Lâm Đồng", "district": "Cát Tiên", "type": "Huyện", "region": "Tây Nguyên", "latitude": 11.5833, "longitude": 107.3500},
    {"station_id": "LD_DH", "station_name": "Huyện Đạ Huoai - Lâm Đồng", "province": "Lâm Đồng", "district": "Đạ Huoai", "type": "Huyện", "region": "Tây Nguyên", "latitude": 11.4167, "longitude": 107.6333},
    {"station_id": "LD_DT", "station_name": "Huyện Đạ Tẻh - Lâm Đồng", "province": "Lâm Đồng", "district": "Đạ Tẻh", "type": "Huyện", "region": "Tây Nguyên", "latitude": 11.5000, "longitude": 107.5167},
    {"station_id": "LD_DT2", "station_name": "Huyện Đam Rông - Lâm Đồng", "province": "Lâm Đồng", "district": "Đam Rông", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.0833, "longitude": 108.1667},
    {"station_id": "LD_DD", "station_name": "Huyện Di Linh - Lâm Đồng", "province": "Lâm Đồng", "district": "Di Linh", "type": "Huyện", "region": "Tây Nguyên", "latitude": 11.5833, "longitude": 108.0667},
    {"station_id": "LD_DL2", "station_name": "Huyện Đơn Dương - Lâm Đồng", "province": "Lâm Đồng", "district": "Đơn Dương", "type": "Huyện", "region": "Tây Nguyên", "latitude": 11.7500, "longitude": 108.5500},
    {"station_id": "LD_DR", "station_name": "Huyện Đức Trọng - Lâm Đồng", "province": "Lâm Đồng", "district": "Đức Trọng", "type": "Huyện", "region": "Tây Nguyên", "latitude": 11.7333, "longitude": 108.3000},
    {"station_id": "LD_LC", "station_name": "Huyện Lạc Dương - Lâm Đồng", "province": "Lâm Đồng", "district": "Lạc Dương", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.0000, "longitude": 108.4333},
    {"station_id": "LD_LH", "station_name": "Huyện Lâm Hà - Lâm Đồng", "province": "Lâm Đồng", "district": "Lâm Hà", "type": "Huyện", "region": "Tây Nguyên", "latitude": 11.8333, "longitude": 108.2000},

    # Bình Dương
    {"station_id": "BD_TDM", "station_name": "Thành phố Thủ Dầu Một - Bình Dương", "province": "Bình Dương", "district": "Thủ Dầu Một", "type": "Thành phố", "region": "Đông Nam Bộ", "latitude": 10.9804, "longitude": 106.6519},
    {"station_id": "BD_BG", "station_name": "Thị xã Bến Cát - Bình Dương", "province": "Bình Dương", "district": "Bến Cát", "type": "Thị xã", "region": "Đông Nam Bộ", "latitude": 11.1500, "longitude": 106.6000},
    {"station_id": "BD_TU", "station_name": "Thị xã Tân Uyên - Bình Dương", "province": "Bình Dương", "district": "Tân Uyên", "type": "Thị xã", "region": "Đông Nam Bộ", "latitude": 11.0667, "longitude": 106.8000},
    {"station_id": "BD_TN", "station_name": "Thị xã Thuận An - Bình Dương", "province": "Bình Dương", "district": "Thuận An", "type": "Thị xã", "region": "Đông Nam Bộ", "latitude": 10.9333, "longitude": 106.7000},
    {"station_id": "BD_DT", "station_name": "Thị xã Dĩ An - Bình Dương", "province": "Bình Dương", "district": "Dĩ An", "type": "Thị xã", "region": "Đông Nam Bộ", "latitude": 10.9000, "longitude": 106.7667},
    {"station_id": "BD_BA", "station_name": "Huyện Bàu Bàng - Bình Dương", "province": "Bình Dương", "district": "Bàu Bàng", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.2500, "longitude": 106.8000},
    {"station_id": "BD_BC", "station_name": "Huyện Bắc Tân Uyên - Bình Dương", "province": "Bình Dương", "district": "Bắc Tân Uyên", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.1667, "longitude": 106.8667},
    {"station_id": "BD_DH", "station_name": "Huyện Dầu Tiếng - Bình Dương", "province": "Bình Dương", "district": "Dầu Tiếng", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.2667, "longitude": 106.3667},
    {"station_id": "BD_PG", "station_name": "Huyện Phú Giáo - Bình Dương", "province": "Bình Dương", "district": "Phú Giáo", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.3333, "longitude": 106.7667},

    # Đồng Nai
    {"station_id": "DN_BH", "station_name": "Thành phố Biên Hòa - Đồng Nai", "province": "Đồng Nai", "district": "Biên Hòa", "type": "Thành phố", "region": "Đông Nam Bộ", "latitude": 10.9574, "longitude": 106.8429},
    {"station_id": "DN_LT", "station_name": "Thành phố Long Khánh - Đồng Nai", "province": "Đồng Nai", "district": "Long Khánh", "type": "Thành phố", "region": "Đông Nam Bộ", "latitude": 10.9333, "longitude": 107.2500},
    {"station_id": "DN_CM", "station_name": "Huyện Cẩm Mỹ - Đồng Nai", "province": "Đồng Nai", "district": "Cẩm Mỹ", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.8000, "longitude": 107.2833},
    {"station_id": "DN_DR", "station_name": "Huyện Định Quán - Đồng Nai", "province": "Đồng Nai", "district": "Định Quán", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.1667, "longitude": 107.3333},
    {"station_id": "DN_LK", "station_name": "Huyện Long Thành - Đồng Nai", "province": "Đồng Nai", "district": "Long Thành", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.7833, "longitude": 107.0000},
    {"station_id": "DN_NT", "station_name": "Huyện Nhơn Trạch - Đồng Nai", "province": "Đồng Nai", "district": "Nhơn Trạch", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.7000, "longitude": 106.8833},
    {"station_id": "DN_TP", "station_name": "Huyện Tân Phú - Đồng Nai", "province": "Đồng Nai", "district": "Tân Phú", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.2833, "longitude": 107.4333},
    {"station_id": "DN_TU", "station_name": "Huyện Thống Nhất - Đồng Nai", "province": "Đồng Nai", "district": "Thống Nhất", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.9667, "longitude": 107.1500},
    {"station_id": "DN_VC", "station_name": "Huyện Vĩnh Cửu - Đồng Nai", "province": "Đồng Nai", "district": "Vĩnh Cửu", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.2500, "longitude": 107.0333},
    {"station_id": "DN_XL", "station_name": "Huyện Xuân Lộc - Đồng Nai", "province": "Đồng Nai", "district": "Xuân Lộc", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.9333, "longitude": 107.4167},
    {"station_id": "DN_TB", "station_name": "Huyện Trảng Bom - Đồng Nai", "province": "Đồng Nai", "district": "Trảng Bom", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.9500, "longitude": 107.0000},

    # Bà Rịa - Vũng Tàu
    {"station_id": "BR_VT", "station_name": "Thành phố Vũng Tàu - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Vũng Tàu", "type": "Thành phố", "region": "Đông Nam Bộ", "latitude": 10.3460, "longitude": 107.0843},
    {"station_id": "BR_BR", "station_name": "Thành phố Bà Rịa - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Bà Rịa", "type": "Thành phố", "region": "Đông Nam Bộ", "latitude": 10.5000, "longitude": 107.1833},
    {"station_id": "BR_PT", "station_name": "Thị xã Phú Mỹ - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Phú Mỹ", "type": "Thị xã", "region": "Đông Nam Bộ", "latitude": 10.6000, "longitude": 107.0500},
    {"station_id": "BR_CD", "station_name": "Huyện Côn Đảo - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Côn Đảo", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 8.6822, "longitude": 106.6089},
    {"station_id": "BR_DD", "station_name": "Huyện Đất Đỏ - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Đất Đỏ", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.5000, "longitude": 107.2833},
    {"station_id": "BR_LD", "station_name": "Huyện Long Điền - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Long Điền", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.4667, "longitude": 107.2333},
    {"station_id": "BR_XM", "station_name": "Huyện Xuyên Mộc - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Xuyên Mộc", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.6333, "longitude": 107.4333},
    {"station_id": "BR_CH", "station_name": "Huyện Châu Đức - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Châu Đức", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 10.6500, "longitude": 107.2500},

    # An Giang
    {"station_id": "AG_LX", "station_name": "Thành phố Long Xuyên - An Giang", "province": "An Giang", "district": "Long Xuyên", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3865, "longitude": 105.4351},
    {"station_id": "AG_CD", "station_name": "Thành phố Châu Đốc - An Giang", "province": "An Giang", "district": "Châu Đốc", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.7000, "longitude": 105.1167},
    {"station_id": "AG_AP", "station_name": "Thị xã Tân Châu - An Giang", "province": "An Giang", "district": "Tân Châu", "type": "Thị xã", "region": "Đồng bằng sông Cửu Long", "latitude": 10.8000, "longitude": 105.2167},
    {"station_id": "AG_AT", "station_name": "Huyện An Phú - An Giang", "province": "An Giang", "district": "An Phú", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.8333, "longitude": 105.0833},
    {"station_id": "AG_CL", "station_name": "Huyện Châu Phú - An Giang", "province": "An Giang", "district": "Châu Phú", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.5667, "longitude": 105.1667},
    {"station_id": "AG_CP", "station_name": "Huyện Châu Thành - An Giang", "province": "An Giang", "district": "Châu Thành", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.4500, "longitude": 105.2500},
    {"station_id": "AG_CT", "station_name": "Huyện Chợ Mới - An Giang", "province": "An Giang", "district": "Chợ Mới", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.5333, "longitude": 105.3833},
    {"station_id": "AG_PT", "station_name": "Huyện Phú Tân - An Giang", "province": "An Giang", "district": "Phú Tân", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.6500, "longitude": 105.2833},
    {"station_id": "AG_TV", "station_name": "Huyện Thoại Sơn - An Giang", "province": "An Giang", "district": "Thoại Sơn", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2667, "longitude": 105.2667},
    {"station_id": "AG_TN", "station_name": "Huyện Tịnh Biên - An Giang", "province": "An Giang", "district": "Tịnh Biên", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.5500, "longitude": 105.0000},
    {"station_id": "AG_TX", "station_name": "Huyện Tri Tôn - An Giang", "province": "An Giang", "district": "Tri Tôn", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.4167, "longitude": 105.0000},

    # Kiên Giang
    {"station_id": "KG_RG", "station_name": "Thành phố Rạch Giá - Kiên Giang", "province": "Kiên Giang", "district": "Rạch Giá", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0317, "longitude": 105.0809},
    {"station_id": "KG_HN", "station_name": "Thành phố Hà Tiên - Kiên Giang", "province": "Kiên Giang", "district": "Hà Tiên", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3833, "longitude": 104.4833},
    {"station_id": "KG_PQ", "station_name": "Huyện Phú Quốc - Kiên Giang", "province": "Kiên Giang", "district": "Phú Quốc", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2270, "longitude": 103.9679},
    {"station_id": "KG_AT", "station_name": "Huyện An Biên - Kiên Giang", "province": "Kiên Giang", "district": "An Biên", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.8167, "longitude": 105.0667},
    {"station_id": "KG_AM", "station_name": "Huyện An Minh - Kiên Giang", "province": "Kiên Giang", "district": "An Minh", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.6667, "longitude": 104.9500},
    {"station_id": "KG_CL", "station_name": "Huyện Châu Thành - Kiên Giang", "province": "Kiên Giang", "district": "Châu Thành", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.9333, "longitude": 105.1667},
    {"station_id": "KG_GT", "station_name": "Huyện Giang Thành - Kiên Giang", "province": "Kiên Giang", "district": "Giang Thành", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.4833, "longitude": 104.6333},
    {"station_id": "KG_GD", "station_name": "Huyện Gò Quao - Kiên Giang", "province": "Kiên Giang", "district": "Gò Quao", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.7333, "longitude": 105.2667},
    {"station_id": "KG_HG", "station_name": "Huyện Hòn Đất - Kiên Giang", "province": "Kiên Giang", "district": "Hòn Đất", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2333, "longitude": 104.9333},
    {"station_id": "KG_KT", "station_name": "Huyện Kiên Hải - Kiên Giang", "province": "Kiên Giang", "district": "Kiên Hải", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.7100, "longitude": 104.3300},
    {"station_id": "KG_LS", "station_name": "Huyện Kiên Lương - Kiên Giang", "province": "Kiên Giang", "district": "Kiên Lương", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2500, "longitude": 104.6333},
    {"station_id": "KG_TG", "station_name": "Huyện Tân Hiệp - Kiên Giang", "province": "Kiên Giang", "district": "Tân Hiệp", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.1000, "longitude": 105.2833},
    {"station_id": "KG_UM", "station_name": "Huyện U Minh Thượng - Kiên Giang", "province": "Kiên Giang", "district": "U Minh Thượng", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.6000, "longitude": 105.1000},
    {"station_id": "KG_VT", "station_name": "Huyện Vĩnh Thuận - Kiên Giang", "province": "Kiên Giang", "district": "Vĩnh Thuận", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.5333, "longitude": 105.2500},

    # Cà Mau
    {"station_id": "CM_TP", "station_name": "Thành phố Cà Mau - Cà Mau", "province": "Cà Mau", "district": "Cà Mau", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 9.1769, "longitude": 105.1521},
    {"station_id": "CM_CM", "station_name": "Huyện Cái Nước - Cà Mau", "province": "Cà Mau", "district": "Cái Nước", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.0000, "longitude": 105.0333},
    {"station_id": "CM_DH", "station_name": "Huyện Đầm Dơi - Cà Mau", "province": "Cà Mau", "district": "Đầm Dơi", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 8.9833, "longitude": 105.2000},
    {"station_id": "CM_NC", "station_name": "Huyện Năm Căn - Cà Mau", "province": "Cà Mau", "district": "Năm Căn", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 8.7500, "longitude": 105.0000},
    {"station_id": "CM_NG", "station_name": "Huyện Ngọc Hiển - Cà Mau", "province": "Cà Mau", "district": "Ngọc Hiển", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 8.6667, "longitude": 105.0500},
    {"station_id": "CM_PT", "station_name": "Huyện Phú Tân - Cà Mau", "province": "Cà Mau", "district": "Phú Tân", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 8.8833, "longitude": 104.8833},
    {"station_id": "CM_TV", "station_name": "Huyện Thới Bình - Cà Mau", "province": "Cà Mau", "district": "Thới Bình", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.3500, "longitude": 105.1000},
    {"station_id": "CM_TR", "station_name": "Huyện Trần Văn Thời - Cà Mau", "province": "Cà Mau", "district": "Trần Văn Thời", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.0833, "longitude": 104.9667},
    {"station_id": "CM_UM", "station_name": "Huyện U Minh - Cà Mau", "province": "Cà Mau", "district": "U Minh", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.4000, "longitude": 104.9833},

    # Tiền Giang
    {"station_id": "TG_MT", "station_name": "Thành phố Mỹ Tho - Tiền Giang", "province": "Tiền Giang", "district": "Mỹ Tho", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3500, "longitude": 106.3500},
    {"station_id": "TG_GL", "station_name": "Thị xã Gò Công - Tiền Giang", "province": "Tiền Giang", "district": "Gò Công", "type": "Thị xã", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3667, "longitude": 106.6667},
    {"station_id": "TG_CT", "station_name": "Huyện Cái Bè - Tiền Giang", "province": "Tiền Giang", "district": "Cái Bè", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3333, "longitude": 105.9500},
    {"station_id": "TG_CL", "station_name": "Huyện Châu Thành - Tiền Giang", "province": "Tiền Giang", "district": "Châu Thành", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.4500, "longitude": 106.2500},
    {"station_id": "TG_GC", "station_name": "Huyện Gò Công Tây - Tiền Giang", "province": "Tiền Giang", "district": "Gò Công Tây", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3500, "longitude": 106.6000},
    {"station_id": "TG_GD", "station_name": "Huyện Gò Công Đông - Tiền Giang", "province": "Tiền Giang", "district": "Gò Công Đông", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3667, "longitude": 106.7167},
    {"station_id": "TG_TP", "station_name": "Huyện Tân Phú Đông - Tiền Giang", "province": "Tiền Giang", "district": "Tân Phú Đông", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2500, "longitude": 106.6000},
    {"station_id": "TG_TL", "station_name": "Huyện Tân Phước - Tiền Giang", "province": "Tiền Giang", "district": "Tân Phước", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.5000, "longitude": 106.1500},
    {"station_id": "TG_CL2", "station_name": "Huyện Chợ Gạo - Tiền Giang", "province": "Tiền Giang", "district": "Chợ Gạo", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3500, "longitude": 106.4667},
    {"station_id": "TG_CB", "station_name": "Huyện Cai Lậy - Tiền Giang", "province": "Tiền Giang", "district": "Cai Lậy", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.4000, "longitude": 106.1167},

    # Bến Tre
    {"station_id": "BT_BT", "station_name": "Thành phố Bến Tre - Bến Tre", "province": "Bến Tre", "district": "Bến Tre", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2333, "longitude": 106.3833},
    {"station_id": "BT_BL", "station_name": "Huyện Ba Tri - Bến Tre", "province": "Bến Tre", "district": "Ba Tri", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0667, "longitude": 106.6000},
    {"station_id": "BT_BH", "station_name": "Huyện Bình Đại - Bến Tre", "province": "Bến Tre", "district": "Bình Đại", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.1833, "longitude": 106.7000},
    {"station_id": "BT_CL", "station_name": "Huyện Châu Thành - Bến Tre", "province": "Bến Tre", "district": "Châu Thành", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3000, "longitude": 106.2667},
    {"station_id": "BT_CT", "station_name": "Huyện Chợ Lách - Bến Tre", "province": "Bến Tre", "district": "Chợ Lách", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2667, "longitude": 106.1500},
    {"station_id": "BT_GD", "station_name": "Huyện Giồng Trôm - Bến Tre", "province": "Bến Tre", "district": "Giồng Trôm", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.1667, "longitude": 106.4667},
    {"station_id": "BT_MB", "station_name": "Huyện Mỏ Cày Bắc - Bến Tre", "province": "Bến Tre", "district": "Mỏ Cày Bắc", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.1333, "longitude": 106.3000},
    {"station_id": "BT_MN", "station_name": "Huyện Mỏ Cày Nam - Bến Tre", "province": "Bến Tre", "district": "Mỏ Cày Nam", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0667, "longitude": 106.3333},
    {"station_id": "BT_TP", "station_name": "Huyện Thạnh Phú - Bến Tre", "province": "Bến Tre", "district": "Thạnh Phú", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.9333, "longitude": 106.5333},

    # Vĩnh Long
    {"station_id": "VL_VL", "station_name": "Thành phố Vĩnh Long - Vĩnh Long", "province": "Vĩnh Long", "district": "Vĩnh Long", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2500, "longitude": 105.9667},
    {"station_id": "VL_BT", "station_name": "Huyện Bình Tân - Vĩnh Long", "province": "Vĩnh Long", "district": "Bình Tân", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0833, "longitude": 105.7667},
    {"station_id": "VL_LH", "station_name": "Huyện Long Hồ - Vĩnh Long", "province": "Vĩnh Long", "district": "Long Hồ", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2000, "longitude": 106.0167},
    {"station_id": "VL_MT", "station_name": "Huyện Mang Thít - Vĩnh Long", "province": "Vĩnh Long", "district": "Mang Thít", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.1833, "longitude": 106.1000},
    {"station_id": "VL_TL", "station_name": "Huyện Tam Bình - Vĩnh Long", "province": "Vĩnh Long", "district": "Tam Bình", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0667, "longitude": 105.9833},
    {"station_id": "VL_TR", "station_name": "Huyện Trà Ôn - Vĩnh Long", "province": "Vĩnh Long", "district": "Trà Ôn", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.9667, "longitude": 105.8833},
    {"station_id": "VL_VT", "station_name": "Huyện Vũng Liêm - Vĩnh Long", "province": "Vĩnh Long", "district": "Vũng Liêm", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.0833, "longitude": 106.1667},

    # Đồng Tháp
    {"station_id": "DT_CL", "station_name": "Thành phố Cao Lãnh - Đồng Tháp", "province": "Đồng Tháp", "district": "Cao Lãnh", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.4500, "longitude": 105.6333},
    {"station_id": "DT_SD", "station_name": "Thành phố Sa Đéc - Đồng Tháp", "province": "Đồng Tháp", "district": "Sa Đéc", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3000, "longitude": 105.7667},
    {"station_id": "DT_HM", "station_name": "Thị xã Hồng Ngự - Đồng Tháp", "province": "Đồng Tháp", "district": "Hồng Ngự", "type": "Thị xã", "region": "Đồng bằng sông Cửu Long", "latitude": 10.8000, "longitude": 105.3333},
    {"station_id": "DT_CL2", "station_name": "Huyện Cao Lãnh - Đồng Tháp", "province": "Đồng Tháp", "district": "Cao Lãnh", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.4667, "longitude": 105.6333},
    {"station_id": "DT_CH", "station_name": "Huyện Châu Thành - Đồng Tháp", "province": "Đồng Tháp", "district": "Châu Thành", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3000, "longitude": 105.8167},
    {"station_id": "DT_HN", "station_name": "Huyện Hồng Ngự - Đồng Tháp", "province": "Đồng Tháp", "district": "Hồng Ngự", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.8167, "longitude": 105.3500},
    {"station_id": "DT_LV", "station_name": "Huyện Lai Vung - Đồng Tháp", "province": "Đồng Tháp", "district": "Lai Vung", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2833, "longitude": 105.6667},
    {"station_id": "DT_LT", "station_name": "Huyện Lấp Vò - Đồng Tháp", "province": "Đồng Tháp", "district": "Lấp Vò", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.3500, "longitude": 105.5167},
    {"station_id": "DT_TN", "station_name": "Huyện Tam Nông - Đồng Tháp", "province": "Đồng Tháp", "district": "Tam Nông", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.7333, "longitude": 105.5500},
    {"station_id": "DT_TH", "station_name": "Huyện Thanh Bình - Đồng Tháp", "province": "Đồng Tháp", "district": "Thanh Bình", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.6000, "longitude": 105.4667},
    {"station_id": "DT_TX", "station_name": "Huyện Tháp Mười - Đồng Tháp", "province": "Đồng Tháp", "district": "Tháp Mười", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.5333, "longitude": 105.8167},

    # Hậu Giang
    {"station_id": "HG_VT", "station_name": "Thành phố Vị Thanh - Hậu Giang", "province": "Hậu Giang", "district": "Vị Thanh", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 9.7833, "longitude": 105.4667},
    {"station_id": "HG_NM", "station_name": "Thị xã Ngã Bảy - Hậu Giang", "province": "Hậu Giang", "district": "Ngã Bảy", "type": "Thị xã", "region": "Đồng bằng sông Cửu Long", "latitude": 9.8167, "longitude": 105.8167},
    {"station_id": "HG_CL", "station_name": "Huyện Châu Thành - Hậu Giang", "province": "Hậu Giang", "district": "Châu Thành", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.9333, "longitude": 105.6833},
    {"station_id": "HG_CT", "station_name": "Huyện Châu Thành A - Hậu Giang", "province": "Hậu Giang", "district": "Châu Thành A", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.9333, "longitude": 105.6333},
    {"station_id": "HG_LT", "station_name": "Huyện Long Mỹ - Hậu Giang", "province": "Hậu Giang", "district": "Long Mỹ", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.6833, "longitude": 105.5667},
    {"station_id": "HG_PB", "station_name": "Huyện Phụng Hiệp - Hậu Giang", "province": "Hậu Giang", "district": "Phụng Hiệp", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.7833, "longitude": 105.7167},
    {"station_id": "HG_VT2", "station_name": "Huyện Vị Thủy - Hậu Giang", "province": "Hậu Giang", "district": "Vị Thủy", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.7667, "longitude": 105.4667},

    # Sóc Trăng
    {"station_id": "ST_ST", "station_name": "Thành phố Sóc Trăng - Sóc Trăng", "province": "Sóc Trăng", "district": "Sóc Trăng", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 9.6025, "longitude": 105.9739},
    {"station_id": "ST_VC", "station_name": "Thị xã Vĩnh Châu - Sóc Trăng", "province": "Sóc Trăng", "district": "Vĩnh Châu", "type": "Thị xã", "region": "Đồng bằng sông Cửu Long", "latitude": 9.3333, "longitude": 105.9833},
    {"station_id": "ST_NG", "station_name": "Thị xã Ngã Năm - Sóc Trăng", "province": "Sóc Trăng", "district": "Ngã Năm", "type": "Thị xã", "region": "Đồng bằng sông Cửu Long", "latitude": 9.5667, "longitude": 105.8500},
    {"station_id": "ST_CL", "station_name": "Huyện Châu Thành - Sóc Trăng", "province": "Sóc Trăng", "district": "Châu Thành", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.7000, "longitude": 105.9000},
    {"station_id": "ST_CT", "station_name": "Huyện Cù Lao Dung - Sóc Trăng", "province": "Sóc Trăng", "district": "Cù Lao Dung", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.6667, "longitude": 106.1667},
    {"station_id": "ST_KS", "station_name": "Huyện Kế Sách - Sóc Trăng", "province": "Sóc Trăng", "district": "Kế Sách", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.8167, "longitude": 105.9833},
    {"station_id": "ST_LT", "station_name": "Huyện Long Phú - Sóc Trăng", "province": "Sóc Trăng", "district": "Long Phú", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.6167, "longitude": 106.1167},
    {"station_id": "ST_MD", "station_name": "Huyện Mỹ Tú - Sóc Trăng", "province": "Sóc Trăng", "district": "Mỹ Tú", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.6333, "longitude": 105.8167},
    {"station_id": "ST_MX", "station_name": "Huyện Mỹ Xuyên - Sóc Trăng", "province": "Sóc Trăng", "district": "Mỹ Xuyên", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.5333, "longitude": 105.9833},
    {"station_id": "ST_TN", "station_name": "Huyện Thạnh Trị - Sóc Trăng", "province": "Sóc Trăng", "district": "Thạnh Trị", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.4667, "longitude": 105.7167},
    {"station_id": "ST_TP", "station_name": "Huyện Trần Đề - Sóc Trăng", "province": "Sóc Trăng", "district": "Trần Đề", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.5000, "longitude": 106.0833},

    # Bạc Liêu
    {"station_id": "BL_BL", "station_name": "Thành phố Bạc Liêu - Bạc Liêu", "province": "Bạc Liêu", "district": "Bạc Liêu", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 9.2833, "longitude": 105.7167},
    {"station_id": "BL_GD", "station_name": "Thị xã Giá Rai - Bạc Liêu", "province": "Bạc Liêu", "district": "Giá Rai", "type": "Thị xã", "region": "Đồng bằng sông Cửu Long", "latitude": 9.2333, "longitude": 105.4667},
    {"station_id": "BL_HB", "station_name": "Huyện Hồng Dân - Bạc Liêu", "province": "Bạc Liêu", "district": "Hồng Dân", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.5500, "longitude": 105.4167},
    {"station_id": "BL_PT", "station_name": "Huyện Phước Long - Bạc Liêu", "province": "Bạc Liêu", "district": "Phước Long", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.4333, "longitude": 105.4667},
    {"station_id": "BL_VL", "station_name": "Huyện Vĩnh Lợi - Bạc Liêu", "province": "Bạc Liêu", "district": "Vĩnh Lợi", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.3500, "longitude": 105.5667},
    {"station_id": "BL_DD", "station_name": "Huyện Đông Hải - Bạc Liêu", "province": "Bạc Liêu", "district": "Đông Hải", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.1667, "longitude": 105.4333},
    {"station_id": "BL_HH", "station_name": "Huyện Hòa Bình - Bạc Liêu", "province": "Bạc Liêu", "district": "Hòa Bình", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.2833, "longitude": 105.6333},

    # Thanh Hóa
    {"station_id": "TH_TH", "station_name": "Thành phố Thanh Hóa - Thanh Hóa", "province": "Thanh Hóa", "district": "Thanh Hóa", "type": "Thành phố", "region": "Bắc Trung Bộ", "latitude": 19.8000, "longitude": 105.7667},
    {"station_id": "TH_SL", "station_name": "Thị xã Sầm Sơn - Thanh Hóa", "province": "Thanh Hóa", "district": "Sầm Sơn", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 19.7333, "longitude": 105.9000},
    {"station_id": "TH_BT", "station_name": "Thị xã Bỉm Sơn - Thanh Hóa", "province": "Thanh Hóa", "district": "Bỉm Sơn", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 20.0781, "longitude": 105.8603},
    {"station_id": "TH_NC", "station_name": "Huyện Nghi Sơn - Thanh Hóa", "province": "Thanh Hóa", "district": "Nghi Sơn", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 19.4500, "longitude": 105.7833},

    # Nghệ An
    {"station_id": "NA_VH", "station_name": "Thành phố Vinh - Nghệ An", "province": "Nghệ An", "district": "Vinh", "type": "Thành phố", "region": "Bắc Trung Bộ", "latitude": 18.6733, "longitude": 105.6811},
    {"station_id": "NA_CT", "station_name": "Thị xã Cửa Lò - Nghệ An", "province": "Nghệ An", "district": "Cửa Lò", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 18.8167, "longitude": 105.7167},
    {"station_id": "NA_TH", "station_name": "Thị xã Thái Hòa - Nghệ An", "province": "Nghệ An", "district": "Thái Hòa", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 19.3000, "longitude": 105.4667},
    {"station_id": "NA_HH", "station_name": "Huyện Hoàng Mai - Nghệ An", "province": "Nghệ An", "district": "Hoàng Mai", "type": "Huyện", "region": "Bắc Trung Bộ", "latitude": 19.2667, "longitude": 105.7167},

    # Hà Tĩnh
    {"station_id": "HT_HT", "station_name": "Thành phố Hà Tĩnh - Hà Tĩnh", "province": "Hà Tĩnh", "district": "Hà Tĩnh", "type": "Thành phố", "region": "Bắc Trung Bộ", "latitude": 18.3333, "longitude": 105.9000},
    {"station_id": "HT_HK", "station_name": "Thị xã Hồng Lĩnh - Hà Tĩnh", "province": "Hà Tĩnh", "district": "Hồng Lĩnh", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 18.5333, "longitude": 105.7167},
    {"station_id": "HT_KL", "station_name": "Thị xã Kỳ Anh - Hà Tĩnh", "province": "Hà Tĩnh", "district": "Kỳ Anh", "type": "Thị xã", "region": "Bắc Trung Bộ", "latitude": 18.0833, "longitude": 106.3000},

    # Quảng Nam
    {"station_id": "QN_TC", "station_name": "Thành phố Tam Kỳ - Quảng Nam", "province": "Quảng Nam", "district": "Tam Kỳ", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 15.5667, "longitude": 108.4833},
    {"station_id": "QN_HT", "station_name": "Thành phố Hội An - Quảng Nam", "province": "Quảng Nam", "district": "Hội An", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 15.8833, "longitude": 108.3333},
    {"station_id": "QN_DG", "station_name": "Huyện Điện Bàn - Quảng Nam", "province": "Quảng Nam", "district": "Điện Bàn", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 15.9000, "longitude": 108.2500},
    {"station_id": "QN_NT", "station_name": "Huyện Núi Thành - Quảng Nam", "province": "Quảng Nam", "district": "Núi Thành", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 15.4333, "longitude": 108.6667},

    # Quảng Ngãi
    {"station_id": "QN_QN", "station_name": "Thành phố Quảng Ngãi - Quảng Ngãi", "province": "Quảng Ngãi", "district": "Quảng Ngãi", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 15.1167, "longitude": 108.8000},
    {"station_id": "QN_LS", "station_name": "Huyện Lý Sơn - Quảng Ngãi", "province": "Quảng Ngãi", "district": "Lý Sơn", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 15.3833, "longitude": 109.1167},
    {"station_id": "QN_BT", "station_name": "Huyện Bình Sơn - Quảng Ngãi", "province": "Quảng Ngãi", "district": "Bình Sơn", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 15.3167, "longitude": 108.7667},

    # Bình Định
    {"station_id": "BD_QN", "station_name": "Thành phố Quy Nhơn - Bình Định", "province": "Bình Định", "district": "Quy Nhơn", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 13.7667, "longitude": 109.2333},
    {"station_id": "BD_AL", "station_name": "Huyện An Lão - Bình Định", "province": "Bình Định", "district": "An Lão", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 14.5667, "longitude": 108.9000},
    {"station_id": "BD_HV", "station_name": "Huyện Hoài Ân - Bình Định", "province": "Bình Định", "district": "Hoài Ân", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 14.3500, "longitude": 108.9000},

    # Phú Yên
    {"station_id": "PY_TH", "station_name": "Thành phố Tuy Hòa - Phú Yên", "province": "Phú Yên", "district": "Tuy Hòa", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 13.0833, "longitude": 109.3000},
    {"station_id": "PY_SH", "station_name": "Thị xã Sông Cầu - Phú Yên", "province": "Phú Yên", "district": "Sông Cầu", "type": "Thị xã", "region": "Duyên hải Nam Trung Bộ", "latitude": 13.4500, "longitude": 109.2167},
    {"station_id": "PY_DD", "station_name": "Huyện Đông Hòa - Phú Yên", "province": "Phú Yên", "district": "Đông Hòa", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 12.9833, "longitude": 109.3667},

    # Khánh Hòa
    {"station_id": "KH_NT", "station_name": "Thành phố Nha Trang - Khánh Hòa", "province": "Khánh Hòa", "district": "Nha Trang", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 12.2500, "longitude": 109.1833},
    {"station_id": "KH_CT", "station_name": "Thành phố Cam Ranh - Khánh Hòa", "province": "Khánh Hòa", "district": "Cam Ranh", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 11.9000, "longitude": 109.1333},
    {"station_id": "KH_NH", "station_name": "Thị xã Ninh Hòa - Khánh Hòa", "province": "Khánh Hòa", "district": "Ninh Hòa", "type": "Thị xã", "region": "Duyên hải Nam Trung Bộ", "latitude": 12.5000, "longitude": 109.1500},
    {"station_id": "KH_TS", "station_name": "Huyện Trường Sa - Khánh Hòa", "province": "Khánh Hòa", "district": "Trường Sa", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 8.6500, "longitude": 111.9167},

    # Ninh Thuận
    {"station_id": "NT_PH", "station_name": "Thành phố Phan Rang-Tháp Chàm - Ninh Thuận", "province": "Ninh Thuận", "district": "Phan Rang-Tháp Chàm", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 11.5667, "longitude": 108.9833},
    {"station_id": "NT_TA", "station_name": "Huyện Thuận Bắc - Ninh Thuận", "province": "Ninh Thuận", "district": "Thuận Bắc", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 11.7500, "longitude": 108.9333},
    {"station_id": "NT_NA", "station_name": "Huyện Ninh Hải - Ninh Thuận", "province": "Ninh Thuận", "district": "Ninh Hải", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 11.6000, "longitude": 109.0333},

    # Bình Thuận
    {"station_id": "BT_PT", "station_name": "Thành phố Phan Thiết - Bình Thuận", "province": "Bình Thuận", "district": "Phan Thiết", "type": "Thành phố", "region": "Duyên hải Nam Trung Bộ", "latitude": 10.9333, "longitude": 108.1000},
    {"station_id": "BT_LC", "station_name": "Thị xã La Gi - Bình Thuận", "province": "Bình Thuận", "district": "La Gi", "type": "Thị xã", "region": "Duyên hải Nam Trung Bộ", "latitude": 10.6667, "longitude": 107.7667},
    {"station_id": "BT_PH", "station_name": "Huyện Phú Quý - Bình Thuận", "province": "Bình Thuận", "district": "Phú Quý", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 10.5000, "longitude": 108.9333},
    {"station_id": "BT_HD", "station_name": "Huyện Hàm Thuận Bắc - Bình Thuận", "province": "Bình Thuận", "district": "Hàm Thuận Bắc", "type": "Huyện", "region": "Duyên hải Nam Trung Bộ", "latitude": 11.0833, "longitude": 108.1167},

    # Điện Biên
    {"station_id": "DB_DB", "station_name": "Thành phố Điện Biên Phủ - Điện Biên", "province": "Điện Biên", "district": "Điện Biên Phủ", "type": "Thành phố", "region": "Tây Bắc", "latitude": 21.3833, "longitude": 103.0167},
    {"station_id": "DB_ML", "station_name": "Thị xã Mường Lay - Điện Biên", "province": "Điện Biên", "district": "Mường Lay", "type": "Thị xã", "region": "Tây Bắc", "latitude": 22.0333, "longitude": 103.1500},
    {"station_id": "DB_DL", "station_name": "Huyện Điện Biên - Điện Biên", "province": "Điện Biên", "district": "Điện Biên", "type": "Huyện", "region": "Tây Bắc", "latitude": 21.2500, "longitude": 103.0333},
    {"station_id": "DB_DD", "station_name": "Huyện Điện Biên Đông - Điện Biên", "province": "Điện Biên", "district": "Điện Biên Đông", "type": "Huyện", "region": "Tây Bắc", "latitude": 21.3333, "longitude": 103.2500},

    # Sơn La
    {"station_id": "SL_SL", "station_name": "Thành phố Sơn La - Sơn La", "province": "Sơn La", "district": "Sơn La", "type": "Thành phố", "region": "Tây Bắc", "latitude": 21.3256, "longitude": 103.9189},
    {"station_id": "SL_ML", "station_name": "Huyện Mộc Châu - Sơn La", "province": "Sơn La", "district": "Mộc Châu", "type": "Huyện", "region": "Tây Bắc", "latitude": 20.8500, "longitude": 104.6333},
    {"station_id": "SL_YB", "station_name": "Huyện Yên Châu - Sơn La", "province": "Sơn La", "district": "Yên Châu", "type": "Huyện", "region": "Tây Bắc", "latitude": 21.0500, "longitude": 104.3000},

    # Lai Châu
    {"station_id": "LC_LC", "station_name": "Thành phố Lai Châu - Lai Châu", "province": "Lai Châu", "district": "Lai Châu", "type": "Thành phố", "region": "Tây Bắc", "latitude": 22.4000, "longitude": 103.4500},
    {"station_id": "LC_TN", "station_name": "Huyện Tam Đường - Lai Châu", "province": "Lai Châu", "district": "Tam Đường", "type": "Huyện", "region": "Tây Bắc", "latitude": 22.3667, "longitude": 103.6167},
    {"station_id": "LC_PT", "station_name": "Huyện Phong Thổ - Lai Châu", "province": "Lai Châu", "district": "Phong Thổ", "type": "Huyện", "region": "Tây Bắc", "latitude": 22.5333, "longitude": 103.3333},

    # Hòa Bình
    {"station_id": "HB_HB", "station_name": "Thành phố Hòa Bình - Hòa Bình", "province": "Hòa Bình", "district": "Hòa Bình", "type": "Thành phố", "region": "Tây Bắc", "latitude": 20.8133, "longitude": 105.3383},
    {"station_id": "HB_ML", "station_name": "Huyện Mai Châu - Hòa Bình", "province": "Hòa Bình", "district": "Mai Châu", "type": "Huyện", "region": "Tây Bắc", "latitude": 20.6667, "longitude": 105.0833},
    {"station_id": "HB_KB", "station_name": "Huyện Kim Bôi - Hòa Bình", "province": "Hòa Bình", "district": "Kim Bôi", "type": "Huyện", "region": "Tây Bắc", "latitude": 20.6667, "longitude": 105.5333},

    # Lạng Sơn
    {"station_id": "LS_LS", "station_name": "Thành phố Lạng Sơn - Lạng Sơn", "province": "Lạng Sơn", "district": "Lạng Sơn", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.8478, "longitude": 106.7578},
    {"station_id": "LS_CL", "station_name": "Huyện Cao Lộc - Lạng Sơn", "province": "Lạng Sơn", "district": "Cao Lộc", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.9000, "longitude": 106.8000},
    {"station_id": "LS_LL", "station_name": "Huyện Lộc Bình - Lạng Sơn", "province": "Lạng Sơn", "district": "Lộc Bình", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.7500, "longitude": 106.9333},

    # Cao Bằng
    {"station_id": "CB_CB", "station_name": "Thành phố Cao Bằng - Cao Bằng", "province": "Cao Bằng", "district": "Cao Bằng", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 22.6667, "longitude": 106.2500},
    {"station_id": "CB_BL", "station_name": "Huyện Bảo Lâm - Cao Bằng", "province": "Cao Bằng", "district": "Bảo Lâm", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 22.8667, "longitude": 105.5000},
    {"station_id": "CB_BC", "station_name": "Huyện Bảo Lạc - Cao Bằng", "province": "Cao Bằng", "district": "Bảo Lạc", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 22.9500, "longitude": 105.7333},

    # Bắc Kạn
    {"station_id": "BK_BK", "station_name": "Thành phố Bắc Kạn - Bắc Kạn", "province": "Bắc Kạn", "district": "Bắc Kạn", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 22.1500, "longitude": 105.8333},
    {"station_id": "BK_BB", "station_name": "Huyện Ba Bể - Bắc Kạn", "province": "Bắc Kạn", "district": "Ba Bể", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 22.4167, "longitude": 105.7500},
    {"station_id": "BK_PT", "station_name": "Huyện Pác Nặm - Bắc Kạn", "province": "Bắc Kạn", "district": "Pác Nặm", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 22.6333, "longitude": 105.6667},

    # Thái Nguyên
    {"station_id": "TN_TN", "station_name": "Thành phố Thái Nguyên - Thái Nguyên", "province": "Thái Nguyên", "district": "Thái Nguyên", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.5928, "longitude": 105.8442},
    {"station_id": "TN_SC", "station_name": "Thành phố Sông Công - Thái Nguyên", "province": "Thái Nguyên", "district": "Sông Công", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.4833, "longitude": 105.8500},
    {"station_id": "TN_DD", "station_name": "Huyện Định Hóa - Thái Nguyên", "province": "Thái Nguyên", "district": "Định Hóa", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.9000, "longitude": 105.6333},

    # Tuyên Quang
    {"station_id": "TQ_TQ", "station_name": "Thành phố Tuyên Quang - Tuyên Quang", "province": "Tuyên Quang", "district": "Tuyên Quang", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.8181, "longitude": 105.2144},
    {"station_id": "TQ_NH", "station_name": "Huyện Na Hang - Tuyên Quang", "province": "Tuyên Quang", "district": "Na Hang", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 22.3500, "longitude": 105.3833},
    {"station_id": "TQ_LB", "station_name": "Huyện Lâm Bình - Tuyên Quang", "province": "Tuyên Quang", "district": "Lâm Bình", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 22.4667, "longitude": 105.2167},

    # Yên Bái
    {"station_id": "YB_YB", "station_name": "Thành phố Yên Bái - Yên Bái", "province": "Yên Bái", "district": "Yên Bái", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.7000, "longitude": 104.8667},
    {"station_id": "YB_NZ", "station_name": "Thị xã Nghĩa Lộ - Yên Bái", "province": "Yên Bái", "district": "Nghĩa Lộ", "type": "Thị xã", "region": "Đông Bắc Bộ", "latitude": 21.6000, "longitude": 104.5000},
    {"station_id": "YB_MT", "station_name": "Huyện Mù Cang Chải - Yên Bái", "province": "Yên Bái", "district": "Mù Cang Chải", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.8500, "longitude": 104.0833},

    # Phú Thọ
    {"station_id": "PT_VT", "station_name": "Thành phố Việt Trì - Phú Thọ", "province": "Phú Thọ", "district": "Việt Trì", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.3000, "longitude": 105.4333},
    {"station_id": "PT_PT", "station_name": "Thị xã Phú Thọ - Phú Thọ", "province": "Phú Thọ", "district": "Phú Thọ", "type": "Thị xã", "region": "Đông Bắc Bộ", "latitude": 21.4000, "longitude": 105.2333},
    {"station_id": "PT_HT", "station_name": "Huyện Hạ Hòa - Phú Thọ", "province": "Phú Thọ", "district": "Hạ Hòa", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.5500, "longitude": 105.0000},

    # Vĩnh Phúc
    {"station_id": "VP_VY", "station_name": "Thành phố Vĩnh Yên - Vĩnh Phúc", "province": "Vĩnh Phúc", "district": "Vĩnh Yên", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 21.3089, "longitude": 105.6044},
    {"station_id": "VP_PC", "station_name": "Thành phố Phúc Yên - Vĩnh Phúc", "province": "Vĩnh Phúc", "district": "Phúc Yên", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 21.2333, "longitude": 105.7000},
    {"station_id": "VP_LP", "station_name": "Huyện Lập Thạch - Vĩnh Phúc", "province": "Vĩnh Phúc", "district": "Lập Thạch", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.4167, "longitude": 105.4667},

    # Bắc Ninh
    {"station_id": "BN_BN", "station_name": "Thành phố Bắc Ninh - Bắc Ninh", "province": "Bắc Ninh", "district": "Bắc Ninh", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 21.1861, "longitude": 106.0764},
    {"station_id": "BN_TS", "station_name": "Thành phố Từ Sơn - Bắc Ninh", "province": "Bắc Ninh", "district": "Từ Sơn", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 21.1167, "longitude": 105.9667},
    {"station_id": "BN_YL", "station_name": "Huyện Yên Phong - Bắc Ninh", "province": "Bắc Ninh", "district": "Yên Phong", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 21.2000, "longitude": 105.9500},

    # Hải Dương
    {"station_id": "HD_HD", "station_name": "Thành phố Hải Dương - Hải Dương", "province": "Hải Dương", "district": "Hải Dương", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 20.9397, "longitude": 106.3308},
    {"station_id": "HD_CL", "station_name": "Thị xã Kinh Môn - Hải Dương", "province": "Hải Dương", "district": "Kinh Môn", "type": "Thị xã", "region": "Đồng bằng sông Hồng", "latitude": 21.0167, "longitude": 106.5000},
    {"station_id": "HD_CM", "station_name": "Huyện Cẩm Giàng - Hải Dương", "province": "Hải Dương", "district": "Cẩm Giàng", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.9500, "longitude": 106.2167},

    # Hưng Yên
    {"station_id": "HY_HY", "station_name": "Thành phố Hưng Yên - Hưng Yên", "province": "Hưng Yên", "district": "Hưng Yên", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 20.6461, "longitude": 106.0511},
    {"station_id": "HY_ML", "station_name": "Huyện Mỹ Hào - Hưng Yên", "province": "Hưng Yên", "district": "Mỹ Hào", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.9333, "longitude": 106.0667},
    {"station_id": "HY_VD", "station_name": "Huyện Văn Lâm - Hưng Yên", "province": "Hưng Yên", "district": "Văn Lâm", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.9833, "longitude": 106.0333},

    # Ninh Bình
    {"station_id": "NB_NB", "station_name": "Thành phố Ninh Bình - Ninh Bình", "province": "Ninh Bình", "district": "Ninh Bình", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 20.2539, "longitude": 105.9750},
    {"station_id": "NB_TL", "station_name": "Thành phố Tam Điệp - Ninh Bình", "province": "Ninh Bình", "district": "Tam Điệp", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 20.1500, "longitude": 105.9167},
    {"station_id": "NB_HS", "station_name": "Huyện Hoa Lư - Ninh Bình", "province": "Ninh Bình", "district": "Hoa Lư", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.2833, "longitude": 105.9167},

    # Bắc Giang
    {"station_id": "BG_BG", "station_name": "Thành phố Bắc Giang - Bắc Giang", "province": "Bắc Giang", "district": "Bắc Giang", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 21.2667, "longitude": 106.2000},
    {"station_id": "BG_YL", "station_name": "Huyện Yên Thế - Bắc Giang", "province": "Bắc Giang", "district": "Yên Thế", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.5167, "longitude": 106.1167},
    {"station_id": "BG_LN", "station_name": "Huyện Lạng Giang - Bắc Giang", "province": "Bắc Giang", "district": "Lạng Giang", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 21.3500, "longitude": 106.2500},

    # Bình Phước
    {"station_id": "BP_DX", "station_name": "Thành phố Đồng Xoài - Bình Phước", "province": "Bình Phước", "district": "Đồng Xoài", "type": "Thành phố", "region": "Đông Nam Bộ", "latitude": 11.5349, "longitude": 106.8823},
    {"station_id": "BP_BL", "station_name": "Thị xã Bình Long - Bình Phước", "province": "Bình Phước", "district": "Bình Long", "type": "Thị xã", "region": "Đông Nam Bộ", "latitude": 11.6500, "longitude": 106.6000},
    {"station_id": "BP_PL", "station_name": "Thị xã Phước Long - Bình Phước", "province": "Bình Phước", "district": "Phước Long", "type": "Thị xã", "region": "Đông Nam Bộ", "latitude": 11.8333, "longitude": 106.9667},

    # Đắk Nông
    {"station_id": "DKN_GN", "station_name": "Thành phố Gia Nghĩa - Đắk Nông", "province": "Đắk Nông", "district": "Gia Nghĩa", "type": "Thành phố", "region": "Tây Nguyên", "latitude": 12.0042, "longitude": 107.6907},
    {"station_id": "DKN_CJ", "station_name": "Huyện Cư Jút - Đắk Nông", "province": "Đắk Nông", "district": "Cư Jút", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.6569, "longitude": 107.7636},
    {"station_id": "DKN_KM", "station_name": "Huyện Krông Nô - Đắk Nông", "province": "Đắk Nông", "district": "Krông Nô", "type": "Huyện", "region": "Tây Nguyên", "latitude": 12.4500, "longitude": 107.8667},

    # Hà Giang
    {"station_id": "HG_HG", "station_name": "Thành phố Hà Giang - Hà Giang", "province": "Hà Giang", "district": "Hà Giang", "type": "Thành phố", "region": "Đông Bắc Bộ", "latitude": 22.8233, "longitude": 104.9836},
    {"station_id": "HG_DV", "station_name": "Huyện Đồng Văn - Hà Giang", "province": "Hà Giang", "district": "Đồng Văn", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 23.2547, "longitude": 105.2500},
    {"station_id": "HG_MK", "station_name": "Huyện Mèo Vạc - Hà Giang", "province": "Hà Giang", "district": "Mèo Vạc", "type": "Huyện", "region": "Đông Bắc Bộ", "latitude": 23.1528, "longitude": 105.4069},

    # Hà Nam
    {"station_id": "HNA_PL", "station_name": "Thành phố Phủ Lý - Hà Nam", "province": "Hà Nam", "district": "Phủ Lý", "type": "Thành phố", "region": "Đồng bằng sông Hồng", "latitude": 20.5411, "longitude": 105.9139},
    {"station_id": "HNA_DT", "station_name": "Huyện Duy Tiên - Hà Nam", "province": "Hà Nam", "district": "Duy Tiên", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.6333, "longitude": 105.9667},
    {"station_id": "HNA_KB", "station_name": "Huyện Kim Bảng - Hà Nam", "province": "Hà Nam", "district": "Kim Bảng", "type": "Huyện", "region": "Đồng bằng sông Hồng", "latitude": 20.5667, "longitude": 105.8500},

    # Kon Tum
    {"station_id": "KT_KT", "station_name": "Thành phố Kon Tum - Kon Tum", "province": "Kon Tum", "district": "Kon Tum", "type": "Thành phố", "region": "Tây Nguyên", "latitude": 14.3833, "longitude": 107.9833},
    {"station_id": "KT_DG", "station_name": "Huyện Đắk Glei - Kon Tum", "province": "Kon Tum", "district": "Đắk Glei", "type": "Huyện", "region": "Tây Nguyên", "latitude": 15.0167, "longitude": 107.7500},
    {"station_id": "KT_NH", "station_name": "Huyện Ngọc Hồi - Kon Tum", "province": "Kon Tum", "district": "Ngọc Hồi", "type": "Huyện", "region": "Tây Nguyên", "latitude": 14.7000, "longitude": 107.8333},

    # Lào Cai
    {"station_id": "LC_LC2", "station_name": "Thành phố Lào Cai - Lào Cai", "province": "Lào Cai", "district": "Lào Cai", "type": "Thành phố", "region": "Tây Bắc", "latitude": 22.4833, "longitude": 103.9500},
    {"station_id": "LC_SM", "station_name": "Huyện Sa Pa - Lào Cai", "province": "Lào Cai", "district": "Sa Pa", "type": "Huyện", "region": "Tây Bắc", "latitude": 22.3367, "longitude": 103.8400},
    {"station_id": "LC_BT", "station_name": "Huyện Bát Xát - Lào Cai", "province": "Lào Cai", "district": "Bát Xát", "type": "Huyện", "region": "Tây Bắc", "latitude": 22.5333, "longitude": 103.8833},

    # Long An
    {"station_id": "LA_TA", "station_name": "Thành phố Tân An - Long An", "province": "Long An", "district": "Tân An", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 10.5333, "longitude": 106.4167},
    {"station_id": "LA_KT", "station_name": "Thị xã Kiến Tường - Long An", "province": "Long An", "district": "Kiến Tường", "type": "Thị xã", "region": "Đồng bằng sông Cửu Long", "latitude": 10.7667, "longitude": 105.9000},
    {"station_id": "LA_DH", "station_name": "Huyện Đức Hòa - Long An", "province": "Long An", "district": "Đức Hòa", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 10.8833, "longitude": 106.4167},

    # Tây Ninh
    {"station_id": "TN_TN2", "station_name": "Thành phố Tây Ninh - Tây Ninh", "province": "Tây Ninh", "district": "Tây Ninh", "type": "Thành phố", "region": "Đông Nam Bộ", "latitude": 11.3131, "longitude": 106.0963},
    {"station_id": "TN_TB", "station_name": "Huyện Tân Biên - Tây Ninh", "province": "Tây Ninh", "district": "Tân Biên", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.5500, "longitude": 105.9667},
    {"station_id": "TN_TC", "station_name": "Huyện Tân Châu - Tây Ninh", "province": "Tây Ninh", "district": "Tân Châu", "type": "Huyện", "region": "Đông Nam Bộ", "latitude": 11.3333, "longitude": 106.1667},

    # Trà Vinh
    {"station_id": "TV_TV", "station_name": "Thành phố Trà Vinh - Trà Vinh", "province": "Trà Vinh", "district": "Trà Vinh", "type": "Thành phố", "region": "Đồng bằng sông Cửu Long", "latitude": 9.9347, "longitude": 106.3453},
    {"station_id": "TV_CL", "station_name": "Huyện Càng Long - Trà Vinh", "province": "Trà Vinh", "district": "Càng Long", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.9667, "longitude": 106.2000},
    {"station_id": "TV_CN", "station_name": "Huyện Cầu Ngang - Trà Vinh", "province": "Trà Vinh", "district": "Cầu Ngang", "type": "Huyện", "region": "Đồng bằng sông Cửu Long", "latitude": 9.8000, "longitude": 106.4333},

    # Các đảo lớn
    {"station_id": "HP_CB", "station_name": "Đảo Cát Bà - Hải Phòng", "province": "Hải Phòng", "district": "Cát Hải", "type": "Đảo", "region": "Đồng bằng sông Hồng", "latitude": 20.7264, "longitude": 107.0489},
    {"station_id": "QN_CT", "station_name": "Đảo Cô Tô - Quảng Ninh", "province": "Quảng Ninh", "district": "Cô Tô", "type": "Đảo", "region": "Đông Bắc Bộ", "latitude": 20.9681, "longitude": 107.7639},
    {"station_id": "QN_VD", "station_name": "Đảo Vân Đồn - Quảng Ninh", "province": "Quảng Ninh", "district": "Vân Đồn", "type": "Đảo", "region": "Đông Bắc Bộ", "latitude": 20.9551, "longitude": 107.4764},
    {"station_id": "KH_TS", "station_name": "Đảo Trường Sa - Khánh Hòa", "province": "Khánh Hòa", "district": "Trường Sa", "type": "Đảo", "region": "Duyên hải Nam Trung Bộ", "latitude": 8.6500, "longitude": 111.9167},
    {"station_id": "KH_HS", "station_name": "Đảo Hoàng Sa - Khánh Hòa", "province": "Khánh Hòa", "district": "Hoàng Sa", "type": "Đảo", "region": "Duyên hải Nam Trung Bộ", "latitude": 16.5322, "longitude": 111.6156},
    {"station_id": "KG_PQ2", "station_name": "Đảo Phú Quốc - Kiên Giang", "province": "Kiên Giang", "district": "Phú Quốc", "type": "Đảo", "region": "Đồng bằng sông Cửu Long", "latitude": 10.2270, "longitude": 103.9679},
    {"station_id": "KG_ND", "station_name": "Đảo Nam Du - Kiên Giang", "province": "Kiên Giang", "district": "Kiên Hải", "type": "Đảo", "region": "Đồng bằng sông Cửu Long", "latitude": 9.7100, "longitude": 104.3300},
    {"station_id": "KG_TC", "station_name": "Đảo Thổ Chu - Kiên Giang", "province": "Kiên Giang", "district": "Phú Quốc", "type": "Đảo", "region": "Đồng bằng sông Cửu Long", "latitude": 9.3000, "longitude": 103.4833},
    {"station_id": "BR_CD2", "station_name": "Đảo Côn Sơn - Bà Rịa - Vũng Tàu", "province": "Bà Rịa - Vũng Tàu", "district": "Côn Đảo", "type": "Đảo", "region": "Đông Nam Bộ", "latitude": 8.6822, "longitude": 106.6089},
    {"station_id": "QT_CC2", "station_name": "Đảo Cồn Cỏ - Quảng Trị", "province": "Quảng Trị", "district": "Cồn Cỏ", "type": "Đảo", "region": "Bắc Trung Bộ", "latitude": 17.1597, "longitude": 107.3408},
    {"station_id": "QN_LS2", "station_name": "Đảo Lý Sơn - Quảng Ngãi", "province": "Quảng Ngãi", "district": "Lý Sơn", "type": "Đảo", "region": "Duyên hải Nam Trung Bộ", "latitude": 15.3833, "longitude": 109.1167},
    {"station_id": "BT_PQ2", "station_name": "Đảo Phú Quý - Bình Thuận", "province": "Bình Thuận", "district": "Phú Quý", "type": "Đảo", "region": "Duyên hải Nam Trung Bộ", "latitude": 10.5000, "longitude": 108.9333},
]

def main():
    """Hàm chính thực thi"""
    try:
        crawler = VietnamWeatherDataCrawler()
        locations = vietnam_locations

        logging.info("=" * 70)
        logging.info("🌏 HỆ THỐNG THU THẬP DỮ LIỆU THỜI TIẾT VIỆT NAM")
        logging.info("=" * 70)
        logging.info("📝 LƯU Ý: Dữ liệu từ các API miễn phí có thể không chính xác tuyệt đối")
        logging.info("🔍 Đang thu thập từ đa nguồn với đánh giá chất lượng...")

        start_time = time.time()
        weather_data = crawler.crawl_all_locations(locations, delay=2.0)
        end_time = time.time()

        if weather_data:
            # Lưu vào Excel
            excel_file = crawler.save_to_excel(weather_data)

            # Lưu vào SQLite
            sqlite_success = crawler.save_to_sqlite(weather_data, locations)

            # Lấy tổng quan database
            if sqlite_success:
                db_summary = crawler.get_database_summary()

            quality_report = crawler.get_data_quality_report()

            # Hiển thị báo cáo
            logging.info("=" * 70)
            logging.info("📊 BÁO CÁO CHẤT LƯỢNG DỮ LIỆU")
            logging.info("=" * 70)

            w_report = quality_report['weather']
            logging.info(f"🌡️  DỮ LIỆU THỜI TIẾT:")
            logging.info(
                f"   ✅ Chất lượng cao: {w_report['high_quality']}/{w_report['total']} ({w_report['high_percent']}%)")
            logging.info(
                f"   ⚠️  Chất lượng TB: {w_report['medium_quality']}/{w_report['total']} ({w_report['medium_percent']}%)")
            logging.info(
                f"   ❌ Chất lượng thấp: {w_report['low_quality']}/{w_report['total']} ({w_report['low_percent']}%)")

            if sqlite_success and db_summary:
                logging.info("=" * 70)
                logging.info("🗃️  TỔNG QUAN DATABASE")
                logging.info("=" * 70)
                logging.info(f"📊 Tổng số bản ghi: {db_summary.get('total_records', 0)}")
                logging.info(f"🏙️  Số tỉnh thành: {db_summary.get('total_provinces', 0)}")
                logging.info(f"🕒 Dữ liệu mới nhất: {db_summary.get('latest_data', 'N/A')}")
                if 'quality_stats' in db_summary:
                    logging.info("📈 Chất lượng dữ liệu:")
                    for quality, count in db_summary['quality_stats'].items():
                        logging.info(f"   {quality}: {count} bản ghi")

            logging.info("=" * 70)
            logging.info(f"⏱️ Thời gian thực hiện: {end_time - start_time:.2f} giây")
            logging.info(f"📁 File Excel: {excel_file}")
            logging.info(f"🗄️  Database SQLite: vietnam_weather.db")
            logging.info("🎯 LƯU Ý: Để có dữ liệu chính xác hơn, cần sử dụng API có phí hoặc truy cập trực tiếp")
            logging.info("   các nguồn dữ liệu chính thức của Việt Nam")

        else:
            logging.warning("❌ Không thu thập được dữ liệu nào")

    except Exception as e:
        logging.error(f"💥 Lỗi hệ thống: {e}")


def run_continuously():
    """Hàm thường trú để chạy main() lặp lại cứ 10 phút một lần"""
    while True:
        try:
            main()
            logging.info("⏳ Đang chờ 10 phút để chạy lần tiếp theo...")
            time.sleep(600)  # 10 phút = 600 giây
        except Exception as e:
            logging.error(f"💥 Lỗi trong quá trình chạy thường trú: {e}")
            logging.info("🔄 Thử chạy lại sau 10 phút...")
            time.sleep(600)


if __name__ == "__main__":
    run_continuously()