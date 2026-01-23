import requests
import pandas as pd
import time
import json
from pathlib import Path
from datetime import datetime, timedelta
import logging
import os
import numpy as np
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sqlite3
import os


OPENWEATHER_API_KEY = os.getenv("OPENWEATHER_API_KEY", "").strip()
WEATHERAPI_KEY = os.getenv("WEATHERAPI_KEY", "").strip()
CRAWL_MODE = os.getenv("CRAWL_MODE", "continuous").lower()
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logging.info(f"OPENWEATHER_API_KEY length = {len(OPENWEATHER_API_KEY)}")
logging.info(f"WEATHERAPI_KEY length = {len(WEATHERAPI_KEY)}")
logging.info(f"CRAWL_MODE = {CRAWL_MODE}")


class SQLiteManager:
    """Qu·∫£n l√Ω k·∫øt n·ªëi v√† thao t√°c v·ªõi SQLite database"""

    def __init__(self, db_path=None):
        if db_path is None:
            db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", "vietnam_weather.db")
        self.db_path = db_path
        self.conn = None
        self.cursor = None

    def connect(self):
        """K·∫øt n·ªëi ƒë·∫øn database"""
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.cursor = self.conn.cursor()
            logging.info(f"‚úÖ ƒê√£ k·∫øt n·ªëi ƒë·∫øn SQLite database: {self.db_path}")
        except Exception as e:
            logging.error(f"‚ùå L·ªói k·∫øt n·ªëi SQLite: {e}")

    def disconnect(self):
        """ƒê√≥ng k·∫øt n·ªëi database"""
        if self.conn:
            self.conn.close()
            logging.info("‚úÖ ƒê√£ ƒë√≥ng k·∫øt n·ªëi SQLite")

    def create_tables(self):
        """T·∫°o c√°c b·∫£ng c·∫ßn thi·∫øt trong database"""
        try:
            # B·∫£ng th√¥ng tin tr·∫°m
            self.cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS weather_stations (
                    station_id TEXT PRIMARY KEY,
                    station_name TEXT NOT NULL,
                    province TEXT NOT NULL,
                    district TEXT NOT NULL,
                    type TEXT,
                    region TEXT,
                    latitude REAL,
                    longitude REAL,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """
            )

            # B·∫£ng d·ªØ li·ªáu th·ªùi ti·∫øt
            self.cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS weather_data (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    station_id TEXT,
                    station_name TEXT,
                    province TEXT,
                    district TEXT,
                    latitude REAL,
                    longitude REAL,
                    timestamp TEXT,
                    data_source TEXT,
                    data_quality TEXT,
                    data_time TEXT,

                    -- Nhi·ªát ƒë·ªô
                    temperature_current REAL,
                    temperature_max REAL,
                    temperature_min REAL,
                    temperature_avg REAL,

                    -- ƒê·ªô ·∫©m
                    humidity_current REAL,
                    humidity_max REAL,
                    humidity_min REAL,
                    humidity_avg REAL,

                    -- √Åp su·∫•t
                    pressure_current REAL,
                    pressure_max REAL,
                    pressure_min REAL,
                    pressure_avg REAL,

                    -- Gi√≥
                    wind_speed_current REAL,
                    wind_speed_max REAL,
                    wind_speed_min REAL,
                    wind_speed_avg REAL,
                    wind_direction_current REAL,
                    wind_direction_avg REAL,

                    -- M∆∞a
                    rain_current REAL,
                    rain_max REAL,
                    rain_min REAL,
                    rain_avg REAL,
                    rain_total REAL,

                    -- C√°c ch·ªâ s·ªë kh√°c
                    cloud_cover_current INTEGER,
                    cloud_cover_max INTEGER,
                    cloud_cover_min INTEGER,
                    cloud_cover_avg INTEGER,

                    visibility_current INTEGER,
                    visibility_max INTEGER,
                    visibility_min INTEGER,
                    visibility_avg INTEGER,

                    thunder_probability INTEGER,
                    error_reason TEXT,

                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

                    FOREIGN KEY (station_id) REFERENCES weather_stations (station_id)
                )
            """
            )

            # B·∫£ng ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu
            self.cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS data_quality_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    run_timestamp TEXT,
                    data_type TEXT,
                    total_records INTEGER,
                    high_quality INTEGER,
                    medium_quality INTEGER,
                    low_quality INTEGER,
                    high_percent REAL,
                    medium_percent REAL,
                    low_percent REAL,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """
            )

            self.conn.commit()
            logging.info("‚úÖ ƒê√£ t·∫°o/x√°c nh·∫≠n c√°c b·∫£ng trong database")

        except Exception as e:
            logging.error(f"‚ùå L·ªói t·∫°o b·∫£ng SQLite: {e}")

    def insert_stations(self, stations):
        """Ch√®n d·ªØ li·ªáu tr·∫°m v√†o database"""
        try:
            for station in stations:
                self.cursor.execute(
                    """
                    INSERT OR REPLACE INTO weather_stations 
                    (station_id, station_name, province, district, type, region, latitude, longitude)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                    (
                        station["station_id"],
                        station["station_name"],
                        station["province"],
                        station["district"],
                        station.get("type", ""),
                        station.get("region", ""),
                        station["latitude"],
                        station["longitude"],
                    ),
                )
            self.conn.commit()
            logging.info(f"‚úÖ ƒê√£ ch√®n {len(stations)} tr·∫°m v√†o database")
        except Exception as e:
            logging.error(f"‚ùå L·ªói ch√®n d·ªØ li·ªáu tr·∫°m: {e}")

    def convert_vietnamese_keys_to_english(self, data):
        """Chuy·ªÉn ƒë·ªïi kh√≥a ti·∫øng Vi·ªát sang ti·∫øng Anh"""
        key_mapping = {
            "M√£ tr·∫°m": "station_id",
            "T√™n tr·∫°m": "station_name",
            "T·ªânh/Th√†nh ph·ªë": "province",
            "Huy·ªán": "district",
            "Vƒ© ƒë·ªô": "latitude",
            "Kinh ƒë·ªô": "longitude",
            "D·∫•u th·ªùi gian": "timestamp",
            "Ngu·ªìn d·ªØ li·ªáu": "data_source",
            "Ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu": "data_quality",
            "Th·ªùi gian c·∫≠p nh·∫≠t": "data_time",
            "Nhi·ªát ƒë·ªô hi·ªán t·∫°i": "temperature_current",
            "Nhi·ªát ƒë·ªô t·ªëi ƒëa": "temperature_max",
            "Nhi·ªát ƒë·ªô t·ªëi thi·ªÉu": "temperature_min",
            "Nhi·ªát ƒë·ªô trung b√¨nh": "temperature_avg",
            "ƒê·ªô ·∫©m hi·ªán t·∫°i": "humidity_current",
            "ƒê·ªô ·∫©m t·ªëi ƒëa": "humidity_max",
            "ƒê·ªô ·∫©m t·ªëi thi·ªÉu": "humidity_min",
            "ƒê·ªô ·∫©m trung b√¨nh": "humidity_avg",
            "√Åp su·∫•t hi·ªán t·∫°i": "pressure_current",
            "√Åp su·∫•t t·ªëi ƒëa": "pressure_max",
            "√Åp su·∫•t t·ªëi thi·ªÉu": "pressure_min",
            "√Åp su·∫•t trung b√¨nh": "pressure_avg",
            "T·ªëc ƒë·ªô gi√≥ hi·ªán t·∫°i": "wind_speed_current",
            "T·ªëc ƒë·ªô gi√≥ t·ªëi ƒëa": "wind_speed_max",
            "T·ªëc ƒë·ªô gi√≥ t·ªëi thi·ªÉu": "wind_speed_min",
            "T·ªëc ƒë·ªô gi√≥ trung b√¨nh": "wind_speed_avg",
            "H∆∞·ªõng gi√≥ hi·ªán t·∫°i": "wind_direction_current",
            "H∆∞·ªõng gi√≥ trung b√¨nh": "wind_direction_avg",
            "L∆∞·ª£ng m∆∞a hi·ªán t·∫°i": "rain_current",
            "L∆∞·ª£ng m∆∞a t·ªëi ƒëa": "rain_max",
            "L∆∞·ª£ng m∆∞a t·ªëi thi·ªÉu": "rain_min",
            "L∆∞·ª£ng m∆∞a trung b√¨nh": "rain_avg",
            "T·ªïng l∆∞·ª£ng m∆∞a": "rain_total",
            "ƒê·ªô che ph·ªß m√¢y hi·ªán t·∫°i": "cloud_cover_current",
            "ƒê·ªô che ph·ªß m√¢y t·ªëi ƒëa": "cloud_cover_max",
            "ƒê·ªô che ph·ªß m√¢y t·ªïi thi·ªÉu": "cloud_cover_min",
            "ƒê·ªô che ph·ªß m√¢y trung b√¨nh": "cloud_cover_avg",
            "T·∫ßm nh√¨n hi·ªán t·∫°i": "visibility_current",
            "T·∫ßm nh√¨n ƒëa": "visibility_max",
            "T·∫ßm nh√¨n t·ªëi thi·ªÉu": "visibility_min",
            "T·∫ßm nh√¨n trung b√¨nh": "visibility_avg",
            "X√°c xu·∫•t s·∫•m s√©t": "thunder_probability",
            "L√Ω do l·ªói": "error_reason",
        }
        
        converted = {}
        for viet_key, eng_key in key_mapping.items():
            converted[eng_key] = data.get(viet_key, None)
        
        return converted

    def insert_weather_data(self, weather_data):
        """Ch√®n d·ªØ li·ªáu th·ªùi ti·∫øt v√†o database"""
        try:
            inserted_count = 0
            for data in weather_data:
                # Chuy·ªÉn ƒë·ªïi kh√≥a ti·∫øng Vi·ªát sang ti·∫øng Anh
                data_converted = self.convert_vietnamese_keys_to_english(data)
                
                self.cursor.execute(
                    """
                    INSERT INTO weather_data (
                        station_id, station_name, province, district, latitude, longitude,
                        timestamp, data_source, data_quality, data_time,
                        temperature_current, temperature_max, temperature_min, temperature_avg,
                        humidity_current, humidity_max, humidity_min, humidity_avg,
                        pressure_current, pressure_max, pressure_min, pressure_avg,
                        wind_speed_current, wind_speed_max, wind_speed_min, wind_speed_avg,
                        wind_direction_current, wind_direction_avg,
                        rain_current, rain_max, rain_min, rain_avg, rain_total,
                        cloud_cover_current, cloud_cover_max, cloud_cover_min, cloud_cover_avg,
                        visibility_current, visibility_max, visibility_min, visibility_avg,
                        thunder_probability, error_reason
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                    (
                        data_converted["station_id"],
                        data_converted["station_name"],
                        data_converted["province"],
                        data_converted["district"],
                        data_converted["latitude"],
                        data_converted["longitude"],
                        data_converted["timestamp"],
                        data_converted["data_source"],
                        data_converted["data_quality"],
                        data_converted["data_time"],
                        data_converted["temperature_current"],
                        data_converted["temperature_max"],
                        data_converted["temperature_min"],
                        data_converted["temperature_avg"],
                        data_converted["humidity_current"],
                        data_converted["humidity_max"],
                        data_converted["humidity_min"],
                        data_converted["humidity_avg"],
                        data_converted["pressure_current"],
                        data_converted["pressure_max"],
                        data_converted["pressure_min"],
                        data_converted["pressure_avg"],
                        data_converted["wind_speed_current"],
                        data_converted["wind_speed_max"],
                        data_converted["wind_speed_min"],
                        data_converted["wind_speed_avg"],
                        data_converted["wind_direction_current"],
                        data_converted["wind_direction_avg"],
                        data_converted["rain_current"],
                        data_converted["rain_max"],
                        data_converted["rain_min"],
                        data_converted["rain_avg"],
                        data_converted["rain_total"],
                        data_converted["cloud_cover_current"],
                        data_converted["cloud_cover_max"],
                        data_converted["cloud_cover_min"],
                        data_converted["cloud_cover_avg"],
                        data_converted["visibility_current"],
                        data_converted["visibility_max"],
                        data_converted["visibility_min"],
                        data_converted["visibility_avg"],
                        data_converted["thunder_probability"],
                        data_converted.get("error_reason", ""),
                    ),
                )
                inserted_count += 1

            self.conn.commit()
            logging.info(f"‚úÖ ƒê√£ ch√®n {inserted_count} b·∫£n ghi th·ªùi ti·∫øt v√†o database")
            return inserted_count

        except Exception as e:
            logging.error(f"‚ùå L·ªói ch√®n d·ªØ li·ªáu th·ªùi ti·∫øt: {e}")
            return 0

    def insert_quality_log(self, quality_report, run_timestamp):
        """Ch√®n log ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu"""
        try:
            weather_report = quality_report["weather"]

            self.cursor.execute(
                """
                INSERT INTO data_quality_log 
                (run_timestamp, data_type, total_records, high_quality, medium_quality, low_quality, high_percent, medium_percent, low_percent)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                (
                    run_timestamp,
                    "weather",
                    weather_report["total"],
                    weather_report["high_quality"],
                    weather_report["medium_quality"],
                    weather_report["low_quality"],
                    weather_report["high_percent"],
                    weather_report["medium_percent"],
                    weather_report["low_percent"],
                ),
            )

            self.conn.commit()
            logging.info("‚úÖ ƒê√£ ch√®n log ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu")

        except Exception as e:
            logging.error(f"‚ùå L·ªói ch√®n log ch·∫•t l∆∞·ª£ng: {e}")

    def get_recent_data(self, limit=10, province=None):
        """L·∫•y d·ªØ li·ªáu g·∫ßn ƒë√¢y t·ª´ database"""
        try:
            if province:
                self.cursor.execute(
                    """
                    SELECT * FROM weather_data 
                    WHERE province = ? 
                    ORDER BY timestamp DESC 
                    LIMIT ?
                """,
                    (province, limit),
                )
            else:
                self.cursor.execute(
                    """
                    SELECT * FROM weather_data 
                    ORDER BY timestamp DESC 
                    LIMIT ?
                """,
                    (limit,),
                )

            columns = [description[0] for description in self.cursor.description]
            results = self.cursor.fetchall()

            data = []
            for row in results:
                data.append(dict(zip(columns, row)))

            return data

        except Exception as e:
            logging.error(f"‚ùå L·ªói l·∫•y d·ªØ li·ªáu: {e}")
            return []

    def get_data_summary(self):
        """L·∫•y t·ªïng quan d·ªØ li·ªáu"""
        try:
            # T·ªïng s·ªë b·∫£n ghi
            self.cursor.execute("SELECT COUNT(*) FROM weather_data")
            total_records = self.cursor.fetchone()[0]

            # S·ªë t·ªânh th√†nh
            self.cursor.execute("SELECT COUNT(DISTINCT province) FROM weather_data")
            total_provinces = self.cursor.fetchone()[0]

            # D·ªØ li·ªáu m·ªõi nh·∫•t
            self.cursor.execute("SELECT MAX(timestamp) FROM weather_data")
            latest_data = self.cursor.fetchone()[0]

            # Ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu trung b√¨nh
            self.cursor.execute(
                """
                SELECT data_quality, COUNT(*) 
                FROM weather_data 
                GROUP BY data_quality
            """
            )
            quality_stats = dict(self.cursor.fetchall())

            return {
                "total_records": total_records,
                "total_provinces": total_provinces,
                "latest_data": latest_data,
                "quality_stats": quality_stats,
            }

        except Exception as e:
            logging.error(f"‚ùå L·ªói l·∫•y t·ªïng quan: {e}")
            return {}

FILE_PATH = Path(__file__).resolve()
APP_DIR = FILE_PATH.parents[1]
OUTPUT_DIR = "/media/voanhnhat/SDD_OUTSIDE5/PROJECT_WEATHER_FORECAST/Weather_Forcast_App/output"

class VietnamWeatherDataCrawler:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }
        )
        self.data_quality_tracking = {
            "weather": {"high_quality": 0, "medium_quality": 0, "low_quality": 0}
        }
        self.db_manager = SQLiteManager()

    def get_data_quality_assessment(self, data_source, data_type):
        """ƒê√°nh gi√° ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu d·ª±a tr√™n ngu·ªìn"""
        quality_metrics = {
            "openmeteo": {"weather": "medium"},
            "openweather": {"weather": "medium"},
            "simulated": {"weather": "low"},
            "statistical": {"weather": "medium"},
        }

        return quality_metrics.get(data_source, {}).get(data_type, "low")

    def get_vietnam_weather_data(self, lat, lon, province):
        """
        C·ªë g·∫Øng l·∫•y d·ªØ li·ªáu th·ªùi ti·∫øt t·ª´ c√°c ngu·ªìn c√≥ s·∫µn cho Vi·ªát Nam
        """
        attempts = [
            self.try_openmeteo_weather,
            self.try_weatherapi_com,
            self.try_openweathermap,
        ]

        for attempt in attempts:
            try:
                data, source = attempt(lat, lon, province)
                if data and data.get("current"):
                    quality = self.get_data_quality_assessment(source, "weather")
                    self.data_quality_tracking["weather"][f"{quality}_quality"] += 1
                    return data, source, quality
            except Exception as e:
                logging.debug(f"Attempt failed: {e}")
                continue

        # Fallback cu·ªëi c√πng
        return (
            self.generate_vietnam_statistical_weather(lat, lon, province),
            "statistical",
            "medium",
        )

    def try_openmeteo_weather(self, lat, lon, province):
        """Th·ª≠ Open-Meteo API"""
        try:
            url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current=temperature_2m,relative_humidity_2m,precipitation,pressure_msl,wind_speed_10m,wind_direction_10m&hourly=temperature_2m,relative_humidity_2m,precipitation,pressure_msl,wind_speed_10m,wind_direction_10m&timezone=auto"
            response = self.session.get(url, timeout=10)

            if response.status_code == 200:
                data = response.json()
                if data.get("current"):
                    logging.info(f"‚úì Open-Meteo data for {province}")
                    return data, "openmeteo"
        except Exception as e:
            logging.debug(f"Open-Meteo failed: {e}")

        return None, None

    def try_weatherapi_com(self, lat, lon, province):
        """Th·ª≠ WeatherAPI v·ªõi key th·∫≠t"""
        try:
            # S·ª≠ d·ª•ng WeatherAPI v·ªõi key th·∫≠t ----------------------------------- S·ª¨ D·ª§NG CHO CH·∫†Y LOCAL
            # api_key = "142f4fa048f24efdad1113219251510"
            # url = f"https://api.weatherapi.com/v1/current.json?key={api_key}&q={lat},{lon}"

            # S·ª¨ D·ª§NG CHO CHA·ª¥ NG·∫¶M B·∫∞NG DOCKER --------------------------------------------------------
            api_key = WEATHERAPI_KEY
            if not api_key:
                logging.debug("WEATHERAPI_KEY ch∆∞a ƒë∆∞·ª£c set, b·ªè qua WeatherAPI")
                return None, None
            url = f"https://api.weatherapi.com/v1/current.json?key={api_key}&q={lat},{lon}"
            response = self.session.get(url, timeout=8)

            if response.status_code == 200:
                data = response.json()
                # Convert format
                converted = self.convert_weatherapi_format(data, lat, lon)
                if converted:
                    logging.info(f"‚úì WeatherAPI data for {province}")
                    return converted, "weatherapi"

        except Exception as e:
            logging.debug(f"WeatherAPI failed: {e}")

        return None, None

    def try_openweathermap(self, lat, lon, province):
        """Th·ª≠ OpenWeatherMap API v·ªõi key th·∫≠t"""
        try:
            # S·ª≠ d·ª•ng OpenWeatherMap API v·ªõi key th·∫≠t ---------------------------- S·ª¨ D·ª§NG CHO CH·∫†Y LOCAL
            # api_key = "b79b0a6a70b8d6ce0fd907a1d893156d"
            # url = f"https://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&appid={api_key}&units=metric"

            # S·ª¨ D·ª§NG CHO CH·∫†Y NG·∫¶M TR√äN M√ÅY ------------------------------------------------------------
            api_key = OPENWEATHER_API_KEY
            if not api_key:
                logging.debug(
                    "OPENWEATHER_API_KEY ch∆∞a ƒë∆∞·ª£c set, b·ªè qua OpenWeatherMap"
                )
                return None, None
            url = (
                f"https://api.openweathermap.org/data/2.5/weather?"
                f"lat={lat}&lon={lon}&appid={api_key}&units=metric"
            )

            response = self.session.get(url, timeout=8)

            if response.status_code == 200:
                data = response.json()
                converted = self.convert_openweathermap_format(data, lat, lon)
                if converted:
                    logging.info(f"‚úì OpenWeatherMap data for {province}")
                    return converted, "openweather"
        except Exception as e:
            logging.debug(f"OpenWeatherMap failed: {e}")

        return None, None

    def convert_weatherapi_format(self, data, lat, lon):
        """Chuy·ªÉn ƒë·ªïi ƒë·ªãnh d·∫°ng WeatherAPI"""
        try:
            current = data.get("current", {})
            return {
                "current": {
                    "time": datetime.utcnow().isoformat(),
                    "temperature_2m": current.get("temp_c", 0),
                    "relative_humidity_2m": current.get("humidity", 0),
                    "pressure_msl": current.get("pressure_mb", 0),
                    "wind_speed_10m": current.get("wind_kph", 0) / 3.6,  # km/h to m/s
                    "wind_direction_10m": current.get("wind_degree", 0),
                    "precipitation": current.get("precip_mm", 0),
                },
                "hourly": {
                    "time": [datetime.utcnow().isoformat()],
                    "temperature_2m": [current.get("temp_c", 0)],
                    "relative_humidity_2m": [current.get("humidity", 0)],
                    "pressure_msl": [current.get("pressure_mb", 0)],
                    "wind_speed_10m": [current.get("wind_kph", 0) / 3.6],
                    "wind_direction_10m": [current.get("wind_degree", 0)],
                    "precipitation": [current.get("precip_mm", 0)],
                },
            }
        except Exception as e:
            logging.debug(f"WeatherAPI conversion failed: {e}")
            return None

    def convert_openweathermap_format(self, data, lat, lon):
        """Chuy·ªÉn ƒë·ªïi ƒë·ªãnh d·∫°ng OpenWeatherMap"""
        try:
            main = data.get("main", {})
            wind = data.get("wind", {})
            return {
                "current": {
                    "time": datetime.utcnow().isoformat(),
                    "temperature_2m": main.get("temp", 0),
                    "relative_humidity_2m": main.get("humidity", 0),
                    "pressure_msl": main.get("pressure", 0),
                    "wind_speed_10m": wind.get("speed", 0),
                    "wind_direction_10m": wind.get("deg", 0),
                    "precipitation": (
                        data.get("rain", {}).get("1h", 0) if data.get("rain") else 0
                    ),
                },
                "hourly": {
                    "time": [datetime.utcnow().isoformat()],
                    "temperature_2m": [main.get("temp", 0)],
                    "relative_humidity_2m": [main.get("humidity", 0)],
                    "pressure_msl": [main.get("pressure", 0)],
                    "wind_speed_10m": [wind.get("speed", 0)],
                    "wind_direction_10m": [wind.get("deg", 0)],
                    "precipitation": [
                        data.get("rain", {}).get("1h", 0) if data.get("rain") else 0
                    ],
                },
            }
        except Exception as e:
            logging.debug(f"OpenWeatherMap conversion failed: {e}")
            return None

    def generate_vietnam_statistical_weather(self, lat, lon, province):
        """
        T·∫°o d·ªØ li·ªáu th·ªëng k√™ d·ª±a tr√™n kh√≠ h·∫≠u Vi·ªát Nam
        S·ª≠ d·ª•ng d·ªØ li·ªáu l·ªãch s·ª≠ v√† ph√¢n v√πng kh√≠ h·∫≠u
        """
        current_time = datetime.now()
        current_month = current_time.month
        current_hour = current_time.hour

        # Ph√¢n v√πng kh√≠ h·∫≠u Vi·ªát Nam
        if lat > 21.0:  # B·∫Øc B·ªô
            if current_month in [12, 1, 2]:  # ƒê√¥ng
                base_temp = 18 + random.uniform(-3, 5)
                rain_prob = 0.1
                humidity_range = (70, 85)
            elif current_month in [3, 4]:  # Xu√¢n
                base_temp = 23 + random.uniform(-2, 4)
                rain_prob = 0.3
                humidity_range = (75, 90)
            else:  # H√®-Thu
                base_temp = 29 + random.uniform(-2, 4)
                rain_prob = 0.5
                humidity_range = (75, 95)
        elif 16.0 <= lat <= 21.0:  # Trung B·ªô
            base_temp = 28 + random.uniform(-2, 4)
            rain_prob = 0.4
            humidity_range = (70, 90)
        else:  # Nam B·ªô
            if current_month in [5, 6, 7, 8, 9, 10]:  # M√πa m∆∞a
                base_temp = 28 + random.uniform(-1, 3)
                rain_prob = 0.7
                humidity_range = (75, 95)
            else:  # M√πa kh√¥
                base_temp = 30 + random.uniform(-1, 3)
                rain_prob = 0.2
                humidity_range = (65, 85)

        # ƒêi·ªÅu ch·ªânh theo gi·ªù trong ng√†y
        if 0 <= current_hour <= 6:  # ƒê√™m
            temperature = base_temp - 4 + random.uniform(-1, 1)
        elif 12 <= current_hour <= 14:  # Tr∆∞a
            temperature = base_temp + 3 + random.uniform(-1, 1)
        else:
            temperature = base_temp + random.uniform(-1, 1)

        # T·∫°o d·ªØ li·ªáu hourly v·ªõi bi·∫øn ƒë·ªông th·ª±c t·∫ø
        hourly_temps = []
        hourly_humidities = []
        hourly_pressures = []
        hourly_wind_speeds = []
        hourly_wind_directions = []
        hourly_rains = []

        for hour in range(24):
            hour_temp = base_temp
            if hour <= 6:
                hour_temp -= 4
            elif 12 <= hour <= 14:
                hour_temp += 3

            hourly_temps.append(hour_temp + random.uniform(-2, 2))
            hourly_humidities.append(
                random.randint(humidity_range[0], humidity_range[1])
            )
            hourly_pressures.append(random.randint(1005, 1020))
            hourly_wind_speeds.append(round(random.uniform(1, 6), 1))
            hourly_wind_directions.append(random.randint(0, 360))
            hourly_rains.append(
                round(random.uniform(0, 8) if random.random() < rain_prob else 0, 1)
            )

        return {
            "current": {
                "time": current_time.isoformat(),
                "temperature_2m": round(temperature, 1),
                "relative_humidity_2m": random.randint(
                    humidity_range[0], humidity_range[1]
                ),
                "pressure_msl": random.randint(1005, 1020),
                "wind_speed_10m": round(random.uniform(1, 6), 1),
                "wind_direction_10m": random.randint(0, 360),
                "precipitation": round(
                    random.uniform(0, 5) if random.random() < rain_prob else 0, 1
                ),
            },
            "hourly": {
                "time": [
                    (current_time - timedelta(hours=23 - hour)).isoformat()
                    for hour in range(24)
                ],
                "temperature_2m": [round(t, 1) for t in hourly_temps],
                "relative_humidity_2m": hourly_humidities,
                "pressure_msl": hourly_pressures,
                "wind_speed_10m": hourly_wind_speeds,
                "wind_direction_10m": hourly_wind_directions,
                "precipitation": hourly_rains,
            },
        }

    def parse_weather_data(self, station_info):
        try:
            weather_data, source, quality = self.get_vietnam_weather_data(
                station_info["latitude"],
                station_info["longitude"],
                station_info["province"],
            )

            if not weather_data or not isinstance(weather_data, dict):
                logging.error(
                    f"L·ªói: weather_data r·ªóng ho·∫∑c kh√¥ng ph·∫£i dict cho tr·∫°m {station_info['station_id']}"
                )
                return self.create_fallback_weather_record(station_info, "no_data")

            current = weather_data.get("current", {})
            hourly = weather_data.get("hourly", {})

            record = self.calculate_weather_metrics(
                station_info, current, hourly, source, quality
            )
            return record

        except Exception as e:
            logging.error(
                f"L·ªói ph√¢n t√≠ch th·ªùi ti·∫øt {station_info['station_name']}: {e}"
            )
            return self.create_fallback_weather_record(station_info, "error")

    def calculate_weather_metrics(self, station_info, current, hourly, source, quality):
        """T√≠nh to√°n c√°c ch·ªâ s·ªë th·ªùi ti·∫øt v·ªõi max, min, avg cho t·ª´ng ch·ªâ s·ªë"""
        # L·∫•y ho·∫∑c t√≠nh to√°n c√°c gi√° tr·ªã
        temp_current = current.get("temperature_2m", 0)
        temp_hourly = hourly.get("temperature_2m", [temp_current] * 24)

        humidity_current = current.get("relative_humidity_2m", 0)
        humidity_hourly = hourly.get("relative_humidity_2m", [humidity_current] * 24)

        pressure_current = current.get("pressure_msl", 1013)
        pressure_hourly = hourly.get("pressure_msl", [pressure_current] * 24)

        wind_speed_current = current.get("wind_speed_10m", 0)
        wind_speed_hourly = hourly.get("wind_speed_10m", [wind_speed_current] * 24)

        wind_direction_current = current.get("wind_direction_10m", 0)
        wind_direction_hourly = hourly.get(
            "wind_direction_10m", [wind_direction_current] * 24
        )

        rain_current = current.get("precipitation", 0)
        rain_hourly = hourly.get("precipitation", [rain_current] * 24)

        # T√≠nh to√°n c√°c gi√° tr·ªã max, min, avg
        temp_max = max(temp_hourly) if temp_hourly else temp_current
        temp_min = min(temp_hourly) if temp_hourly else temp_current
        temp_avg = sum(temp_hourly) / len(temp_hourly) if temp_hourly else temp_current

        humidity_max = max(humidity_hourly) if humidity_hourly else humidity_current
        humidity_min = min(humidity_hourly) if humidity_hourly else humidity_current
        humidity_avg = (
            sum(humidity_hourly) / len(humidity_hourly)
            if humidity_hourly
            else humidity_current
        )

        pressure_max = max(pressure_hourly) if pressure_hourly else pressure_current
        pressure_min = min(pressure_hourly) if pressure_hourly else pressure_current
        pressure_avg = (
            sum(pressure_hourly) / len(pressure_hourly)
            if pressure_hourly
            else pressure_current
        )

        wind_speed_max = (
            max(wind_speed_hourly) if wind_speed_hourly else wind_speed_current
        )
        wind_speed_min = (
            min(wind_speed_hourly) if wind_speed_hourly else wind_speed_current
        )
        wind_speed_avg = (
            sum(wind_speed_hourly) / len(wind_speed_hourly)
            if wind_speed_hourly
            else wind_speed_current
        )

        wind_direction_avg = self.calculate_avg_wind_direction(wind_direction_hourly)

        rain_max = max(rain_hourly) if rain_hourly else rain_current
        rain_min = min(rain_hourly) if rain_hourly else rain_current
        rain_total = sum(rain_hourly) if rain_hourly else rain_current * 24
        rain_avg = rain_total / 24 if rain_hourly else rain_current

        return {
    "M√£ tr·∫°m": station_info["station_id"],
    "T√™n tr·∫°m": station_info["station_name"],
    "T·ªânh/Th√†nh ph·ªë": station_info["province"],
    "Huy·ªán": station_info["district"],
    "Vƒ© ƒë·ªô": station_info["latitude"],
    "Kinh ƒë·ªô": station_info["longitude"],

    "D·∫•u th·ªùi gian": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "Ngu·ªìn d·ªØ li·ªáu": source,
    "Ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu": quality,
    "Th·ªùi gian c·∫≠p nh·∫≠t": current.get("time", datetime.now().isoformat()),

    # üå°Ô∏è Nhi·ªát ƒë·ªô
    "Nhi·ªát ƒë·ªô hi·ªán t·∫°i": round(temp_current, 1),
    "Nhi·ªát ƒë·ªô t·ªëi ƒëa": round(temp_max, 1),
    "Nhi·ªát ƒë·ªô t·ªëi thi·ªÉu": round(temp_min, 1),
    "Nhi·ªát ƒë·ªô trung b√¨nh": round(temp_avg, 1),

    # üíß ƒê·ªô ·∫©m
    "ƒê·ªô ·∫©m hi·ªán t·∫°i": round(humidity_current, 1),
    "ƒê·ªô ·∫©m t·ªëi ƒëa": round(humidity_max, 1),
    "ƒê·ªô ·∫©m t·ªëi thi·ªÉu": round(humidity_min, 1),
    "ƒê·ªô ·∫©m trung b√¨nh": round(humidity_avg, 1),

    # üß≠ √Åp su·∫•t
    "√Åp su·∫•t hi·ªán t·∫°i": round(pressure_current, 1),
    "√Åp su·∫•t t·ªëi ƒëa": round(pressure_max, 1),
    "√Åp su·∫•t t·ªëi thi·ªÉu": round(pressure_min, 1),
    "√Åp su·∫•t trung b√¨nh": round(pressure_avg, 1),

    # üå¨Ô∏è Gi√≥
    "T·ªëc ƒë·ªô gi√≥ hi·ªán t·∫°i": round(wind_speed_current, 1),
    "T·ªëc ƒë·ªô gi√≥ t·ªëi ƒëa": round(wind_speed_max, 1),
    "T·ªëc ƒë·ªô gi√≥ t·ªëi thi·ªÉu": round(wind_speed_min, 1),
    "T·ªëc ƒë·ªô gi√≥ trung b√¨nh": round(wind_speed_avg, 1),
    "H∆∞·ªõng gi√≥ hi·ªán t·∫°i": round(wind_direction_current, 1),
    "H∆∞·ªõng gi√≥ trung b√¨nh": round(wind_direction_avg, 1),

    # üåßÔ∏è M∆∞a
    "L∆∞·ª£ng m∆∞a hi·ªán t·∫°i": round(rain_current, 1),
    "L∆∞·ª£ng m∆∞a t·ªëi ƒëa": round(rain_max, 1),
    "L∆∞·ª£ng m∆∞a t·ªëi thi·ªÉu": round(rain_min, 1),
    "L∆∞·ª£ng m∆∞a trung b√¨nh": round(rain_avg, 1),
    "T·ªïng l∆∞·ª£ng m∆∞a": round(rain_total, 1),

    # ‚òÅÔ∏è C√°c ch·ªâ s·ªë ∆∞·ªõc t√≠nh
    "ƒê·ªô che ph·ªß m√¢y hi·ªán t·∫°i": random.randint(20, 80),
    "ƒê·ªô che ph·ªß m√¢y t·ªëi ƒëa": random.randint(60, 95),
    "ƒê·ªô che ph·ªß m√¢y t·ªïi thi·ªÉu": random.randint(10, 40),
    "ƒê·ªô che ph·ªß m√¢y trung b√¨nh": random.randint(30, 70),

    "T·∫ßm nh√¨n hi·ªán t·∫°i": random.randint(5, 15),
    "T·∫ßm nh√¨n ƒëa": random.randint(10, 20),
    "T·∫ßm nh√¨n t·ªëi thi·ªÉu": random.randint(2, 8),
    "T·∫ßm nh√¨n trung b√¨nh": random.randint(6, 12),

    "X√°c xu·∫•t s·∫•m s√©t": random.randint(0, 30),
}


    def calculate_avg_wind_direction(self, directions):
        """T√≠nh h∆∞·ªõng gi√≥ trung b√¨nh"""
        if not directions:
            return 0
        try:
            rads = [d * np.pi / 180 for d in directions]
            x_sum = sum([np.sin(r) for r in rads])
            y_sum = sum([np.cos(r) for r in rads])
            avg_rad = np.arctan2(x_sum, y_sum)
            return (avg_rad * 180 / np.pi) % 360
        except:
            return sum(directions) / len(directions)

    def create_fallback_weather_record(self, station_info, reason):
        """T·∫°o b·∫£n ghi th·ªùi ti·∫øt fallback v·ªõi ƒë·∫ßy ƒë·ªß max, min, avg"""
        return {
            "M√£ tr·∫°m": station_info["station_id"],
            "T√™n tr·∫°m": station_info["station_name"],
            "T·ªânh/Th√†nh ph·ªë": station_info["province"],
            "Huy·ªán": station_info["district"],
            "Vƒ© ƒë·ªô": station_info["latitude"],
            "Kinh ƒë·ªô": station_info["longitude"],
            "D·∫•u th·ªùi gian": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Ngu·ªìn d·ªØ li·ªáu": "fallback",
            "Ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu": "low",
            "Th·ªùi gian c·∫≠p nh·∫≠t": datetime.now().isoformat(),
            "L√Ω do l·ªói": reason,
            # Nhi·ªát ƒë·ªô
            "Nhi·ªát ƒë·ªô hi·ªán t·∫°i": 25.0,
            "Nhi·ªát ƒë·ªô t·ªëi ƒëa": 30.0,
            "Nhi·ªát ƒë·ªô t·ªëi thi·ªÉu": 20.0,
            "Nhi·ªát ƒë·ªô trung b√¨nh": 25.0,
            # ƒê·ªô ·∫©m
            "ƒê·ªô ·∫©m hi·ªán t·∫°i": 75.0,
            "ƒê·ªô ·∫©m t·ªëi ƒëa": 85.0,
            "ƒê·ªô ·∫©m t·ªëi thi·ªÉu": 65.0,
            "ƒê·ªô ·∫©m trung b√¨nh": 75.0,
            # √Åp su·∫•t
            "√Åp su·∫•t hi·ªán t·∫°i": 1013.0,
            "√Åp su·∫•t t·ªëi ƒëa": 1020.0,
            "√Åp su·∫•t t·ªëi thi·ªÉu": 1005.0,
            "√Åp su·∫•t·∫°irung b√¨nh": 1013.0,
            # Gi√≥
            "T·ªëc ƒë·ªô gi√≥ hi·ªán t·∫°i": 3.0,
            "T·ªëc ƒë·ªô gi√≥ t·ªëi ƒëa": 6.0,
            "T·ªëc ƒë·ªô gi√≥ t·ªëi thi·ªÉu": 1.0,
            "T·ªëc ƒë·ªô gi√≥ trung b√¨nh": 3.5,
            "H∆∞·ªõng gi√≥ hi·ªán t·∫°i": 180,
            "H∆∞·ªõng gi√≥ trung b√¨nh": 180,
            # M∆∞a
            "L∆∞·ª£ng m∆∞a hi·ªán t·∫°i": 0,
            "L∆∞·ª£ng m∆∞a t·ªëi ƒëa": 0,
            "L∆∞·ª£ng m∆∞a t·ªëi thi·ªÉu": 0,
            "L∆∞·ª£ng m∆∞a trung b√¨nh": 0,
            "T·ªïng l∆∞·ª£ng m∆∞a": 0,
            # C√°c ch·ªâ s·ªë ∆∞·ªõc t√≠nh
            "ƒê·ªô che ph·ªß m√¢y hi·ªán t·∫°i": 50,
            "ƒê·ªô che ph·ªß m√¢y t·ªëi ƒëa": 80,
            "ƒê·ªô che ph·ªß m√¢y t·ªïi thi·ªÉu": 20,
            "ƒê·ªô che ph·ªß m√¢y trung b√¨nh": 50,
            "T·∫ßm nh√¨n hi·ªán t·∫°i": 10,
            "T·∫ßm nh√¨n ƒëa": 15,
            "T·∫ßm nh√¨n t·ªëi thi·ªÉu": 5,
            "T·∫ßm nh√¨n trung b√¨nh": 10,
            "X√°c xu·∫•t s·∫•m s√©t": 5,
        }

    def crawl_all_locations(self, locations, delay=2.0):
        """Crawl d·ªØ li·ªáu cho t·∫•t c·∫£ ƒë·ªãa ƒëi·ªÉm"""
        all_weather_data = []

        total_locations = len(locations)

        logging.info(
            f"üîÑ B·∫Øt ƒë·∫ßu thu th·∫≠p d·ªØ li·ªáu cho {total_locations} ƒë·ªãa ƒëi·ªÉm t·∫°i Vi·ªát Nam"
        )
        logging.info("üìä S·ª≠ d·ª•ng ƒëa ngu·ªìn d·ªØ li·ªáu v·ªõi ƒë√°nh gi√° ch·∫•t l∆∞·ª£ng")

        for i, location in enumerate(locations):
            logging.info(
                f"üìç ƒêang x·ª≠ l√Ω {i + 1}/{total_locations}: {location['station_name']}"
            )

            try:
                # Thu th·∫≠p d·ªØ li·ªáu th·ªùi ti·∫øt
                weather_data = self.parse_weather_data(location)
                if weather_data:
                    all_weather_data.append(weather_data)
                    logging.info(
                        f"  ‚úÖ Th·ªùi ti·∫øt: {weather_data['Ngu·ªìn d·ªØ li·ªáu']} ({weather_data['Ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu']})"
                    )

                time.sleep(delay)

            except Exception as e:
                logging.error(f"‚ùå L·ªói x·ª≠ l√Ω {location['station_name']}: {e}")
                continue

        return all_weather_data

    def get_data_quality_report(self):
        """T·∫°o b√°o c√°o ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu"""
        weather_total = sum(self.data_quality_tracking["weather"].values())

        report = {
            "weather": {
                "total": weather_total,
                "high_quality": self.data_quality_tracking["weather"]["high_quality"],
                "medium_quality": self.data_quality_tracking["weather"][
                    "medium_quality"
                ],
                "low_quality": self.data_quality_tracking["weather"]["low_quality"],
                "high_percent": (
                    round(
                        (
                            self.data_quality_tracking["weather"]["high_quality"]
                            / weather_total
                            * 100
                        ),
                        2,
                    )
                    if weather_total > 0
                    else 0
                ),
                "medium_percent": (
                    round(
                        (
                            self.data_quality_tracking["weather"]["medium_quality"]
                            / weather_total
                            * 100
                        ),
                        2,
                    )
                    if weather_total > 0
                    else 0
                ),
                "low_percent": (
                    round(
                        (
                            self.data_quality_tracking["weather"]["low_quality"]
                            / weather_total
                            * 100
                        ),
                        2,
                    )
                    if weather_total > 0
                    else 0
                ),
            }
        }

        return report

    def save_to_excel(self, weather_data, output_dir=None):        
        if output_dir is None:
            output_dir = OUTPUT_DIR

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_dir / f"Bao_cao_{timestamp}.xlsx"

        wb = Workbook()

        # Sheet th·ªùi ti·∫øt
        if weather_data:
            weather_df = pd.DataFrame(weather_data)
            ws_weather = wb.active
            ws_weather.title = "Weather Data"

            for r in dataframe_to_rows(weather_df, index=False, header=True):
                ws_weather.append(r)

            # ƒê·ªãnh d·∫°ng header
            for cell in ws_weather[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )

            # T·ª± ƒë·ªông ƒëi·ªÅu ch·ªânh ƒë·ªô r·ªông c·ªôt
            for column in ws_weather.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_weather.column_dimensions[column_letter].width = adjusted_width
        else:
            ws_weather = wb.active
            ws_weather.title = "Weather Data"
            ws_weather["A1"] = "Kh√¥ng c√≥ d·ªØ li·ªáu th·ªùi ti·∫øt"

        # Sheet ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu
        ws_quality = wb.create_sheet("Data Quality Report")
        quality_report = self.get_data_quality_report()

        ws_quality["A1"] = "B√ÅO C√ÅO CH·∫§T L∆Ø·ª¢NG D·ªÆ LI·ªÜU"
        ws_quality["A1"].font = Font(bold=True, size=14, color="366092")

        # Th·ªëng k√™ ch·∫•t l∆∞·ª£ng
        stats_data = [
            [
                "Lo·∫°i d·ªØ li·ªáu",
                "T·ªïng s·ªë",
                "Ch·∫•t l∆∞·ª£ng cao",
                "Ch·∫•t l∆∞·ª£ng trung b√¨nh",
                "Ch·∫•t l∆∞·ª£ng th·∫•p",
                "T·ª∑ l·ªá cao (%)",
            ],
            [
                "Th·ªùi ti·∫øt",
                quality_report["weather"]["total"],
                quality_report["weather"]["high_quality"],
                quality_report["weather"]["medium_quality"],
                quality_report["weather"]["low_quality"],
                quality_report["weather"]["high_percent"],
            ],
        ]

        for row_idx, row_data in enumerate(stats_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_quality.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="70AD47", end_color="70AD47", fill_type="solid"
                    )

        # ƒêi·ªÅu ch·ªânh ƒë·ªô r·ªông c·ªôt
        for column in ws_quality.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_quality.column_dimensions[column_letter].width = adjusted_width

        wb.save(excel_file)
        logging.info(f"üíæ ƒê√£ l∆∞u d·ªØ li·ªáu v√†o: {excel_file}")

        return excel_file

    def save_to_sqlite(self, weather_data, locations):
        """L∆∞u d·ªØ li·ªáu v√†o SQLite database"""
        try:
            # K·∫øt n·ªëi database
            self.db_manager.connect()

            # T·∫°o c√°c b·∫£ng
            self.db_manager.create_tables()

            # Ch√®n d·ªØ li·ªáu tr·∫°m
            self.db_manager.insert_stations(locations)

            # Ch√®n d·ªØ li·ªáu th·ªùi ti·∫øt
            inserted_count = self.db_manager.insert_weather_data(weather_data)

            # Ch√®n log ch·∫•t l∆∞·ª£ng
            quality_report = self.get_data_quality_report()
            run_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.db_manager.insert_quality_log(quality_report, run_timestamp)

            # ƒê√≥ng k·∫øt n·ªëi
            self.db_manager.disconnect()

            logging.info(f"üíæ ƒê√£ l∆∞u {inserted_count} b·∫£n ghi v√†o SQLite database")
            return True

        except Exception as e:
            logging.error(f"‚ùå L·ªói l∆∞u d·ªØ li·ªáu v√†o SQLite: {e}")
            return False

    def get_database_summary(self):
        """L·∫•y t·ªïng quan d·ªØ li·ªáu t·ª´ database"""
        try:
            self.db_manager.connect()
            summary = self.db_manager.get_data_summary()
            self.db_manager.disconnect()
            return summary
        except Exception as e:
            logging.error(f"‚ùå L·ªói l·∫•y t·ªïng quan database: {e}")
            return {}


# DANH S√ÅCH ƒê·ªäA ƒêI·ªÇM VI·ªÜT NAM
vietnam_locations = [
    # H√† N·ªôi - 30 qu·∫≠n huy·ªán
    {
        "station_id": "HN_BD",
        "station_name": "Qu·∫≠n Ba ƒê√¨nh - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ba ƒê√¨nh",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0338,
        "longitude": 105.8142,
    },
    {
        "station_id": "HN_HK",
        "station_name": "Qu·∫≠n Ho√†n Ki·∫øm - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ho√†n Ki·∫øm",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0285,
        "longitude": 105.8542,
    },
    {
        "station_id": "HN_HBT",
        "station_name": "Qu·∫≠n Hai B√† Tr∆∞ng - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Hai B√† Tr∆∞ng",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0091,
        "longitude": 105.8606,
    },
    {
        "station_id": "HN_DD",
        "station_name": "Qu·∫≠n ƒê·ªëng ƒêa - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "ƒê·ªëng ƒêa",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0190,
        "longitude": 105.8315,
    },
    {
        "station_id": "HN_CG",
        "station_name": "Qu·∫≠n C·∫ßu Gi·∫•y - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "C·∫ßu Gi·∫•y",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0301,
        "longitude": 105.8022,
    },
    {
        "station_id": "HN_TX",
        "station_name": "Qu·∫≠n Thanh Xu√¢n - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Thanh Xu√¢n",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0010,
        "longitude": 105.8093,
    },
    {
        "station_id": "HN_HM",
        "station_name": "Qu·∫≠n Ho√†ng Mai - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ho√†ng Mai",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.9837,
        "longitude": 105.8636,
    },
    {
        "station_id": "HN_LB",
        "station_name": "Qu·∫≠n Long Bi√™n - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Long Bi√™n",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0540,
        "longitude": 105.8959,
    },
    {
        "station_id": "HN_NTL",
        "station_name": "Qu·∫≠n Nam T·ª´ Li√™m - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Nam T·ª´ Li√™m",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0113,
        "longitude": 105.7547,
    },
    {
        "station_id": "HN_BTL",
        "station_name": "Qu·∫≠n B·∫Øc T·ª´ Li√™m - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "B·∫Øc T·ª´ Li√™m",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0772,
        "longitude": 105.7730,
    },
    {
        "station_id": "HN_HD",
        "station_name": "Qu·∫≠n H√† ƒê√¥ng - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "H√† ƒê√¥ng",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.9714,
        "longitude": 105.7788,
    },
    {
        "station_id": "HN_ST",
        "station_name": "Th·ªã x√£ S∆°n T√¢y - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "S∆°n T√¢y",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.1376,
        "longitude": 105.5070,
    },
    {
        "station_id": "HN_DP",
        "station_name": "Huy·ªán ƒêan Ph∆∞·ª£ng - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "ƒêan Ph∆∞·ª£ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.1394,
        "longitude": 105.6736,
    },
    {
        "station_id": "HN_Hƒê",
        "station_name": "Huy·ªán Ho√†i ƒê·ª©c - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ho√†i ƒê·ª©c",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0392,
        "longitude": 105.7108,
    },
    {
        "station_id": "HN_QO",
        "station_name": "Huy·ªán Qu·ªëc Oai - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Qu·ªëc Oai",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.9956,
        "longitude": 105.6119,
    },
    {
        "station_id": "HN_TT",
        "station_name": "Huy·ªán Th·∫°ch Th·∫•t - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Th·∫°ch Th·∫•t",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0278,
        "longitude": 105.5564,
    },
    {
        "station_id": "HN_CM",
        "station_name": "Huy·ªán Ch∆∞∆°ng M·ªπ - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ch∆∞∆°ng M·ªπ",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8964,
        "longitude": 105.7039,
    },
    {
        "station_id": "HN_TO",
        "station_name": "Huy·ªán Thanh Oai - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Thanh Oai",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8556,
        "longitude": 105.7722,
    },
    {
        "station_id": "HN_TT2",
        "station_name": "Huy·ªán Th∆∞·ªùng T√≠n - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Th∆∞·ªùng T√≠n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8700,
        "longitude": 105.8583,
    },
    {
        "station_id": "HN_PX",
        "station_name": "Huy·ªán Ph√∫ Xuy√™n - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ph√∫ Xuy√™n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7392,
        "longitude": 105.9139,
    },
    {
        "station_id": "HN_UH",
        "station_name": "Huy·ªán ·ª®ng H√≤a - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "·ª®ng H√≤a",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7292,
        "longitude": 105.8236,
    },
    {
        "station_id": "HN_MD",
        "station_name": "Huy·ªán M·ªπ ƒê·ª©c - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "M·ªπ ƒê·ª©c",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.6472,
        "longitude": 105.7194,
    },
    {
        "station_id": "HN_SL",
        "station_name": "Huy·ªán S√≥c S∆°n - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "S√≥c S∆°n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.2578,
        "longitude": 105.8489,
    },
    {
        "station_id": "HN_DH",
        "station_name": "Huy·ªán ƒê√¥ng Anh - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "ƒê√¥ng Anh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.1467,
        "longitude": 105.8464,
    },
    {
        "station_id": "HN_GL",
        "station_name": "Huy·ªán Gia L√¢m - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Gia L√¢m",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0408,
        "longitude": 105.9500,
    },
    {
        "station_id": "HN_MT",
        "station_name": "Huy·ªán M√™ Linh - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "M√™ Linh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.1753,
        "longitude": 105.7219,
    },
    {
        "station_id": "HN_PT",
        "station_name": "Huy·ªán Ph√∫c Th·ªç - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ph√∫c Th·ªç",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.1031,
        "longitude": 105.5514,
    },
    {
        "station_id": "HN_BG",
        "station_name": "Huy·ªán Ba V√¨ - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ba V√¨",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.1992,
        "longitude": 105.4239,
    },
    {
        "station_id": "HN_ML",
        "station_name": "Huy·ªán M·ªπ ƒê·ª©c - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "M·ªπ ƒê·ª©c",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.6472,
        "longitude": 105.7194,
    },
    {
        "station_id": "HN_PH",
        "station_name": "Huy·ªán Ph√∫ Xuy√™n - H√† N·ªôi",
        "province": "H√† N·ªôi",
        "district": "Ph√∫ Xuy√™n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7392,
        "longitude": 105.9139,
    },
    # TP H·ªì Ch√≠ Minh - 22 qu·∫≠n huy·ªán (ƒë√£ b·ªï sung ƒë·ªß)
    {
        "station_id": "HCM_1",
        "station_name": "Qu·∫≠n 1 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 1",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7757,
        "longitude": 106.7004,
    },
    {
        "station_id": "HCM_2",
        "station_name": "Qu·∫≠n 2 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 2",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7872,
        "longitude": 106.7498,
    },
    {
        "station_id": "HCM_3",
        "station_name": "Qu·∫≠n 3 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 3",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7823,
        "longitude": 106.6848,
    },
    {
        "station_id": "HCM_4",
        "station_name": "Qu·∫≠n 4 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 4",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7642,
        "longitude": 106.7053,
    },
    {
        "station_id": "HCM_5",
        "station_name": "Qu·∫≠n 5 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 5",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7540,
        "longitude": 106.6690,
    },
    {
        "station_id": "HCM_6",
        "station_name": "Qu·∫≠n 6 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 6",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7464,
        "longitude": 106.6492,
    },
    {
        "station_id": "HCM_7",
        "station_name": "Qu·∫≠n 7 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 7",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7329,
        "longitude": 106.7269,
    },
    {
        "station_id": "HCM_8",
        "station_name": "Qu·∫≠n 8 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 8",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7241,
        "longitude": 106.6286,
    },
    {
        "station_id": "HCM_9",
        "station_name": "Qu·∫≠n 9 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 9",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.8428,
        "longitude": 106.8287,
    },
    {
        "station_id": "HCM_10",
        "station_name": "Qu·∫≠n 10 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 10",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7679,
        "longitude": 106.6668,
    },
    {
        "station_id": "HCM_11",
        "station_name": "Qu·∫≠n 11 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 11",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7639,
        "longitude": 106.6474,
    },
    {
        "station_id": "HCM_12",
        "station_name": "Qu·∫≠n 12 - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Qu·∫≠n 12",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.8642,
        "longitude": 106.6543,
    },
    {
        "station_id": "HCM_GB",
        "station_name": "Qu·∫≠n G√≤ V·∫•p - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "G√≤ V·∫•p",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.8383,
        "longitude": 106.6653,
    },
    {
        "station_id": "HCM_TB",
        "station_name": "Qu·∫≠n T√¢n B√¨nh - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "T√¢n B√¨nh",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.8014,
        "longitude": 106.6526,
    },
    {
        "station_id": "HCM_TP",
        "station_name": "Qu·∫≠n T√¢n Ph√∫ - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "T√¢n Ph√∫",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7902,
        "longitude": 106.6289,
    },
    {
        "station_id": "HCM_BT",
        "station_name": "Qu·∫≠n B√¨nh T√¢n - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "B√¨nh T√¢n",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7654,
        "longitude": 106.6032,
    },
    {
        "station_id": "HCM_BT2",
        "station_name": "Qu·∫≠n B√¨nh Th·∫°nh - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "B√¨nh Th·∫°nh",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.8010,
        "longitude": 106.6959,
    },
    {
        "station_id": "HCM_PN",
        "station_name": "Qu·∫≠n Ph√∫ Nhu·∫≠n - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Ph√∫ Nhu·∫≠n",
        "type": "Qu·∫≠n",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7992,
        "longitude": 106.6753,
    },
    {
        "station_id": "HCM_TD",
        "station_name": "Th√†nh ph·ªë Th·ªß ƒê·ª©c - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Th·ªß ƒê·ª©c",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.8494,
        "longitude": 106.7717,
    },
    {
        "station_id": "HCM_CC",
        "station_name": "Huy·ªán C·ªß Chi - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "C·ªß Chi",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9733,
        "longitude": 106.4939,
    },
    {
        "station_id": "HCM_HM",
        "station_name": "Huy·ªán H√≥c M√¥n - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "H√≥c M√¥n",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.8894,
        "longitude": 106.5953,
    },
    {
        "station_id": "HCM_BC",
        "station_name": "Huy·ªán B√¨nh Ch√°nh - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "B√¨nh Ch√°nh",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7269,
        "longitude": 106.5672,
    },
    {
        "station_id": "HCM_NB",
        "station_name": "Huy·ªán Nh√† B√® - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "Nh√† B√®",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.6956,
        "longitude": 106.7350,
    },
    {
        "station_id": "HCM_CG",
        "station_name": "Huy·ªán C·∫ßn Gi·ªù - TP H·ªì Ch√≠ Minh",
        "province": "TP H·ªì Ch√≠ Minh",
        "district": "C·∫ßn Gi·ªù",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.4111,
        "longitude": 106.9547,
    },
    # H·∫£i Ph√≤ng
    {
        "station_id": "HP_LC",
        "station_name": "Qu·∫≠n L√™ Ch√¢n - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "L√™ Ch√¢n",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8516,
        "longitude": 106.6822,
    },
    {
        "station_id": "HP_NQ",
        "station_name": "Qu·∫≠n Ng√¥ Quy·ªÅn - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "Ng√¥ Quy·ªÅn",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8580,
        "longitude": 106.7053,
    },
    {
        "station_id": "HP_HA",
        "station_name": "Qu·∫≠n H·∫£i An - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "H·∫£i An",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8378,
        "longitude": 106.7377,
    },
    {
        "station_id": "HP_KS",
        "station_name": "Qu·∫≠n Ki·∫øn An - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "Ki·∫øn An",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8075,
        "longitude": 106.6264,
    },
    {
        "station_id": "HP_DS",
        "station_name": "Qu·∫≠n D∆∞∆°ng Kinh - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "D∆∞∆°ng Kinh",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8025,
        "longitude": 106.6669,
    },
    {
        "station_id": "HP_DK",
        "station_name": "Qu·∫≠n ƒê·ªì S∆°n - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "ƒê·ªì S∆°n",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7278,
        "longitude": 106.7733,
    },
    {
        "station_id": "HP_TL",
        "station_name": "Huy·ªán Th·ªßy Nguy√™n - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "Th·ªßy Nguy√™n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.9583,
        "longitude": 106.6667,
    },
    {
        "station_id": "HP_AC",
        "station_name": "Huy·ªán An D∆∞∆°ng - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "An D∆∞∆°ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8667,
        "longitude": 106.6167,
    },
    {
        "station_id": "HP_AL",
        "station_name": "Huy·ªán An L√£o - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "An L√£o",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.8167,
        "longitude": 106.5500,
    },
    {
        "station_id": "HP_KL",
        "station_name": "Huy·ªán Ki·∫øn Th·ª•y - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "Ki·∫øn Th·ª•y",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7500,
        "longitude": 106.6667,
    },
    {
        "station_id": "HP_TH",
        "station_name": "Huy·ªán Ti√™n L√£ng - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "Ti√™n L√£ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7333,
        "longitude": 106.5500,
    },
    {
        "station_id": "HP_VB",
        "station_name": "Huy·ªán Vƒ©nh B·∫£o - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "Vƒ©nh B·∫£o",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7000,
        "longitude": 106.4667,
    },
    {
        "station_id": "HP_CH",
        "station_name": "Huy·ªán C√°t H·∫£i - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "C√°t H·∫£i",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7264,
        "longitude": 107.0489,
    },
    {
        "station_id": "HP_BH",
        "station_name": "Huy·ªán B·∫°ch Long Vƒ© - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "B·∫°ch Long Vƒ©",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.1333,
        "longitude": 107.7333,
    },
    # ƒê√† N·∫µng
    {
        "station_id": "DN_HC",
        "station_name": "Qu·∫≠n H·∫£i Ch√¢u - ƒê√† N·∫µng",
        "province": "ƒê√† N·∫µng",
        "district": "H·∫£i Ch√¢u",
        "type": "Qu·∫≠n",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.0592,
        "longitude": 108.2208,
    },
    {
        "station_id": "DN_ST",
        "station_name": "Qu·∫≠n S∆°n Tr√† - ƒê√† N·∫µng",
        "province": "ƒê√† N·∫µng",
        "district": "S∆°n Tr√†",
        "type": "Qu·∫≠n",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.1060,
        "longitude": 108.2493,
    },
    {
        "station_id": "DN_NH",
        "station_name": "Qu·∫≠n Ng≈© H√†nh S∆°n - ƒê√† N·∫µng",
        "province": "ƒê√† N·∫µng",
        "district": "Ng≈© H√†nh S∆°n",
        "type": "Qu·∫≠n",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.0159,
        "longitude": 108.2579,
    },
    {
        "station_id": "DN_LC",
        "station_name": "Qu·∫≠n Li√™n Chi·ªÉu - ƒê√† N·∫µng",
        "province": "ƒê√† N·∫µng",
        "district": "Li√™n Chi·ªÉu",
        "type": "Qu·∫≠n",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.0717,
        "longitude": 108.1503,
    },
    {
        "station_id": "DN_TS",
        "station_name": "Qu·∫≠n Thanh Kh√™ - ƒê√† N·∫µng",
        "province": "ƒê√† N·∫µng",
        "district": "Thanh Kh√™",
        "type": "Qu·∫≠n",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.0647,
        "longitude": 108.1911,
    },
    {
        "station_id": "DN_CL",
        "station_name": "Qu·∫≠n C·∫©m L·ªá - ƒê√† N·∫µng",
        "province": "ƒê√† N·∫µng",
        "district": "C·∫©m L·ªá",
        "type": "Qu·∫≠n",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.0156,
        "longitude": 108.2028,
    },
    {
        "station_id": "DN_HV",
        "station_name": "Huy·ªán H√≤a Vang - ƒê√† N·∫µng",
        "province": "ƒê√† N·∫µng",
        "district": "H√≤a Vang",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.0333,
        "longitude": 108.0833,
    },
    {
        "station_id": "DN_HH",
        "station_name": "Huy·ªán Ho√†ng Sa - ƒê√† N·∫µng",
        "province": "ƒê√† N·∫µng",
        "district": "Ho√†ng Sa",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.5322,
        "longitude": 111.6156,
    },
    # C·∫ßn Th∆°
    {
        "station_id": "CT_NK",
        "station_name": "Qu·∫≠n Ninh Ki·ªÅu - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "Ninh Ki·ªÅu",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0452,
        "longitude": 105.7469,
    },
    {
        "station_id": "CT_BT",
        "station_name": "Qu·∫≠n B√¨nh Th·ªßy - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "B√¨nh Th·ªßy",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0816,
        "longitude": 105.7378,
    },
    {
        "station_id": "CT_CR",
        "station_name": "Qu·∫≠n C√°i RƒÉng - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "C√°i RƒÉng",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0000,
        "longitude": 105.7667,
    },
    {
        "station_id": "CT_OT",
        "station_name": "Qu·∫≠n √î M√¥n - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "√î M√¥n",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.1167,
        "longitude": 105.6333,
    },
    {
        "station_id": "CT_TL",
        "station_name": "Qu·∫≠n Th·ªët N·ªët - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "Th·ªët N·ªët",
        "type": "Qu·∫≠n",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2667,
        "longitude": 105.5333,
    },
    {
        "station_id": "CT_VT",
        "station_name": "Huy·ªán Vƒ©nh Th·∫°nh - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "Vƒ©nh Th·∫°nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2167,
        "longitude": 105.4000,
    },
    {
        "station_id": "CT_CC",
        "station_name": "Huy·ªán C·ªù ƒê·ªè - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "C·ªù ƒê·ªè",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.1000,
        "longitude": 105.4333,
    },
    {
        "station_id": "CT_PT",
        "station_name": "Huy·ªán Phong ƒêi·ªÅn - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "Phong ƒêi·ªÅn",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2500,
        "longitude": 105.6667,
    },
    {
        "station_id": "CT_TN",
        "station_name": "Huy·ªán Th·ªõi Lai - C·∫ßn Th∆°",
        "province": "C·∫ßn Th∆°",
        "district": "Th·ªõi Lai",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0667,
        "longitude": 105.5500,
    },
    # Qu·∫£ng Ninh - ƒê√É S·ª¨A V√ôNG MI·ªÄN
    {
        "station_id": "QN_HL",
        "station_name": "Th√†nh ph·ªë H·∫° Long - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "H·∫° Long",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 20.9582,
        "longitude": 107.0758,
    },
    {
        "station_id": "QN_CP",
        "station_name": "Th√†nh ph·ªë C·∫©m Ph·∫£ - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "C·∫©m Ph·∫£",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.0167,
        "longitude": 107.3167,
    },
    {
        "station_id": "QN_UB",
        "station_name": "Th√†nh ph·ªë U√¥ng B√≠ - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "U√¥ng B√≠",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.0361,
        "longitude": 106.7633,
    },
    {
        "station_id": "QN_MC",
        "station_name": "Th√†nh ph·ªë M√≥ng C√°i - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "M√≥ng C√°i",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.5247,
        "longitude": 107.9664,
    },
    {
        "station_id": "QN_DK",
        "station_name": "Th·ªã x√£ ƒê√¥ng Tri·ªÅu - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "ƒê√¥ng Tri·ªÅu",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.0833,
        "longitude": 106.5000,
    },
    {
        "station_id": "QN_QY",
        "station_name": "Th·ªã x√£ Qu·∫£ng Y√™n - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "Qu·∫£ng Y√™n",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 20.9333,
        "longitude": 106.8167,
    },
    {
        "station_id": "QN_BG",
        "station_name": "Huy·ªán Ba Ch·∫Ω - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "Ba Ch·∫Ω",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.2833,
        "longitude": 107.1833,
    },
    {
        "station_id": "QN_BQ",
        "station_name": "Huy·ªán B√¨nh Li√™u - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "B√¨nh Li√™u",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.5333,
        "longitude": 107.4333,
    },
    {
        "station_id": "QN_CT",
        "station_name": "Huy·ªán C√¥ T√¥ - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "C√¥ T√¥",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 20.9681,
        "longitude": 107.7639,
    },
    {
        "station_id": "QN_DH",
        "station_name": "Huy·ªán ƒê·∫ßm H√† - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "ƒê·∫ßm H√†",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.3667,
        "longitude": 107.6000,
    },
    {
        "station_id": "QN_HD",
        "station_name": "Huy·ªán H·∫£i H√† - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "H·∫£i H√†",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.4333,
        "longitude": 107.7167,
    },
    {
        "station_id": "QN_TY",
        "station_name": "Huy·ªán Ti√™n Y√™n - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "Ti√™n Y√™n",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.3833,
        "longitude": 107.3833,
    },
    {
        "station_id": "QN_VD",
        "station_name": "Huy·ªán V√¢n ƒê·ªìn - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "V√¢n ƒê·ªìn",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 20.9551,
        "longitude": 107.4764,
    },
    # Th√°i B√¨nh
    {
        "station_id": "TB_TP",
        "station_name": "Th√†nh ph·ªë Th√°i B√¨nh - Th√°i B√¨nh",
        "province": "Th√°i B√¨nh",
        "district": "Th√°i B√¨nh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.4461,
        "longitude": 106.3366,
    },
    {
        "station_id": "TB_DH",
        "station_name": "Huy·ªán ƒê√¥ng H∆∞ng - Th√°i B√¨nh",
        "province": "Th√°i B√¨nh",
        "district": "ƒê√¥ng H∆∞ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.5500,
        "longitude": 106.3500,
    },
    {
        "station_id": "TB_HH",
        "station_name": "Huy·ªán H∆∞ng H√† - Th√°i B√¨nh",
        "province": "Th√°i B√¨nh",
        "district": "H∆∞ng H√†",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.6000,
        "longitude": 106.2167,
    },
    {
        "station_id": "TB_KL",
        "station_name": "Huy·ªán Ki·∫øn X∆∞∆°ng - Th√°i B√¨nh",
        "province": "Th√°i B√¨nh",
        "district": "Ki·∫øn X∆∞∆°ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.4000,
        "longitude": 106.4167,
    },
    {
        "station_id": "TB_QH",
        "station_name": "Huy·ªán Qu·ª≥nh Ph·ª• - Th√°i B√¨nh",
        "province": "Th√°i B√¨nh",
        "district": "Qu·ª≥nh Ph·ª•",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.6500,
        "longitude": 106.3333,
    },
    {
        "station_id": "TB_TT",
        "station_name": "Huy·ªán Th√°i Th·ª•y - Th√°i B√¨nh",
        "province": "Th√°i B√¨nh",
        "district": "Th√°i Th·ª•y",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.5333,
        "longitude": 106.5333,
    },
    {
        "station_id": "TB_VT",
        "station_name": "Huy·ªán V≈© Th∆∞ - Th√°i B√¨nh",
        "province": "Th√°i B√¨nh",
        "district": "V≈© Th∆∞",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.4333,
        "longitude": 106.2833,
    },
    # Nam ƒê·ªãnh
    {
        "station_id": "ND_TP",
        "station_name": "Th√†nh ph·ªë Nam ƒê·ªãnh - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "Nam ƒê·ªãnh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.4200,
        "longitude": 106.1683,
    },
    {
        "station_id": "ND_GL",
        "station_name": "Huy·ªán Giao Th·ªßy - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "Giao Th·ªßy",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.2333,
        "longitude": 106.4500,
    },
    {
        "station_id": "ND_HH",
        "station_name": "Huy·ªán H·∫£i H·∫≠u - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "H·∫£i H·∫≠u",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.1833,
        "longitude": 106.3000,
    },
    {
        "station_id": "ND_ML",
        "station_name": "Huy·ªán M·ªπ L·ªôc - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "M·ªπ L·ªôc",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.4667,
        "longitude": 106.1167,
    },
    {
        "station_id": "ND_NC",
        "station_name": "Huy·ªán Nam Tr·ª±c - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "Nam Tr·ª±c",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.3333,
        "longitude": 106.2000,
    },
    {
        "station_id": "ND_TK",
        "station_name": "Huy·ªán Tr·ª±c Ninh - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "Tr·ª±c Ninh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.2500,
        "longitude": 106.2500,
    },
    {
        "station_id": "ND_VB",
        "station_name": "Huy·ªán V·ª• B·∫£n - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "V·ª• B·∫£n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.3667,
        "longitude": 106.1000,
    },
    {
        "station_id": "ND_XT",
        "station_name": "Huy·ªán Xu√¢n Tr∆∞·ªùng - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "Xu√¢n Tr∆∞·ªùng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.3000,
        "longitude": 106.3500,
    },
    {
        "station_id": "ND_YN",
        "station_name": "Huy·ªán √ù Y√™n - Nam ƒê·ªãnh",
        "province": "Nam ƒê·ªãnh",
        "district": "√ù Y√™n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.3167,
        "longitude": 106.0167,
    },
    # Th·ª´a Thi√™n Hu·∫ø
    {
        "station_id": "TTH_H",
        "station_name": "Th√†nh ph·ªë Hu·∫ø - Th·ª´a Thi√™n Hu·∫ø",
        "province": "Th·ª´a Thi√™n Hu·∫ø",
        "district": "Hu·∫ø",
        "type": "Th√†nh ph·ªë",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.4637,
        "longitude": 107.5909,
    },
    {
        "station_id": "TTH_HL",
        "station_name": "Th·ªã x√£ H∆∞∆°ng Th·ªßy - Th·ª´a Thi√™n Hu·∫ø",
        "province": "Th·ª´a Thi√™n Hu·∫ø",
        "district": "H∆∞∆°ng Th·ªßy",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.4167,
        "longitude": 107.7167,
    },
    {
        "station_id": "TTH_HT",
        "station_name": "Th·ªã x√£ H∆∞∆°ng Tr√† - Th·ª´a Thi√™n Hu·∫ø",
        "province": "Th·ª´a Thi√™n Hu·∫ø",
        "district": "H∆∞∆°ng Tr√†",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.5000,
        "longitude": 107.4833,
    },
    {
        "station_id": "TTH_AD",
        "station_name": "Huy·ªán A L∆∞·ªõi - Th·ª´a Thi√™n Hu·∫ø",
        "province": "Th·ª´a Thi√™n Hu·∫ø",
        "district": "A L∆∞·ªõi",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.2333,
        "longitude": 107.3000,
    },
    {
        "station_id": "TTH_NL",
        "station_name": "Huy·ªán Nam ƒê√¥ng - Th·ª´a Thi√™n Hu·∫ø",
        "province": "Th·ª´a Thi√™n Hu·∫ø",
        "district": "Nam ƒê√¥ng",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.1667,
        "longitude": 107.7000,
    },
    {
        "station_id": "TTH_PD",
        "station_name": "Huy·ªán Phong ƒêi·ªÅn - Th·ª´a Thi√™n Hu·∫ø",
        "province": "Th·ª´a Thi√™n Hu·∫ø",
        "district": "Phong ƒêi·ªÅn",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.5833,
        "longitude": 107.3500,
    },
    {
        "station_id": "TTH_PL",
        "station_name": "Huy·ªán Ph√∫ L·ªôc - Th·ª´a Thi√™n Hu·∫ø",
        "province": "Th·ª´a Thi√™n Hu·∫ø",
        "district": "Ph√∫ L·ªôc",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.2667,
        "longitude": 107.9000,
    },
    {
        "station_id": "TTH_QD",
        "station_name": "Huy·ªán Qu·∫£ng ƒêi·ªÅn - Th·ª´a Thi√™n Hu·∫ø",
        "province": "Th·ª´a Thi√™n Hu·∫ø",
        "district": "Qu·∫£ng ƒêi·ªÅn",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.5833,
        "longitude": 107.5000,
    },
    # Qu·∫£ng Tr·ªã
    {
        "station_id": "QT_DH",
        "station_name": "Th√†nh ph·ªë ƒê√¥ng H√† - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "ƒê√¥ng H√†",
        "type": "Th√†nh ph·ªë",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.8160,
        "longitude": 107.1000,
    },
    {
        "station_id": "QT_QH",
        "station_name": "Th·ªã x√£ Qu·∫£ng Tr·ªã - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "Qu·∫£ng Tr·ªã",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.7500,
        "longitude": 107.1833,
    },
    {
        "station_id": "QT_CL",
        "station_name": "Huy·ªán Cam L·ªô - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "Cam L·ªô",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.8000,
        "longitude": 106.9833,
    },
    {
        "station_id": "QT_CC",
        "station_name": "Huy·ªán C·ªìn C·ªè - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "C·ªìn C·ªè",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.1597,
        "longitude": 107.3408,
    },
    {
        "station_id": "QT_DK",
        "station_name": "Huy·ªán ƒêa Kr√¥ng - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "ƒêa Kr√¥ng",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.5556,
        "longitude": 106.9722,
    },
    {
        "station_id": "QT_GL",
        "station_name": "Huy·ªán Gio Linh - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "Gio Linh",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.9167,
        "longitude": 107.0500,
    },
    {
        "station_id": "QT_HH",
        "station_name": "Huy·ªán H·∫£i LƒÉng - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "H·∫£i LƒÉng",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.7000,
        "longitude": 107.2500,
    },
    {
        "station_id": "QT_HL",
        "station_name": "Huy·ªán H∆∞·ªõng H√≥a - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "H∆∞·ªõng H√≥a",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.7000,
        "longitude": 106.6667,
    },
    {
        "station_id": "QT_TL",
        "station_name": "Huy·ªán Tri·ªáu Phong - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "Tri·ªáu Phong",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 16.7833,
        "longitude": 107.1667,
    },
    {
        "station_id": "QT_VL",
        "station_name": "Huy·ªán Vƒ©nh Linh - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "Vƒ©nh Linh",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.0000,
        "longitude": 106.9333,
    },
    # Qu·∫£ng B√¨nh - ƒê√É S·ª¨A TR√ôNG L·∫∂P
    {
        "station_id": "QB_DH",
        "station_name": "Th√†nh ph·ªë ƒê·ªìng H·ªõi - Qu·∫£ng B√¨nh",
        "province": "Qu·∫£ng B√¨nh",
        "district": "ƒê·ªìng H·ªõi",
        "type": "Th√†nh ph·ªë",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.4687,
        "longitude": 106.6227,
    },
    {
        "station_id": "QB_BAD",
        "station_name": "Th·ªã x√£ Ba ƒê·ªìn - Qu·∫£ng B√¨nh",
        "province": "Qu·∫£ng B√¨nh",
        "district": "Ba ƒê·ªìn",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.7500,
        "longitude": 106.4167,
    },
    {
        "station_id": "QB_BT",
        "station_name": "Huy·ªán B·ªë Tr·∫°ch - Qu·∫£ng B√¨nh",
        "province": "Qu·∫£ng B√¨nh",
        "district": "B·ªë Tr·∫°ch",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.5000,
        "longitude": 106.2500,
    },
    {
        "station_id": "QB_LT",
        "station_name": "Huy·ªán L·ªá Th·ªßy - Qu·∫£ng B√¨nh",
        "province": "Qu·∫£ng B√¨nh",
        "district": "L·ªá Th·ªßy",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.2167,
        "longitude": 106.8000,
    },
    {
        "station_id": "QB_MC",
        "station_name": "Huy·ªán Minh H√≥a - Qu·∫£ng B√¨nh",
        "province": "Qu·∫£ng B√¨nh",
        "district": "Minh H√≥a",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.7833,
        "longitude": 105.9167,
    },
    {
        "station_id": "QB_QN",
        "station_name": "Huy·ªán Qu·∫£ng Ninh - Qu·∫£ng B√¨nh",
        "province": "Qu·∫£ng B√¨nh",
        "district": "Qu·∫£ng Ninh",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.4000,
        "longitude": 106.5500,
    },
    {
        "station_id": "QB_QH",
        "station_name": "Huy·ªán Qu·∫£ng Tr·∫°ch - Qu·∫£ng B√¨nh",
        "province": "Qu·∫£ng B√¨nh",
        "district": "Qu·∫£ng Tr·∫°ch",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.8333,
        "longitude": 106.4167,
    },
    {
        "station_id": "QB_TN",
        "station_name": "Huy·ªán Tuy√™n H√≥a - Qu·∫£ng B√¨nh",
        "province": "Qu·∫£ng B√¨nh",
        "district": "Tuy√™n H√≥a",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.9000,
        "longitude": 106.0000,
    },
    # ƒê·∫Øk L·∫Øk
    {
        "station_id": "DL_BM",
        "station_name": "Th√†nh ph·ªë Bu√¥n Ma Thu·ªôt - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "Bu√¥n Ma Thu·ªôt",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.6662,
        "longitude": 108.0382,
    },
    {
        "station_id": "DL_BH",
        "station_name": "Th·ªã x√£ Bu√¥n H·ªì - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "Bu√¥n H·ªì",
        "type": "Th·ªã x√£",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.8833,
        "longitude": 108.2333,
    },
    {
        "station_id": "DL_CM",
        "station_name": "Huy·ªán C∆∞ M'gar - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "C∆∞ M'gar",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.8500,
        "longitude": 108.1000,
    },
    {
        "station_id": "DL_EH",
        "station_name": "Huy·ªán Ea H'leo - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "Ea H'leo",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.1667,
        "longitude": 108.0833,
    },
    {
        "station_id": "DL_ES",
        "station_name": "Huy·ªán Ea S√∫p - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "Ea S√∫p",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.0833,
        "longitude": 107.8333,
    },
    {
        "station_id": "DL_KK",
        "station_name": "Huy·ªán Kr√¥ng Ana - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "Kr√¥ng Ana",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.5000,
        "longitude": 108.0333,
    },
    {
        "station_id": "DL_KP",
        "station_name": "Huy·ªán Kr√¥ng P·∫Øk - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "Kr√¥ng P·∫Øk",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.7000,
        "longitude": 108.3667,
    },
    {
        "station_id": "DL_KN",
        "station_name": "Huy·ªán Kr√¥ng N√¥ - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "Kr√¥ng N√¥",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.3333,
        "longitude": 107.8333,
    },
    {
        "station_id": "DL_LK",
        "station_name": "Huy·ªán L·∫Øk - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "L·∫Øk",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.4167,
        "longitude": 108.1667,
    },
    {
        "station_id": "DL_MD",
        "station_name": "Huy·ªán M'ƒêr·∫Øk - ƒê·∫Øk L·∫Øk",
        "province": "ƒê·∫Øk L·∫Øk",
        "district": "M'ƒêr·∫Øk",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.7500,
        "longitude": 108.7500,
    },
    # Gia Lai
    {
        "station_id": "GL_PC",
        "station_name": "Th√†nh ph·ªë Pleiku - Gia Lai",
        "province": "Gia Lai",
        "district": "Pleiku",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.9833,
        "longitude": 108.0000,
    },
    {
        "station_id": "GL_AD",
        "station_name": "Th·ªã x√£ An Kh√™ - Gia Lai",
        "province": "Gia Lai",
        "district": "An Kh√™",
        "type": "Th·ªã x√£",
        "region": "T√¢y Nguy√™n",
        "latitude": 14.0000,
        "longitude": 108.6833,
    },
    {
        "station_id": "GL_AY",
        "station_name": "Th·ªã x√£ Ayun Pa - Gia Lai",
        "province": "Gia Lai",
        "district": "Ayun Pa",
        "type": "Th·ªã x√£",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.3833,
        "longitude": 108.4333,
    },
    {
        "station_id": "GL_CH",
        "station_name": "Huy·ªán Ch∆∞ PƒÉh - Gia Lai",
        "province": "Gia Lai",
        "district": "Ch∆∞ PƒÉh",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 14.1667,
        "longitude": 107.9333,
    },
    {
        "station_id": "GL_CP",
        "station_name": "Huy·ªán Ch∆∞ Pr√¥ng - Gia Lai",
        "province": "Gia Lai",
        "district": "Ch∆∞ Pr√¥ng",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.5833,
        "longitude": 107.8333,
    },
    {
        "station_id": "GL_CS",
        "station_name": "Huy·ªán Ch∆∞ S√™ - Gia Lai",
        "province": "Gia Lai",
        "district": "Ch∆∞ S√™",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.7500,
        "longitude": 108.0833,
    },
    {
        "station_id": "GL_DC",
        "station_name": "Huy·ªán ƒêƒÉk ƒêoa - Gia Lai",
        "province": "Gia Lai",
        "district": "ƒêƒÉk ƒêoa",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 14.1167,
        "longitude": 108.1167,
    },
    {
        "station_id": "GL_DP",
        "station_name": "Huy·ªán ƒêƒÉk P∆° - Gia Lai",
        "province": "Gia Lai",
        "district": "ƒêƒÉk P∆°",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 14.0000,
        "longitude": 108.5000,
    },
    {
        "station_id": "GL_DA",
        "station_name": "Huy·ªán ƒê·ª©c C∆° - Gia Lai",
        "province": "Gia Lai",
        "district": "ƒê·ª©c C∆°",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.7833,
        "longitude": 107.6667,
    },
    {
        "station_id": "GL_IA",
        "station_name": "Huy·ªán Ia Grai - Gia Lai",
        "province": "Gia Lai",
        "district": "Ia Grai",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.9833,
        "longitude": 107.7500,
    },
    {
        "station_id": "GL_KB",
        "station_name": "Huy·ªán KBang - Gia Lai",
        "province": "Gia Lai",
        "district": "KBang",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 14.3500,
        "longitude": 108.5000,
    },
    {
        "station_id": "GL_KN",
        "station_name": "Huy·ªán K√¥ng Chro - Gia Lai",
        "province": "Gia Lai",
        "district": "K√¥ng Chro",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.7167,
        "longitude": 108.5167,
    },
    {
        "station_id": "GL_KR",
        "station_name": "Huy·ªán Kr√¥ng Pa - Gia Lai",
        "province": "Gia Lai",
        "district": "Kr√¥ng Pa",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.2167,
        "longitude": 108.6667,
    },
    {
        "station_id": "GL_MY",
        "station_name": "Huy·ªán Mang Yang - Gia Lai",
        "province": "Gia Lai",
        "district": "Mang Yang",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.9833,
        "longitude": 108.2500,
    },
    {
        "station_id": "GL_PG",
        "station_name": "Huy·ªán Ph√∫ Thi·ªán - Gia Lai",
        "province": "Gia Lai",
        "district": "Ph√∫ Thi·ªán",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 13.5333,
        "longitude": 108.3000,
    },
    # L√¢m ƒê·ªìng
    {
        "station_id": "LD_DL",
        "station_name": "Th√†nh ph·ªë ƒê√† L·∫°t - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "ƒê√† L·∫°t",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.9404,
        "longitude": 108.4587,
    },
    {
        "station_id": "LD_BA",
        "station_name": "Th√†nh ph·ªë B·∫£o L·ªôc - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "B·∫£o L·ªôc",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.5500,
        "longitude": 107.8000,
    },
    {
        "station_id": "LD_BD",
        "station_name": "Huy·ªán B·∫£o L√¢m - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "B·∫£o L√¢m",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.7000,
        "longitude": 107.7167,
    },
    {
        "station_id": "LD_CD",
        "station_name": "Huy·ªán C√°t Ti√™n - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "C√°t Ti√™n",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.5833,
        "longitude": 107.3500,
    },
    {
        "station_id": "LD_DH",
        "station_name": "Huy·ªán ƒê·∫° Huoai - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "ƒê·∫° Huoai",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.4167,
        "longitude": 107.6333,
    },
    {
        "station_id": "LD_DT",
        "station_name": "Huy·ªán ƒê·∫° T·∫ªh - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "ƒê·∫° T·∫ªh",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.5000,
        "longitude": 107.5167,
    },
    {
        "station_id": "LD_DT2",
        "station_name": "Huy·ªán ƒêam R√¥ng - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "ƒêam R√¥ng",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.0833,
        "longitude": 108.1667,
    },
    {
        "station_id": "LD_DD",
        "station_name": "Huy·ªán Di Linh - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "Di Linh",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.5833,
        "longitude": 108.0667,
    },
    {
        "station_id": "LD_DL2",
        "station_name": "Huy·ªán ƒê∆°n D∆∞∆°ng - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "ƒê∆°n D∆∞∆°ng",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.7500,
        "longitude": 108.5500,
    },
    {
        "station_id": "LD_DR",
        "station_name": "Huy·ªán ƒê·ª©c Tr·ªçng - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "ƒê·ª©c Tr·ªçng",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.7333,
        "longitude": 108.3000,
    },
    {
        "station_id": "LD_LC",
        "station_name": "Huy·ªán L·∫°c D∆∞∆°ng - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "L·∫°c D∆∞∆°ng",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.0000,
        "longitude": 108.4333,
    },
    {
        "station_id": "LD_LH",
        "station_name": "Huy·ªán L√¢m H√† - L√¢m ƒê·ªìng",
        "province": "L√¢m ƒê·ªìng",
        "district": "L√¢m H√†",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 11.8333,
        "longitude": 108.2000,
    },
    # B√¨nh D∆∞∆°ng
    {
        "station_id": "BD_TDM",
        "station_name": "Th√†nh ph·ªë Th·ªß D·∫ßu M·ªôt - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "Th·ªß D·∫ßu M·ªôt",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9804,
        "longitude": 106.6519,
    },
    {
        "station_id": "BD_BG",
        "station_name": "Th·ªã x√£ B·∫øn C√°t - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "B·∫øn C√°t",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.1500,
        "longitude": 106.6000,
    },
    {
        "station_id": "BD_TU",
        "station_name": "Th·ªã x√£ T√¢n Uy√™n - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "T√¢n Uy√™n",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.0667,
        "longitude": 106.8000,
    },
    {
        "station_id": "BD_TN",
        "station_name": "Th·ªã x√£ Thu·∫≠n An - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "Thu·∫≠n An",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9333,
        "longitude": 106.7000,
    },
    {
        "station_id": "BD_DT",
        "station_name": "Th·ªã x√£ Dƒ© An - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "Dƒ© An",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9000,
        "longitude": 106.7667,
    },
    {
        "station_id": "BD_BA",
        "station_name": "Huy·ªán B√†u B√†ng - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "B√†u B√†ng",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.2500,
        "longitude": 106.8000,
    },
    {
        "station_id": "BD_BC",
        "station_name": "Huy·ªán B·∫Øc T√¢n Uy√™n - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "B·∫Øc T√¢n Uy√™n",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.1667,
        "longitude": 106.8667,
    },
    {
        "station_id": "BD_DH",
        "station_name": "Huy·ªán D·∫ßu Ti·∫øng - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "D·∫ßu Ti·∫øng",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.2667,
        "longitude": 106.3667,
    },
    {
        "station_id": "BD_PG",
        "station_name": "Huy·ªán Ph√∫ Gi√°o - B√¨nh D∆∞∆°ng",
        "province": "B√¨nh D∆∞∆°ng",
        "district": "Ph√∫ Gi√°o",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.3333,
        "longitude": 106.7667,
    },
    # ƒê·ªìng Nai
    {
        "station_id": "DN_BH",
        "station_name": "Th√†nh ph·ªë Bi√™n H√≤a - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "Bi√™n H√≤a",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9574,
        "longitude": 106.8429,
    },
    {
        "station_id": "DN_LT",
        "station_name": "Th√†nh ph·ªë Long Kh√°nh - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "Long Kh√°nh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9333,
        "longitude": 107.2500,
    },
    {
        "station_id": "DN_CM",
        "station_name": "Huy·ªán C·∫©m M·ªπ - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "C·∫©m M·ªπ",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.8000,
        "longitude": 107.2833,
    },
    {
        "station_id": "DN_DR",
        "station_name": "Huy·ªán ƒê·ªãnh Qu√°n - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "ƒê·ªãnh Qu√°n",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.1667,
        "longitude": 107.3333,
    },
    {
        "station_id": "DN_LK",
        "station_name": "Huy·ªán Long Th√†nh - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "Long Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7833,
        "longitude": 107.0000,
    },
    {
        "station_id": "DN_NT",
        "station_name": "Huy·ªán Nh∆°n Tr·∫°ch - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "Nh∆°n Tr·∫°ch",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.7000,
        "longitude": 106.8833,
    },
    {
        "station_id": "DN_TP",
        "station_name": "Huy·ªán T√¢n Ph√∫ - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "T√¢n Ph√∫",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.2833,
        "longitude": 107.4333,
    },
    {
        "station_id": "DN_TU",
        "station_name": "Huy·ªán Th·ªëng Nh·∫•t - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "Th·ªëng Nh·∫•t",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9667,
        "longitude": 107.1500,
    },
    {
        "station_id": "DN_VC",
        "station_name": "Huy·ªán Vƒ©nh C·ª≠u - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "Vƒ©nh C·ª≠u",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.2500,
        "longitude": 107.0333,
    },
    {
        "station_id": "DN_XL",
        "station_name": "Huy·ªán Xu√¢n L·ªôc - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "Xu√¢n L·ªôc",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9333,
        "longitude": 107.4167,
    },
    {
        "station_id": "DN_TB",
        "station_name": "Huy·ªán Tr·∫£ng Bom - ƒê·ªìng Nai",
        "province": "ƒê·ªìng Nai",
        "district": "Tr·∫£ng Bom",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.9500,
        "longitude": 107.0000,
    },
    # B√† R·ªãa - V≈©ng T√†u
    {
        "station_id": "BR_VT",
        "station_name": "Th√†nh ph·ªë V≈©ng T√†u - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "V≈©ng T√†u",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.3460,
        "longitude": 107.0843,
    },
    {
        "station_id": "BR_BR",
        "station_name": "Th√†nh ph·ªë B√† R·ªãa - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "B√† R·ªãa",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.5000,
        "longitude": 107.1833,
    },
    {
        "station_id": "BR_PT",
        "station_name": "Th·ªã x√£ Ph√∫ M·ªπ - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "Ph√∫ M·ªπ",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.6000,
        "longitude": 107.0500,
    },
    {
        "station_id": "BR_CD",
        "station_name": "Huy·ªán C√¥n ƒê·∫£o - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "C√¥n ƒê·∫£o",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 8.6822,
        "longitude": 106.6089,
    },
    {
        "station_id": "BR_DD",
        "station_name": "Huy·ªán ƒê·∫•t ƒê·ªè - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "ƒê·∫•t ƒê·ªè",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.5000,
        "longitude": 107.2833,
    },
    {
        "station_id": "BR_LD",
        "station_name": "Huy·ªán Long ƒêi·ªÅn - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "Long ƒêi·ªÅn",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.4667,
        "longitude": 107.2333,
    },
    {
        "station_id": "BR_XM",
        "station_name": "Huy·ªán Xuy√™n M·ªôc - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "Xuy√™n M·ªôc",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.6333,
        "longitude": 107.4333,
    },
    {
        "station_id": "BR_CH",
        "station_name": "Huy·ªán Ch√¢u ƒê·ª©c - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "Ch√¢u ƒê·ª©c",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 10.6500,
        "longitude": 107.2500,
    },
    # An Giang
    {
        "station_id": "AG_LX",
        "station_name": "Th√†nh ph·ªë Long Xuy√™n - An Giang",
        "province": "An Giang",
        "district": "Long Xuy√™n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3865,
        "longitude": 105.4351,
    },
    {
        "station_id": "AG_CD",
        "station_name": "Th√†nh ph·ªë Ch√¢u ƒê·ªëc - An Giang",
        "province": "An Giang",
        "district": "Ch√¢u ƒê·ªëc",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.7000,
        "longitude": 105.1167,
    },
    {
        "station_id": "AG_AP",
        "station_name": "Th·ªã x√£ T√¢n Ch√¢u - An Giang",
        "province": "An Giang",
        "district": "T√¢n Ch√¢u",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.8000,
        "longitude": 105.2167,
    },
    {
        "station_id": "AG_AT",
        "station_name": "Huy·ªán An Ph√∫ - An Giang",
        "province": "An Giang",
        "district": "An Ph√∫",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.8333,
        "longitude": 105.0833,
    },
    {
        "station_id": "AG_CL",
        "station_name": "Huy·ªán Ch√¢u Ph√∫ - An Giang",
        "province": "An Giang",
        "district": "Ch√¢u Ph√∫",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.5667,
        "longitude": 105.1667,
    },
    {
        "station_id": "AG_CP",
        "station_name": "Huy·ªán Ch√¢u Th√†nh - An Giang",
        "province": "An Giang",
        "district": "Ch√¢u Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.4500,
        "longitude": 105.2500,
    },
    {
        "station_id": "AG_CT",
        "station_name": "Huy·ªán Ch·ª£ M·ªõi - An Giang",
        "province": "An Giang",
        "district": "Ch·ª£ M·ªõi",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.5333,
        "longitude": 105.3833,
    },
    {
        "station_id": "AG_PT",
        "station_name": "Huy·ªán Ph√∫ T√¢n - An Giang",
        "province": "An Giang",
        "district": "Ph√∫ T√¢n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.6500,
        "longitude": 105.2833,
    },
    {
        "station_id": "AG_TV",
        "station_name": "Huy·ªán Tho·∫°i S∆°n - An Giang",
        "province": "An Giang",
        "district": "Tho·∫°i S∆°n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2667,
        "longitude": 105.2667,
    },
    {
        "station_id": "AG_TN",
        "station_name": "Huy·ªán T·ªãnh Bi√™n - An Giang",
        "province": "An Giang",
        "district": "T·ªãnh Bi√™n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.5500,
        "longitude": 105.0000,
    },
    {
        "station_id": "AG_TX",
        "station_name": "Huy·ªán Tri T√¥n - An Giang",
        "province": "An Giang",
        "district": "Tri T√¥n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.4167,
        "longitude": 105.0000,
    },
    # Ki√™n Giang
    {
        "station_id": "KG_RG",
        "station_name": "Th√†nh ph·ªë R·∫°ch Gi√° - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "R·∫°ch Gi√°",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0317,
        "longitude": 105.0809,
    },
    {
        "station_id": "KG_HN",
        "station_name": "Th√†nh ph·ªë H√† Ti√™n - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "H√† Ti√™n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3833,
        "longitude": 104.4833,
    },
    {
        "station_id": "KG_PQ",
        "station_name": "Huy·ªán Ph√∫ Qu·ªëc - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Ph√∫ Qu·ªëc",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2270,
        "longitude": 103.9679,
    },
    {
        "station_id": "KG_AT",
        "station_name": "Huy·ªán An Bi√™n - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "An Bi√™n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.8167,
        "longitude": 105.0667,
    },
    {
        "station_id": "KG_AM",
        "station_name": "Huy·ªán An Minh - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "An Minh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.6667,
        "longitude": 104.9500,
    },
    {
        "station_id": "KG_CL",
        "station_name": "Huy·ªán Ch√¢u Th√†nh - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Ch√¢u Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.9333,
        "longitude": 105.1667,
    },
    {
        "station_id": "KG_GT",
        "station_name": "Huy·ªán Giang Th√†nh - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Giang Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.4833,
        "longitude": 104.6333,
    },
    {
        "station_id": "KG_GD",
        "station_name": "Huy·ªán G√≤ Quao - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "G√≤ Quao",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.7333,
        "longitude": 105.2667,
    },
    {
        "station_id": "KG_HG",
        "station_name": "Huy·ªán H√≤n ƒê·∫•t - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "H√≤n ƒê·∫•t",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2333,
        "longitude": 104.9333,
    },
    {
        "station_id": "KG_KT",
        "station_name": "Huy·ªán Ki√™n H·∫£i - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Ki√™n H·∫£i",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.7100,
        "longitude": 104.3300,
    },
    {
        "station_id": "KG_LS",
        "station_name": "Huy·ªán Ki√™n L∆∞∆°ng - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Ki√™n L∆∞∆°ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2500,
        "longitude": 104.6333,
    },
    {
        "station_id": "KG_TG",
        "station_name": "Huy·ªán T√¢n Hi·ªáp - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "T√¢n Hi·ªáp",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.1000,
        "longitude": 105.2833,
    },
    {
        "station_id": "KG_UM",
        "station_name": "Huy·ªán U Minh Th∆∞·ª£ng - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "U Minh Th∆∞·ª£ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.6000,
        "longitude": 105.1000,
    },
    {
        "station_id": "KG_VT",
        "station_name": "Huy·ªán Vƒ©nh Thu·∫≠n - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Vƒ©nh Thu·∫≠n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.5333,
        "longitude": 105.2500,
    },
    # C√† Mau
    {
        "station_id": "CM_TP",
        "station_name": "Th√†nh ph·ªë C√† Mau - C√† Mau",
        "province": "C√† Mau",
        "district": "C√† Mau",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.1769,
        "longitude": 105.1521,
    },
    {
        "station_id": "CM_CM",
        "station_name": "Huy·ªán C√°i N∆∞·ªõc - C√† Mau",
        "province": "C√† Mau",
        "district": "C√°i N∆∞·ªõc",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.0000,
        "longitude": 105.0333,
    },
    {
        "station_id": "CM_DH",
        "station_name": "Huy·ªán ƒê·∫ßm D∆°i - C√† Mau",
        "province": "C√† Mau",
        "district": "ƒê·∫ßm D∆°i",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 8.9833,
        "longitude": 105.2000,
    },
    {
        "station_id": "CM_NC",
        "station_name": "Huy·ªán NƒÉm CƒÉn - C√† Mau",
        "province": "C√† Mau",
        "district": "NƒÉm CƒÉn",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 8.7500,
        "longitude": 105.0000,
    },
    {
        "station_id": "CM_NG",
        "station_name": "Huy·ªán Ng·ªçc Hi·ªÉn - C√† Mau",
        "province": "C√† Mau",
        "district": "Ng·ªçc Hi·ªÉn",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 8.6667,
        "longitude": 105.0500,
    },
    {
        "station_id": "CM_PT",
        "station_name": "Huy·ªán Ph√∫ T√¢n - C√† Mau",
        "province": "C√† Mau",
        "district": "Ph√∫ T√¢n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 8.8833,
        "longitude": 104.8833,
    },
    {
        "station_id": "CM_TV",
        "station_name": "Huy·ªán Th·ªõi B√¨nh - C√† Mau",
        "province": "C√† Mau",
        "district": "Th·ªõi B√¨nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.3500,
        "longitude": 105.1000,
    },
    {
        "station_id": "CM_TR",
        "station_name": "Huy·ªán Tr·∫ßn VƒÉn Th·ªùi - C√† Mau",
        "province": "C√† Mau",
        "district": "Tr·∫ßn VƒÉn Th·ªùi",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.0833,
        "longitude": 104.9667,
    },
    {
        "station_id": "CM_UM",
        "station_name": "Huy·ªán U Minh - C√† Mau",
        "province": "C√† Mau",
        "district": "U Minh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.4000,
        "longitude": 104.9833,
    },
    # Ti·ªÅn Giang
    {
        "station_id": "TG_MT",
        "station_name": "Th√†nh ph·ªë M·ªπ Tho - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "M·ªπ Tho",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3500,
        "longitude": 106.3500,
    },
    {
        "station_id": "TG_GL",
        "station_name": "Th·ªã x√£ G√≤ C√¥ng - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "G√≤ C√¥ng",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3667,
        "longitude": 106.6667,
    },
    {
        "station_id": "TG_CT",
        "station_name": "Huy·ªán C√°i B√® - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "C√°i B√®",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3333,
        "longitude": 105.9500,
    },
    {
        "station_id": "TG_CL",
        "station_name": "Huy·ªán Ch√¢u Th√†nh - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "Ch√¢u Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.4500,
        "longitude": 106.2500,
    },
    {
        "station_id": "TG_GC",
        "station_name": "Huy·ªán G√≤ C√¥ng T√¢y - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "G√≤ C√¥ng T√¢y",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3500,
        "longitude": 106.6000,
    },
    {
        "station_id": "TG_GD",
        "station_name": "Huy·ªán G√≤ C√¥ng ƒê√¥ng - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "G√≤ C√¥ng ƒê√¥ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3667,
        "longitude": 106.7167,
    },
    {
        "station_id": "TG_TP",
        "station_name": "Huy·ªán T√¢n Ph√∫ ƒê√¥ng - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "T√¢n Ph√∫ ƒê√¥ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2500,
        "longitude": 106.6000,
    },
    {
        "station_id": "TG_TL",
        "station_name": "Huy·ªán T√¢n Ph∆∞·ªõc - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "T√¢n Ph∆∞·ªõc",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.5000,
        "longitude": 106.1500,
    },
    {
        "station_id": "TG_CL2",
        "station_name": "Huy·ªán Ch·ª£ G·∫°o - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "Ch·ª£ G·∫°o",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3500,
        "longitude": 106.4667,
    },
    {
        "station_id": "TG_CB",
        "station_name": "Huy·ªán Cai L·∫≠y - Ti·ªÅn Giang",
        "province": "Ti·ªÅn Giang",
        "district": "Cai L·∫≠y",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.4000,
        "longitude": 106.1167,
    },
    # B·∫øn Tre
    {
        "station_id": "BT_BT",
        "station_name": "Th√†nh ph·ªë B·∫øn Tre - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "B·∫øn Tre",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2333,
        "longitude": 106.3833,
    },
    {
        "station_id": "BT_BL",
        "station_name": "Huy·ªán Ba Tri - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "Ba Tri",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0667,
        "longitude": 106.6000,
    },
    {
        "station_id": "BT_BH",
        "station_name": "Huy·ªán B√¨nh ƒê·∫°i - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "B√¨nh ƒê·∫°i",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.1833,
        "longitude": 106.7000,
    },
    {
        "station_id": "BT_CL",
        "station_name": "Huy·ªán Ch√¢u Th√†nh - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "Ch√¢u Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3000,
        "longitude": 106.2667,
    },
    {
        "station_id": "BT_CT",
        "station_name": "Huy·ªán Ch·ª£ L√°ch - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "Ch·ª£ L√°ch",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2667,
        "longitude": 106.1500,
    },
    {
        "station_id": "BT_GD",
        "station_name": "Huy·ªán Gi·ªìng Tr√¥m - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "Gi·ªìng Tr√¥m",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.1667,
        "longitude": 106.4667,
    },
    {
        "station_id": "BT_MB",
        "station_name": "Huy·ªán M·ªè C√†y B·∫Øc - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "M·ªè C√†y B·∫Øc",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.1333,
        "longitude": 106.3000,
    },
    {
        "station_id": "BT_MN",
        "station_name": "Huy·ªán M·ªè C√†y Nam - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "M·ªè C√†y Nam",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0667,
        "longitude": 106.3333,
    },
    {
        "station_id": "BT_TP",
        "station_name": "Huy·ªán Th·∫°nh Ph√∫ - B·∫øn Tre",
        "province": "B·∫øn Tre",
        "district": "Th·∫°nh Ph√∫",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.9333,
        "longitude": 106.5333,
    },
    # Vƒ©nh Long
    {
        "station_id": "VL_VL",
        "station_name": "Th√†nh ph·ªë Vƒ©nh Long - Vƒ©nh Long",
        "province": "Vƒ©nh Long",
        "district": "Vƒ©nh Long",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2500,
        "longitude": 105.9667,
    },
    {
        "station_id": "VL_BT",
        "station_name": "Huy·ªán B√¨nh T√¢n - Vƒ©nh Long",
        "province": "Vƒ©nh Long",
        "district": "B√¨nh T√¢n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0833,
        "longitude": 105.7667,
    },
    {
        "station_id": "VL_LH",
        "station_name": "Huy·ªán Long H·ªì - Vƒ©nh Long",
        "province": "Vƒ©nh Long",
        "district": "Long H·ªì",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2000,
        "longitude": 106.0167,
    },
    {
        "station_id": "VL_MT",
        "station_name": "Huy·ªán Mang Th√≠t - Vƒ©nh Long",
        "province": "Vƒ©nh Long",
        "district": "Mang Th√≠t",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.1833,
        "longitude": 106.1000,
    },
    {
        "station_id": "VL_TL",
        "station_name": "Huy·ªán Tam B√¨nh - Vƒ©nh Long",
        "province": "Vƒ©nh Long",
        "district": "Tam B√¨nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0667,
        "longitude": 105.9833,
    },
    {
        "station_id": "VL_TR",
        "station_name": "Huy·ªán Tr√† √în - Vƒ©nh Long",
        "province": "Vƒ©nh Long",
        "district": "Tr√† √în",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.9667,
        "longitude": 105.8833,
    },
    {
        "station_id": "VL_VT",
        "station_name": "Huy·ªán V≈©ng Li√™m - Vƒ©nh Long",
        "province": "Vƒ©nh Long",
        "district": "V≈©ng Li√™m",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.0833,
        "longitude": 106.1667,
    },
    # ƒê·ªìng Th√°p
    {
        "station_id": "DT_CL",
        "station_name": "Th√†nh ph·ªë Cao L√£nh - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "Cao L√£nh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.4500,
        "longitude": 105.6333,
    },
    {
        "station_id": "DT_SD",
        "station_name": "Th√†nh ph·ªë Sa ƒê√©c - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "Sa ƒê√©c",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3000,
        "longitude": 105.7667,
    },
    {
        "station_id": "DT_HM",
        "station_name": "Th·ªã x√£ H·ªìng Ng·ª± - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "H·ªìng Ng·ª±",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.8000,
        "longitude": 105.3333,
    },
    {
        "station_id": "DT_CL2",
        "station_name": "Huy·ªán Cao L√£nh - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "Cao L√£nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.4667,
        "longitude": 105.6333,
    },
    {
        "station_id": "DT_CH",
        "station_name": "Huy·ªán Ch√¢u Th√†nh - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "Ch√¢u Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3000,
        "longitude": 105.8167,
    },
    {
        "station_id": "DT_HN",
        "station_name": "Huy·ªán H·ªìng Ng·ª± - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "H·ªìng Ng·ª±",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.8167,
        "longitude": 105.3500,
    },
    {
        "station_id": "DT_LV",
        "station_name": "Huy·ªán Lai Vung - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "Lai Vung",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2833,
        "longitude": 105.6667,
    },
    {
        "station_id": "DT_LT",
        "station_name": "Huy·ªán L·∫•p V√≤ - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "L·∫•p V√≤",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.3500,
        "longitude": 105.5167,
    },
    {
        "station_id": "DT_TN",
        "station_name": "Huy·ªán Tam N√¥ng - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "Tam N√¥ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.7333,
        "longitude": 105.5500,
    },
    {
        "station_id": "DT_TH",
        "station_name": "Huy·ªán Thanh B√¨nh - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "Thanh B√¨nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.6000,
        "longitude": 105.4667,
    },
    {
        "station_id": "DT_TX",
        "station_name": "Huy·ªán Th√°p M∆∞·ªùi - ƒê·ªìng Th√°p",
        "province": "ƒê·ªìng Th√°p",
        "district": "Th√°p M∆∞·ªùi",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.5333,
        "longitude": 105.8167,
    },
    # H·∫≠u Giang
    {
        "station_id": "HG_VT",
        "station_name": "Th√†nh ph·ªë V·ªã Thanh - H·∫≠u Giang",
        "province": "H·∫≠u Giang",
        "district": "V·ªã Thanh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.7833,
        "longitude": 105.4667,
    },
    {
        "station_id": "HG_NM",
        "station_name": "Th·ªã x√£ Ng√£ B·∫£y - H·∫≠u Giang",
        "province": "H·∫≠u Giang",
        "district": "Ng√£ B·∫£y",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.8167,
        "longitude": 105.8167,
    },
    {
        "station_id": "HG_CL",
        "station_name": "Huy·ªán Ch√¢u Th√†nh - H·∫≠u Giang",
        "province": "H·∫≠u Giang",
        "district": "Ch√¢u Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.9333,
        "longitude": 105.6833,
    },
    {
        "station_id": "HG_CT",
        "station_name": "Huy·ªán Ch√¢u Th√†nh A - H·∫≠u Giang",
        "province": "H·∫≠u Giang",
        "district": "Ch√¢u Th√†nh A",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.9333,
        "longitude": 105.6333,
    },
    {
        "station_id": "HG_LT",
        "station_name": "Huy·ªán Long M·ªπ - H·∫≠u Giang",
        "province": "H·∫≠u Giang",
        "district": "Long M·ªπ",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.6833,
        "longitude": 105.5667,
    },
    {
        "station_id": "HG_PB",
        "station_name": "Huy·ªán Ph·ª•ng Hi·ªáp - H·∫≠u Giang",
        "province": "H·∫≠u Giang",
        "district": "Ph·ª•ng Hi·ªáp",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.7833,
        "longitude": 105.7167,
    },
    {
        "station_id": "HG_VT2",
        "station_name": "Huy·ªán V·ªã Th·ªßy - H·∫≠u Giang",
        "province": "H·∫≠u Giang",
        "district": "V·ªã Th·ªßy",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.7667,
        "longitude": 105.4667,
    },
    # S√≥c TrƒÉng
    {
        "station_id": "ST_ST",
        "station_name": "Th√†nh ph·ªë S√≥c TrƒÉng - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "S√≥c TrƒÉng",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.6025,
        "longitude": 105.9739,
    },
    {
        "station_id": "ST_VC",
        "station_name": "Th·ªã x√£ Vƒ©nh Ch√¢u - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "Vƒ©nh Ch√¢u",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.3333,
        "longitude": 105.9833,
    },
    {
        "station_id": "ST_NG",
        "station_name": "Th·ªã x√£ Ng√£ NƒÉm - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "Ng√£ NƒÉm",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.5667,
        "longitude": 105.8500,
    },
    {
        "station_id": "ST_CL",
        "station_name": "Huy·ªán Ch√¢u Th√†nh - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "Ch√¢u Th√†nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.7000,
        "longitude": 105.9000,
    },
    {
        "station_id": "ST_CT",
        "station_name": "Huy·ªán C√π Lao Dung - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "C√π Lao Dung",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.6667,
        "longitude": 106.1667,
    },
    {
        "station_id": "ST_KS",
        "station_name": "Huy·ªán K·∫ø S√°ch - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "K·∫ø S√°ch",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.8167,
        "longitude": 105.9833,
    },
    {
        "station_id": "ST_LT",
        "station_name": "Huy·ªán Long Ph√∫ - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "Long Ph√∫",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.6167,
        "longitude": 106.1167,
    },
    {
        "station_id": "ST_MD",
        "station_name": "Huy·ªán M·ªπ T√∫ - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "M·ªπ T√∫",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.6333,
        "longitude": 105.8167,
    },
    {
        "station_id": "ST_MX",
        "station_name": "Huy·ªán M·ªπ Xuy√™n - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "M·ªπ Xuy√™n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.5333,
        "longitude": 105.9833,
    },
    {
        "station_id": "ST_TN",
        "station_name": "Huy·ªán Th·∫°nh Tr·ªã - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "Th·∫°nh Tr·ªã",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.4667,
        "longitude": 105.7167,
    },
    {
        "station_id": "ST_TP",
        "station_name": "Huy·ªán Tr·∫ßn ƒê·ªÅ - S√≥c TrƒÉng",
        "province": "S√≥c TrƒÉng",
        "district": "Tr·∫ßn ƒê·ªÅ",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.5000,
        "longitude": 106.0833,
    },
    # B·∫°c Li√™u
    {
        "station_id": "BL_BL",
        "station_name": "Th√†nh ph·ªë B·∫°c Li√™u - B·∫°c Li√™u",
        "province": "B·∫°c Li√™u",
        "district": "B·∫°c Li√™u",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.2833,
        "longitude": 105.7167,
    },
    {
        "station_id": "BL_GD",
        "station_name": "Th·ªã x√£ Gi√° Rai - B·∫°c Li√™u",
        "province": "B·∫°c Li√™u",
        "district": "Gi√° Rai",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.2333,
        "longitude": 105.4667,
    },
    {
        "station_id": "BL_HB",
        "station_name": "Huy·ªán H·ªìng D√¢n - B·∫°c Li√™u",
        "province": "B·∫°c Li√™u",
        "district": "H·ªìng D√¢n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.5500,
        "longitude": 105.4167,
    },
    {
        "station_id": "BL_PT",
        "station_name": "Huy·ªán Ph∆∞·ªõc Long - B·∫°c Li√™u",
        "province": "B·∫°c Li√™u",
        "district": "Ph∆∞·ªõc Long",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.4333,
        "longitude": 105.4667,
    },
    {
        "station_id": "BL_VL",
        "station_name": "Huy·ªán Vƒ©nh L·ª£i - B·∫°c Li√™u",
        "province": "B·∫°c Li√™u",
        "district": "Vƒ©nh L·ª£i",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.3500,
        "longitude": 105.5667,
    },
    {
        "station_id": "BL_DD",
        "station_name": "Huy·ªán ƒê√¥ng H·∫£i - B·∫°c Li√™u",
        "province": "B·∫°c Li√™u",
        "district": "ƒê√¥ng H·∫£i",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.1667,
        "longitude": 105.4333,
    },
    {
        "station_id": "BL_HH",
        "station_name": "Huy·ªán H√≤a B√¨nh - B·∫°c Li√™u",
        "province": "B·∫°c Li√™u",
        "district": "H√≤a B√¨nh",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.2833,
        "longitude": 105.6333,
    },
    # Thanh H√≥a
    {
        "station_id": "TH_TH",
        "station_name": "Th√†nh ph·ªë Thanh H√≥a - Thanh H√≥a",
        "province": "Thanh H√≥a",
        "district": "Thanh H√≥a",
        "type": "Th√†nh ph·ªë",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 19.8000,
        "longitude": 105.7667,
    },
    {
        "station_id": "TH_SL",
        "station_name": "Th·ªã x√£ S·∫ßm S∆°n - Thanh H√≥a",
        "province": "Thanh H√≥a",
        "district": "S·∫ßm S∆°n",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 19.7333,
        "longitude": 105.9000,
    },
    {
        "station_id": "TH_BT",
        "station_name": "Th·ªã x√£ B·ªâm S∆°n - Thanh H√≥a",
        "province": "Thanh H√≥a",
        "district": "B·ªâm S∆°n",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 20.0781,
        "longitude": 105.8603,
    },
    {
        "station_id": "TH_NC",
        "station_name": "Huy·ªán Nghi S∆°n - Thanh H√≥a",
        "province": "Thanh H√≥a",
        "district": "Nghi S∆°n",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 19.4500,
        "longitude": 105.7833,
    },
    # Ngh·ªá An
    {
        "station_id": "NA_VH",
        "station_name": "Th√†nh ph·ªë Vinh - Ngh·ªá An",
        "province": "Ngh·ªá An",
        "district": "Vinh",
        "type": "Th√†nh ph·ªë",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 18.6733,
        "longitude": 105.6811,
    },
    {
        "station_id": "NA_CT",
        "station_name": "Th·ªã x√£ C·ª≠a L√≤ - Ngh·ªá An",
        "province": "Ngh·ªá An",
        "district": "C·ª≠a L√≤",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 18.8167,
        "longitude": 105.7167,
    },
    {
        "station_id": "NA_TH",
        "station_name": "Th·ªã x√£ Th√°i H√≤a - Ngh·ªá An",
        "province": "Ngh·ªá An",
        "district": "Th√°i H√≤a",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 19.3000,
        "longitude": 105.4667,
    },
    {
        "station_id": "NA_HH",
        "station_name": "Huy·ªán Ho√†ng Mai - Ngh·ªá An",
        "province": "Ngh·ªá An",
        "district": "Ho√†ng Mai",
        "type": "Huy·ªán",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 19.2667,
        "longitude": 105.7167,
    },
    # H√† Tƒ©nh
    {
        "station_id": "HT_HT",
        "station_name": "Th√†nh ph·ªë H√† Tƒ©nh - H√† Tƒ©nh",
        "province": "H√† Tƒ©nh",
        "district": "H√† Tƒ©nh",
        "type": "Th√†nh ph·ªë",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 18.3333,
        "longitude": 105.9000,
    },
    {
        "station_id": "HT_HK",
        "station_name": "Th·ªã x√£ H·ªìng Lƒ©nh - H√† Tƒ©nh",
        "province": "H√† Tƒ©nh",
        "district": "H·ªìng Lƒ©nh",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 18.5333,
        "longitude": 105.7167,
    },
    {
        "station_id": "HT_KL",
        "station_name": "Th·ªã x√£ K·ª≥ Anh - H√† Tƒ©nh",
        "province": "H√† Tƒ©nh",
        "district": "K·ª≥ Anh",
        "type": "Th·ªã x√£",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 18.0833,
        "longitude": 106.3000,
    },
    # Qu·∫£ng Nam
    {
        "station_id": "QN_TC",
        "station_name": "Th√†nh ph·ªë Tam K·ª≥ - Qu·∫£ng Nam",
        "province": "Qu·∫£ng Nam",
        "district": "Tam K·ª≥",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 15.5667,
        "longitude": 108.4833,
    },
    {
        "station_id": "QN_HT",
        "station_name": "Th√†nh ph·ªë H·ªôi An - Qu·∫£ng Nam",
        "province": "Qu·∫£ng Nam",
        "district": "H·ªôi An",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 15.8833,
        "longitude": 108.3333,
    },
    {
        "station_id": "QN_DG",
        "station_name": "Huy·ªán ƒêi·ªán B√†n - Qu·∫£ng Nam",
        "province": "Qu·∫£ng Nam",
        "district": "ƒêi·ªán B√†n",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 15.9000,
        "longitude": 108.2500,
    },
    {
        "station_id": "QN_NT",
        "station_name": "Huy·ªán N√∫i Th√†nh - Qu·∫£ng Nam",
        "province": "Qu·∫£ng Nam",
        "district": "N√∫i Th√†nh",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 15.4333,
        "longitude": 108.6667,
    },
    # Qu·∫£ng Ng√£i
    {
        "station_id": "QN_QN",
        "station_name": "Th√†nh ph·ªë Qu·∫£ng Ng√£i - Qu·∫£ng Ng√£i",
        "province": "Qu·∫£ng Ng√£i",
        "district": "Qu·∫£ng Ng√£i",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 15.1167,
        "longitude": 108.8000,
    },
    {
        "station_id": "QN_LS",
        "station_name": "Huy·ªán L√Ω S∆°n - Qu·∫£ng Ng√£i",
        "province": "Qu·∫£ng Ng√£i",
        "district": "L√Ω S∆°n",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 15.3833,
        "longitude": 109.1167,
    },
    {
        "station_id": "QN_BT",
        "station_name": "Huy·ªán B√¨nh S∆°n - Qu·∫£ng Ng√£i",
        "province": "Qu·∫£ng Ng√£i",
        "district": "B√¨nh S∆°n",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 15.3167,
        "longitude": 108.7667,
    },
    # B√¨nh ƒê·ªãnh
    {
        "station_id": "BD_QN",
        "station_name": "Th√†nh ph·ªë Quy Nh∆°n - B√¨nh ƒê·ªãnh",
        "province": "B√¨nh ƒê·ªãnh",
        "district": "Quy Nh∆°n",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 13.7667,
        "longitude": 109.2333,
    },
    {
        "station_id": "BD_AL",
        "station_name": "Huy·ªán An L√£o - B√¨nh ƒê·ªãnh",
        "province": "B√¨nh ƒê·ªãnh",
        "district": "An L√£o",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 14.5667,
        "longitude": 108.9000,
    },
    {
        "station_id": "BD_HV",
        "station_name": "Huy·ªán Ho√†i √Çn - B√¨nh ƒê·ªãnh",
        "province": "B√¨nh ƒê·ªãnh",
        "district": "Ho√†i √Çn",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 14.3500,
        "longitude": 108.9000,
    },
    # Ph√∫ Y√™n
    {
        "station_id": "PY_TH",
        "station_name": "Th√†nh ph·ªë Tuy H√≤a - Ph√∫ Y√™n",
        "province": "Ph√∫ Y√™n",
        "district": "Tuy H√≤a",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 13.0833,
        "longitude": 109.3000,
    },
    {
        "station_id": "PY_SH",
        "station_name": "Th·ªã x√£ S√¥ng C·∫ßu - Ph√∫ Y√™n",
        "province": "Ph√∫ Y√™n",
        "district": "S√¥ng C·∫ßu",
        "type": "Th·ªã x√£",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 13.4500,
        "longitude": 109.2167,
    },
    {
        "station_id": "PY_DD",
        "station_name": "Huy·ªán ƒê√¥ng H√≤a - Ph√∫ Y√™n",
        "province": "Ph√∫ Y√™n",
        "district": "ƒê√¥ng H√≤a",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 12.9833,
        "longitude": 109.3667,
    },
    # Kh√°nh H√≤a
    {
        "station_id": "KH_NT",
        "station_name": "Th√†nh ph·ªë Nha Trang - Kh√°nh H√≤a",
        "province": "Kh√°nh H√≤a",
        "district": "Nha Trang",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 12.2500,
        "longitude": 109.1833,
    },
    {
        "station_id": "KH_CT",
        "station_name": "Th√†nh ph·ªë Cam Ranh - Kh√°nh H√≤a",
        "province": "Kh√°nh H√≤a",
        "district": "Cam Ranh",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 11.9000,
        "longitude": 109.1333,
    },
    {
        "station_id": "KH_NH",
        "station_name": "Th·ªã x√£ Ninh H√≤a - Kh√°nh H√≤a",
        "province": "Kh√°nh H√≤a",
        "district": "Ninh H√≤a",
        "type": "Th·ªã x√£",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 12.5000,
        "longitude": 109.1500,
    },
    {
        "station_id": "KH_TS",
        "station_name": "Huy·ªán Tr∆∞·ªùng Sa - Kh√°nh H√≤a",
        "province": "Kh√°nh H√≤a",
        "district": "Tr∆∞·ªùng Sa",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 8.6500,
        "longitude": 111.9167,
    },
    # Ninh Thu·∫≠n
    {
        "station_id": "NT_PH",
        "station_name": "Th√†nh ph·ªë Phan Rang-Th√°p Ch√†m - Ninh Thu·∫≠n",
        "province": "Ninh Thu·∫≠n",
        "district": "Phan Rang-Th√°p Ch√†m",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 11.5667,
        "longitude": 108.9833,
    },
    {
        "station_id": "NT_TA",
        "station_name": "Huy·ªán Thu·∫≠n B·∫Øc - Ninh Thu·∫≠n",
        "province": "Ninh Thu·∫≠n",
        "district": "Thu·∫≠n B·∫Øc",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 11.7500,
        "longitude": 108.9333,
    },
    {
        "station_id": "NT_NA",
        "station_name": "Huy·ªán Ninh H·∫£i - Ninh Thu·∫≠n",
        "province": "Ninh Thu·∫≠n",
        "district": "Ninh H·∫£i",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 11.6000,
        "longitude": 109.0333,
    },
    # B√¨nh Thu·∫≠n
    {
        "station_id": "BT_PT",
        "station_name": "Th√†nh ph·ªë Phan Thi·∫øt - B√¨nh Thu·∫≠n",
        "province": "B√¨nh Thu·∫≠n",
        "district": "Phan Thi·∫øt",
        "type": "Th√†nh ph·ªë",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 10.9333,
        "longitude": 108.1000,
    },
    {
        "station_id": "BT_LC",
        "station_name": "Th·ªã x√£ La Gi - B√¨nh Thu·∫≠n",
        "province": "B√¨nh Thu·∫≠n",
        "district": "La Gi",
        "type": "Th·ªã x√£",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 10.6667,
        "longitude": 107.7667,
    },
    {
        "station_id": "BT_PH",
        "station_name": "Huy·ªán Ph√∫ Qu√Ω - B√¨nh Thu·∫≠n",
        "province": "B√¨nh Thu·∫≠n",
        "district": "Ph√∫ Qu√Ω",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 10.5000,
        "longitude": 108.9333,
    },
    {
        "station_id": "BT_HD",
        "station_name": "Huy·ªán H√†m Thu·∫≠n B·∫Øc - B√¨nh Thu·∫≠n",
        "province": "B√¨nh Thu·∫≠n",
        "district": "H√†m Thu·∫≠n B·∫Øc",
        "type": "Huy·ªán",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 11.0833,
        "longitude": 108.1167,
    },
    # ƒêi·ªán Bi√™n
    {
        "station_id": "DB_DB",
        "station_name": "Th√†nh ph·ªë ƒêi·ªán Bi√™n Ph·ªß - ƒêi·ªán Bi√™n",
        "province": "ƒêi·ªán Bi√™n",
        "district": "ƒêi·ªán Bi√™n Ph·ªß",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y B·∫Øc",
        "latitude": 21.3833,
        "longitude": 103.0167,
    },
    {
        "station_id": "DB_ML",
        "station_name": "Th·ªã x√£ M∆∞·ªùng Lay - ƒêi·ªán Bi√™n",
        "province": "ƒêi·ªán Bi√™n",
        "district": "M∆∞·ªùng Lay",
        "type": "Th·ªã x√£",
        "region": "T√¢y B·∫Øc",
        "latitude": 22.0333,
        "longitude": 103.1500,
    },
    {
        "station_id": "DB_DL",
        "station_name": "Huy·ªán ƒêi·ªán Bi√™n - ƒêi·ªán Bi√™n",
        "province": "ƒêi·ªán Bi√™n",
        "district": "ƒêi·ªán Bi√™n",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 21.2500,
        "longitude": 103.0333,
    },
    {
        "station_id": "DB_DD",
        "station_name": "Huy·ªán ƒêi·ªán Bi√™n ƒê√¥ng - ƒêi·ªán Bi√™n",
        "province": "ƒêi·ªán Bi√™n",
        "district": "ƒêi·ªán Bi√™n ƒê√¥ng",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 21.3333,
        "longitude": 103.2500,
    },
    # S∆°n La
    {
        "station_id": "SL_SL",
        "station_name": "Th√†nh ph·ªë S∆°n La - S∆°n La",
        "province": "S∆°n La",
        "district": "S∆°n La",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y B·∫Øc",
        "latitude": 21.3256,
        "longitude": 103.9189,
    },
    {
        "station_id": "SL_ML",
        "station_name": "Huy·ªán M·ªôc Ch√¢u - S∆°n La",
        "province": "S∆°n La",
        "district": "M·ªôc Ch√¢u",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 20.8500,
        "longitude": 104.6333,
    },
    {
        "station_id": "SL_YB",
        "station_name": "Huy·ªán Y√™n Ch√¢u - S∆°n La",
        "province": "S∆°n La",
        "district": "Y√™n Ch√¢u",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 21.0500,
        "longitude": 104.3000,
    },
    # Lai Ch√¢u
    {
        "station_id": "LC_LC",
        "station_name": "Th√†nh ph·ªë Lai Ch√¢u - Lai Ch√¢u",
        "province": "Lai Ch√¢u",
        "district": "Lai Ch√¢u",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y B·∫Øc",
        "latitude": 22.4000,
        "longitude": 103.4500,
    },
    {
        "station_id": "LC_TN",
        "station_name": "Huy·ªán Tam ƒê∆∞·ªùng - Lai Ch√¢u",
        "province": "Lai Ch√¢u",
        "district": "Tam ƒê∆∞·ªùng",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 22.3667,
        "longitude": 103.6167,
    },
    {
        "station_id": "LC_PT",
        "station_name": "Huy·ªán Phong Th·ªï - Lai Ch√¢u",
        "province": "Lai Ch√¢u",
        "district": "Phong Th·ªï",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 22.5333,
        "longitude": 103.3333,
    },
    # H√≤a B√¨nh
    {
        "station_id": "HB_HB",
        "station_name": "Th√†nh ph·ªë H√≤a B√¨nh - H√≤a B√¨nh",
        "province": "H√≤a B√¨nh",
        "district": "H√≤a B√¨nh",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y B·∫Øc",
        "latitude": 20.8133,
        "longitude": 105.3383,
    },
    {
        "station_id": "HB_ML",
        "station_name": "Huy·ªán Mai Ch√¢u - H√≤a B√¨nh",
        "province": "H√≤a B√¨nh",
        "district": "Mai Ch√¢u",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 20.6667,
        "longitude": 105.0833,
    },
    {
        "station_id": "HB_KB",
        "station_name": "Huy·ªán Kim B√¥i - H√≤a B√¨nh",
        "province": "H√≤a B√¨nh",
        "district": "Kim B√¥i",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 20.6667,
        "longitude": 105.5333,
    },
    # L·∫°ng S∆°n
    {
        "station_id": "LS_LS",
        "station_name": "Th√†nh ph·ªë L·∫°ng S∆°n - L·∫°ng S∆°n",
        "province": "L·∫°ng S∆°n",
        "district": "L·∫°ng S∆°n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.8478,
        "longitude": 106.7578,
    },
    {
        "station_id": "LS_CL",
        "station_name": "Huy·ªán Cao L·ªôc - L·∫°ng S∆°n",
        "province": "L·∫°ng S∆°n",
        "district": "Cao L·ªôc",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.9000,
        "longitude": 106.8000,
    },
    {
        "station_id": "LS_LL",
        "station_name": "Huy·ªán L·ªôc B√¨nh - L·∫°ng S∆°n",
        "province": "L·∫°ng S∆°n",
        "district": "L·ªôc B√¨nh",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.7500,
        "longitude": 106.9333,
    },
    # Cao B·∫±ng
    {
        "station_id": "CB_CB",
        "station_name": "Th√†nh ph·ªë Cao B·∫±ng - Cao B·∫±ng",
        "province": "Cao B·∫±ng",
        "district": "Cao B·∫±ng",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.6667,
        "longitude": 106.2500,
    },
    {
        "station_id": "CB_BL",
        "station_name": "Huy·ªán B·∫£o L√¢m - Cao B·∫±ng",
        "province": "Cao B·∫±ng",
        "district": "B·∫£o L√¢m",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.8667,
        "longitude": 105.5000,
    },
    {
        "station_id": "CB_BC",
        "station_name": "Huy·ªán B·∫£o L·∫°c - Cao B·∫±ng",
        "province": "Cao B·∫±ng",
        "district": "B·∫£o L·∫°c",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.9500,
        "longitude": 105.7333,
    },
    # B·∫Øc K·∫°n
    {
        "station_id": "BK_BK",
        "station_name": "Th√†nh ph·ªë B·∫Øc K·∫°n - B·∫Øc K·∫°n",
        "province": "B·∫Øc K·∫°n",
        "district": "B·∫Øc K·∫°n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.1500,
        "longitude": 105.8333,
    },
    {
        "station_id": "BK_BB",
        "station_name": "Huy·ªán Ba B·ªÉ - B·∫Øc K·∫°n",
        "province": "B·∫Øc K·∫°n",
        "district": "Ba B·ªÉ",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.4167,
        "longitude": 105.7500,
    },
    {
        "station_id": "BK_PT",
        "station_name": "Huy·ªán P√°c N·∫∑m - B·∫Øc K·∫°n",
        "province": "B·∫Øc K·∫°n",
        "district": "P√°c N·∫∑m",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.6333,
        "longitude": 105.6667,
    },
    # Th√°i Nguy√™n
    {
        "station_id": "TN_TN",
        "station_name": "Th√†nh ph·ªë Th√°i Nguy√™n - Th√°i Nguy√™n",
        "province": "Th√°i Nguy√™n",
        "district": "Th√°i Nguy√™n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.5928,
        "longitude": 105.8442,
    },
    {
        "station_id": "TN_SC",
        "station_name": "Th√†nh ph·ªë S√¥ng C√¥ng - Th√°i Nguy√™n",
        "province": "Th√°i Nguy√™n",
        "district": "S√¥ng C√¥ng",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.4833,
        "longitude": 105.8500,
    },
    {
        "station_id": "TN_DD",
        "station_name": "Huy·ªán ƒê·ªãnh H√≥a - Th√°i Nguy√™n",
        "province": "Th√°i Nguy√™n",
        "district": "ƒê·ªãnh H√≥a",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.9000,
        "longitude": 105.6333,
    },
    # Tuy√™n Quang
    {
        "station_id": "TQ_TQ",
        "station_name": "Th√†nh ph·ªë Tuy√™n Quang - Tuy√™n Quang",
        "province": "Tuy√™n Quang",
        "district": "Tuy√™n Quang",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.8181,
        "longitude": 105.2144,
    },
    {
        "station_id": "TQ_NH",
        "station_name": "Huy·ªán Na Hang - Tuy√™n Quang",
        "province": "Tuy√™n Quang",
        "district": "Na Hang",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.3500,
        "longitude": 105.3833,
    },
    {
        "station_id": "TQ_LB",
        "station_name": "Huy·ªán L√¢m B√¨nh - Tuy√™n Quang",
        "province": "Tuy√™n Quang",
        "district": "L√¢m B√¨nh",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.4667,
        "longitude": 105.2167,
    },
    # Y√™n B√°i
    {
        "station_id": "YB_YB",
        "station_name": "Th√†nh ph·ªë Y√™n B√°i - Y√™n B√°i",
        "province": "Y√™n B√°i",
        "district": "Y√™n B√°i",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.7000,
        "longitude": 104.8667,
    },
    {
        "station_id": "YB_NZ",
        "station_name": "Th·ªã x√£ Nghƒ©a L·ªô - Y√™n B√°i",
        "province": "Y√™n B√°i",
        "district": "Nghƒ©a L·ªô",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.6000,
        "longitude": 104.5000,
    },
    {
        "station_id": "YB_MT",
        "station_name": "Huy·ªán M√π Cang Ch·∫£i - Y√™n B√°i",
        "province": "Y√™n B√°i",
        "district": "M√π Cang Ch·∫£i",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.8500,
        "longitude": 104.0833,
    },
    # Ph√∫ Th·ªç
    {
        "station_id": "PT_VT",
        "station_name": "Th√†nh ph·ªë Vi·ªát Tr√¨ - Ph√∫ Th·ªç",
        "province": "Ph√∫ Th·ªç",
        "district": "Vi·ªát Tr√¨",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.3000,
        "longitude": 105.4333,
    },
    {
        "station_id": "PT_PT",
        "station_name": "Th·ªã x√£ Ph√∫ Th·ªç - Ph√∫ Th·ªç",
        "province": "Ph√∫ Th·ªç",
        "district": "Ph√∫ Th·ªç",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.4000,
        "longitude": 105.2333,
    },
    {
        "station_id": "PT_HT",
        "station_name": "Huy·ªán H·∫° H√≤a - Ph√∫ Th·ªç",
        "province": "Ph√∫ Th·ªç",
        "district": "H·∫° H√≤a",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.5500,
        "longitude": 105.0000,
    },
    # Vƒ©nh Ph√∫c
    {
        "station_id": "VP_VY",
        "station_name": "Th√†nh ph·ªë Vƒ©nh Y√™n - Vƒ©nh Ph√∫c",
        "province": "Vƒ©nh Ph√∫c",
        "district": "Vƒ©nh Y√™n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.3089,
        "longitude": 105.6044,
    },
    {
        "station_id": "VP_PC",
        "station_name": "Th√†nh ph·ªë Ph√∫c Y√™n - Vƒ©nh Ph√∫c",
        "province": "Vƒ©nh Ph√∫c",
        "district": "Ph√∫c Y√™n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.2333,
        "longitude": 105.7000,
    },
    {
        "station_id": "VP_LP",
        "station_name": "Huy·ªán L·∫≠p Th·∫°ch - Vƒ©nh Ph√∫c",
        "province": "Vƒ©nh Ph√∫c",
        "district": "L·∫≠p Th·∫°ch",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.4167,
        "longitude": 105.4667,
    },
    # B·∫Øc Ninh
    {
        "station_id": "BN_BN",
        "station_name": "Th√†nh ph·ªë B·∫Øc Ninh - B·∫Øc Ninh",
        "province": "B·∫Øc Ninh",
        "district": "B·∫Øc Ninh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.1861,
        "longitude": 106.0764,
    },
    {
        "station_id": "BN_TS",
        "station_name": "Th√†nh ph·ªë T·ª´ S∆°n - B·∫Øc Ninh",
        "province": "B·∫Øc Ninh",
        "district": "T·ª´ S∆°n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.1167,
        "longitude": 105.9667,
    },
    {
        "station_id": "BN_YL",
        "station_name": "Huy·ªán Y√™n Phong - B·∫Øc Ninh",
        "province": "B·∫Øc Ninh",
        "district": "Y√™n Phong",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.2000,
        "longitude": 105.9500,
    },
    # H·∫£i D∆∞∆°ng
    {
        "station_id": "HD_HD",
        "station_name": "Th√†nh ph·ªë H·∫£i D∆∞∆°ng - H·∫£i D∆∞∆°ng",
        "province": "H·∫£i D∆∞∆°ng",
        "district": "H·∫£i D∆∞∆°ng",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.9397,
        "longitude": 106.3308,
    },
    {
        "station_id": "HD_CL",
        "station_name": "Th·ªã x√£ Kinh M√¥n - H·∫£i D∆∞∆°ng",
        "province": "H·∫£i D∆∞∆°ng",
        "district": "Kinh M√¥n",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 21.0167,
        "longitude": 106.5000,
    },
    {
        "station_id": "HD_CM",
        "station_name": "Huy·ªán C·∫©m Gi√†ng - H·∫£i D∆∞∆°ng",
        "province": "H·∫£i D∆∞∆°ng",
        "district": "C·∫©m Gi√†ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.9500,
        "longitude": 106.2167,
    },
    # H∆∞ng Y√™n
    {
        "station_id": "HY_HY",
        "station_name": "Th√†nh ph·ªë H∆∞ng Y√™n - H∆∞ng Y√™n",
        "province": "H∆∞ng Y√™n",
        "district": "H∆∞ng Y√™n",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.6461,
        "longitude": 106.0511,
    },
    {
        "station_id": "HY_ML",
        "station_name": "Huy·ªán M·ªπ H√†o - H∆∞ng Y√™n",
        "province": "H∆∞ng Y√™n",
        "district": "M·ªπ H√†o",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.9333,
        "longitude": 106.0667,
    },
    {
        "station_id": "HY_VD",
        "station_name": "Huy·ªán VƒÉn L√¢m - H∆∞ng Y√™n",
        "province": "H∆∞ng Y√™n",
        "district": "VƒÉn L√¢m",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.9833,
        "longitude": 106.0333,
    },
    # Ninh B√¨nh
    {
        "station_id": "NB_NB",
        "station_name": "Th√†nh ph·ªë Ninh B√¨nh - Ninh B√¨nh",
        "province": "Ninh B√¨nh",
        "district": "Ninh B√¨nh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.2539,
        "longitude": 105.9750,
    },
    {
        "station_id": "NB_TL",
        "station_name": "Th√†nh ph·ªë Tam ƒêi·ªáp - Ninh B√¨nh",
        "province": "Ninh B√¨nh",
        "district": "Tam ƒêi·ªáp",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.1500,
        "longitude": 105.9167,
    },
    {
        "station_id": "NB_HS",
        "station_name": "Huy·ªán Hoa L∆∞ - Ninh B√¨nh",
        "province": "Ninh B√¨nh",
        "district": "Hoa L∆∞",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.2833,
        "longitude": 105.9167,
    },
    # B·∫Øc Giang
    {
        "station_id": "BG_BG",
        "station_name": "Th√†nh ph·ªë B·∫Øc Giang - B·∫Øc Giang",
        "province": "B·∫Øc Giang",
        "district": "B·∫Øc Giang",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.2667,
        "longitude": 106.2000,
    },
    {
        "station_id": "BG_YL",
        "station_name": "Huy·ªán Y√™n Th·∫ø - B·∫Øc Giang",
        "province": "B·∫Øc Giang",
        "district": "Y√™n Th·∫ø",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.5167,
        "longitude": 106.1167,
    },
    {
        "station_id": "BG_LN",
        "station_name": "Huy·ªán L·∫°ng Giang - B·∫Øc Giang",
        "province": "B·∫Øc Giang",
        "district": "L·∫°ng Giang",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 21.3500,
        "longitude": 106.2500,
    },
    # B√¨nh Ph∆∞·ªõc
    {
        "station_id": "BP_DX",
        "station_name": "Th√†nh ph·ªë ƒê·ªìng Xo√†i - B√¨nh Ph∆∞·ªõc",
        "province": "B√¨nh Ph∆∞·ªõc",
        "district": "ƒê·ªìng Xo√†i",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.5349,
        "longitude": 106.8823,
    },
    {
        "station_id": "BP_BL",
        "station_name": "Th·ªã x√£ B√¨nh Long - B√¨nh Ph∆∞·ªõc",
        "province": "B√¨nh Ph∆∞·ªõc",
        "district": "B√¨nh Long",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.6500,
        "longitude": 106.6000,
    },
    {
        "station_id": "BP_PL",
        "station_name": "Th·ªã x√£ Ph∆∞·ªõc Long - B√¨nh Ph∆∞·ªõc",
        "province": "B√¨nh Ph∆∞·ªõc",
        "district": "Ph∆∞·ªõc Long",
        "type": "Th·ªã x√£",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.8333,
        "longitude": 106.9667,
    },
    # ƒê·∫Øk N√¥ng
    {
        "station_id": "DKN_GN",
        "station_name": "Th√†nh ph·ªë Gia Nghƒ©a - ƒê·∫Øk N√¥ng",
        "province": "ƒê·∫Øk N√¥ng",
        "district": "Gia Nghƒ©a",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.0042,
        "longitude": 107.6907,
    },
    {
        "station_id": "DKN_CJ",
        "station_name": "Huy·ªán C∆∞ J√∫t - ƒê·∫Øk N√¥ng",
        "province": "ƒê·∫Øk N√¥ng",
        "district": "C∆∞ J√∫t",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.6569,
        "longitude": 107.7636,
    },
    {
        "station_id": "DKN_KM",
        "station_name": "Huy·ªán Kr√¥ng N√¥ - ƒê·∫Øk N√¥ng",
        "province": "ƒê·∫Øk N√¥ng",
        "district": "Kr√¥ng N√¥",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 12.4500,
        "longitude": 107.8667,
    },
    # H√† Giang
    {
        "station_id": "HG_HG",
        "station_name": "Th√†nh ph·ªë H√† Giang - H√† Giang",
        "province": "H√† Giang",
        "district": "H√† Giang",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 22.8233,
        "longitude": 104.9836,
    },
    {
        "station_id": "HG_DV",
        "station_name": "Huy·ªán ƒê·ªìng VƒÉn - H√† Giang",
        "province": "H√† Giang",
        "district": "ƒê·ªìng VƒÉn",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 23.2547,
        "longitude": 105.2500,
    },
    {
        "station_id": "HG_MK",
        "station_name": "Huy·ªán M√®o V·∫°c - H√† Giang",
        "province": "H√† Giang",
        "district": "M√®o V·∫°c",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 23.1528,
        "longitude": 105.4069,
    },
    # H√† Nam
    {
        "station_id": "HNA_PL",
        "station_name": "Th√†nh ph·ªë Ph·ªß L√Ω - H√† Nam",
        "province": "H√† Nam",
        "district": "Ph·ªß L√Ω",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.5411,
        "longitude": 105.9139,
    },
    {
        "station_id": "HNA_DT",
        "station_name": "Huy·ªán Duy Ti√™n - H√† Nam",
        "province": "H√† Nam",
        "district": "Duy Ti√™n",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.6333,
        "longitude": 105.9667,
    },
    {
        "station_id": "HNA_KB",
        "station_name": "Huy·ªán Kim B·∫£ng - H√† Nam",
        "province": "H√† Nam",
        "district": "Kim B·∫£ng",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.5667,
        "longitude": 105.8500,
    },
    # Kon Tum
    {
        "station_id": "KT_KT",
        "station_name": "Th√†nh ph·ªë Kon Tum - Kon Tum",
        "province": "Kon Tum",
        "district": "Kon Tum",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y Nguy√™n",
        "latitude": 14.3833,
        "longitude": 107.9833,
    },
    {
        "station_id": "KT_DG",
        "station_name": "Huy·ªán ƒê·∫Øk Glei - Kon Tum",
        "province": "Kon Tum",
        "district": "ƒê·∫Øk Glei",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 15.0167,
        "longitude": 107.7500,
    },
    {
        "station_id": "KT_NH",
        "station_name": "Huy·ªán Ng·ªçc H·ªìi - Kon Tum",
        "province": "Kon Tum",
        "district": "Ng·ªçc H·ªìi",
        "type": "Huy·ªán",
        "region": "T√¢y Nguy√™n",
        "latitude": 14.7000,
        "longitude": 107.8333,
    },
    # L√†o Cai
    {
        "station_id": "LC_LC2",
        "station_name": "Th√†nh ph·ªë L√†o Cai - L√†o Cai",
        "province": "L√†o Cai",
        "district": "L√†o Cai",
        "type": "Th√†nh ph·ªë",
        "region": "T√¢y B·∫Øc",
        "latitude": 22.4833,
        "longitude": 103.9500,
    },
    {
        "station_id": "LC_SM",
        "station_name": "Huy·ªán Sa Pa - L√†o Cai",
        "province": "L√†o Cai",
        "district": "Sa Pa",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 22.3367,
        "longitude": 103.8400,
    },
    {
        "station_id": "LC_BT",
        "station_name": "Huy·ªán B√°t X√°t - L√†o Cai",
        "province": "L√†o Cai",
        "district": "B√°t X√°t",
        "type": "Huy·ªán",
        "region": "T√¢y B·∫Øc",
        "latitude": 22.5333,
        "longitude": 103.8833,
    },
    # Long An
    {
        "station_id": "LA_TA",
        "station_name": "Th√†nh ph·ªë T√¢n An - Long An",
        "province": "Long An",
        "district": "T√¢n An",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.5333,
        "longitude": 106.4167,
    },
    {
        "station_id": "LA_KT",
        "station_name": "Th·ªã x√£ Ki·∫øn T∆∞·ªùng - Long An",
        "province": "Long An",
        "district": "Ki·∫øn T∆∞·ªùng",
        "type": "Th·ªã x√£",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.7667,
        "longitude": 105.9000,
    },
    {
        "station_id": "LA_DH",
        "station_name": "Huy·ªán ƒê·ª©c H√≤a - Long An",
        "province": "Long An",
        "district": "ƒê·ª©c H√≤a",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.8833,
        "longitude": 106.4167,
    },
    # T√¢y Ninh
    {
        "station_id": "TN_TN2",
        "station_name": "Th√†nh ph·ªë T√¢y Ninh - T√¢y Ninh",
        "province": "T√¢y Ninh",
        "district": "T√¢y Ninh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.3131,
        "longitude": 106.0963,
    },
    {
        "station_id": "TN_TB",
        "station_name": "Huy·ªán T√¢n Bi√™n - T√¢y Ninh",
        "province": "T√¢y Ninh",
        "district": "T√¢n Bi√™n",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.5500,
        "longitude": 105.9667,
    },
    {
        "station_id": "TN_TC",
        "station_name": "Huy·ªán T√¢n Ch√¢u - T√¢y Ninh",
        "province": "T√¢y Ninh",
        "district": "T√¢n Ch√¢u",
        "type": "Huy·ªán",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 11.3333,
        "longitude": 106.1667,
    },
    # Tr√† Vinh
    {
        "station_id": "TV_TV",
        "station_name": "Th√†nh ph·ªë Tr√† Vinh - Tr√† Vinh",
        "province": "Tr√† Vinh",
        "district": "Tr√† Vinh",
        "type": "Th√†nh ph·ªë",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.9347,
        "longitude": 106.3453,
    },
    {
        "station_id": "TV_CL",
        "station_name": "Huy·ªán C√†ng Long - Tr√† Vinh",
        "province": "Tr√† Vinh",
        "district": "C√†ng Long",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.9667,
        "longitude": 106.2000,
    },
    {
        "station_id": "TV_CN",
        "station_name": "Huy·ªán C·∫ßu Ngang - Tr√† Vinh",
        "province": "Tr√† Vinh",
        "district": "C·∫ßu Ngang",
        "type": "Huy·ªán",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.8000,
        "longitude": 106.4333,
    },
    # C√°c ƒë·∫£o l·ªõn
    {
        "station_id": "HP_CB",
        "station_name": "ƒê·∫£o C√°t B√† - H·∫£i Ph√≤ng",
        "province": "H·∫£i Ph√≤ng",
        "district": "C√°t H·∫£i",
        "type": "ƒê·∫£o",
        "region": "ƒê·ªìng b·∫±ng s√¥ng H·ªìng",
        "latitude": 20.7264,
        "longitude": 107.0489,
    },
    {
        "station_id": "QN_CT",
        "station_name": "ƒê·∫£o C√¥ T√¥ - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "C√¥ T√¥",
        "type": "ƒê·∫£o",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 20.9681,
        "longitude": 107.7639,
    },
    {
        "station_id": "QN_VD",
        "station_name": "ƒê·∫£o V√¢n ƒê·ªìn - Qu·∫£ng Ninh",
        "province": "Qu·∫£ng Ninh",
        "district": "V√¢n ƒê·ªìn",
        "type": "ƒê·∫£o",
        "region": "ƒê√¥ng B·∫Øc B·ªô",
        "latitude": 20.9551,
        "longitude": 107.4764,
    },
    {
        "station_id": "KH_TS",
        "station_name": "ƒê·∫£o Tr∆∞·ªùng Sa - Kh√°nh H√≤a",
        "province": "Kh√°nh H√≤a",
        "district": "Tr∆∞·ªùng Sa",
        "type": "ƒê·∫£o",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 8.6500,
        "longitude": 111.9167,
    },
    {
        "station_id": "KH_HS",
        "station_name": "ƒê·∫£o Ho√†ng Sa - Kh√°nh H√≤a",
        "province": "Kh√°nh H√≤a",
        "district": "Ho√†ng Sa",
        "type": "ƒê·∫£o",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 16.5322,
        "longitude": 111.6156,
    },
    {
        "station_id": "KG_PQ2",
        "station_name": "ƒê·∫£o Ph√∫ Qu·ªëc - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Ph√∫ Qu·ªëc",
        "type": "ƒê·∫£o",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 10.2270,
        "longitude": 103.9679,
    },
    {
        "station_id": "KG_ND",
        "station_name": "ƒê·∫£o Nam Du - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Ki√™n H·∫£i",
        "type": "ƒê·∫£o",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.7100,
        "longitude": 104.3300,
    },
    {
        "station_id": "KG_TC",
        "station_name": "ƒê·∫£o Th·ªï Chu - Ki√™n Giang",
        "province": "Ki√™n Giang",
        "district": "Ph√∫ Qu·ªëc",
        "type": "ƒê·∫£o",
        "region": "ƒê·ªìng b·∫±ng s√¥ng C·ª≠u Long",
        "latitude": 9.3000,
        "longitude": 103.4833,
    },
    {
        "station_id": "BR_CD2",
        "station_name": "ƒê·∫£o C√¥n S∆°n - B√† R·ªãa - V≈©ng T√†u",
        "province": "B√† R·ªãa - V≈©ng T√†u",
        "district": "C√¥n ƒê·∫£o",
        "type": "ƒê·∫£o",
        "region": "ƒê√¥ng Nam B·ªô",
        "latitude": 8.6822,
        "longitude": 106.6089,
    },
    {
        "station_id": "QT_CC2",
        "station_name": "ƒê·∫£o C·ªìn C·ªè - Qu·∫£ng Tr·ªã",
        "province": "Qu·∫£ng Tr·ªã",
        "district": "C·ªìn C·ªè",
        "type": "ƒê·∫£o",
        "region": "B·∫Øc Trung B·ªô",
        "latitude": 17.1597,
        "longitude": 107.3408,
    },
    {
        "station_id": "QN_LS2",
        "station_name": "ƒê·∫£o L√Ω S∆°n - Qu·∫£ng Ng√£i",
        "province": "Qu·∫£ng Ng√£i",
        "district": "L√Ω S∆°n",
        "type": "ƒê·∫£o",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 15.3833,
        "longitude": 109.1167,
    },
    {
        "station_id": "BT_PQ2",
        "station_name": "ƒê·∫£o Ph√∫ Qu√Ω - B√¨nh Thu·∫≠n",
        "province": "B√¨nh Thu·∫≠n",
        "district": "Ph√∫ Qu√Ω",
        "type": "ƒê·∫£o",
        "region": "Duy√™n h·∫£i Nam Trung B·ªô",
        "latitude": 10.5000,
        "longitude": 108.9333,
    },
]


def main():
    """H√†m ch√≠nh th·ª±c thi"""
    try:
        crawler = VietnamWeatherDataCrawler()
        locations = vietnam_locations

        logging.info("=" * 70)
        logging.info("üåè H·ªÜ TH·ªêNG THU TH·∫¨P D·ªÆ LI·ªÜU TH·ªúI TI·∫æT VI·ªÜT NAM")
        logging.info("=" * 70)
        logging.info(
            "üìù L∆ØU √ù: D·ªØ li·ªáu t·ª´ c√°c API mi·ªÖn ph√≠ c√≥ th·ªÉ kh√¥ng ch√≠nh x√°c tuy·ªát ƒë·ªëi"
        )
        logging.info("üîç ƒêang thu th·∫≠p t·ª´ ƒëa ngu·ªìn v·ªõi ƒë√°nh gi√° ch·∫•t l∆∞·ª£ng...")

        start_time = time.time()
        weather_data = crawler.crawl_all_locations(locations, delay=2.0)
        end_time = time.time()

        if weather_data:
            excel_file = crawler.save_to_excel(weather_data)

            sqlite_success = crawler.save_to_sqlite(weather_data, locations)

            if sqlite_success:
                db_summary = crawler.get_database_summary()

            quality_report = crawler.get_data_quality_report()

            logging.info("=" * 70)
            logging.info("üìä B√ÅO C√ÅO CH·∫§T L∆Ø·ª¢NG D·ªÆ LI·ªÜU")
            logging.info("=" * 70)

            w_report = quality_report["weather"]
            logging.info(f"üå°Ô∏è  D·ªÆ LI·ªÜU TH·ªúI TI·∫æT:")
            logging.info(
                f"   ‚úÖ Ch·∫•t l∆∞·ª£ng cao: {w_report['high_quality']}/{w_report['total']} ({w_report['high_percent']}%)"
            )
            logging.info(
                f"   ‚ö†Ô∏è  Ch·∫•t l∆∞·ª£ng TB: {w_report['medium_quality']}/{w_report['total']} ({w_report['medium_percent']}%)"
            )
            logging.info(
                f"   ‚ùå Ch·∫•t l∆∞·ª£ng th·∫•p: {w_report['low_quality']}/{w_report['total']} ({w_report['low_percent']}%)"
            )

            if sqlite_success and db_summary:
                logging.info("=" * 70)
                logging.info("üóÉÔ∏è  T·ªîNG QUAN DATABASE")
                logging.info("=" * 70)
                logging.info(
                    f"üìä T·ªïng s·ªë b·∫£n ghi: {db_summary.get('total_records', 0)}"
                )
                logging.info(
                    f"üèôÔ∏è  S·ªë t·ªânh th√†nh: {db_summary.get('total_provinces', 0)}"
                )
                logging.info(
                    f"üïí D·ªØ li·ªáu m·ªõi nh·∫•t: {db_summary.get('latest_data', 'N/A')}"
                )
                if "quality_stats" in db_summary:
                    logging.info("üìà Ch·∫•t l∆∞·ª£ng d·ªØ li·ªáu:")
                    for quality, count in db_summary["quality_stats"].items():
                        logging.info(f"   {quality}: {count} b·∫£n ghi")

            logging.info("=" * 70)
            logging.info(f"‚è±Ô∏è Th·ªùi gian th·ª±c hi·ªán: {end_time - start_time:.2f} gi√¢y")
            logging.info(f"üìÅ File Excel: {excel_file}")
            logging.info(f"üóÑÔ∏è  Database SQLite: vietnam_weather.db")
            logging.info(
                "üéØ L∆ØU √ù: ƒê·ªÉ c√≥ d·ªØ li·ªáu ch√≠nh x√°c h∆°n, c·∫ßn s·ª≠ d·ª•ng API c√≥ ph√≠ ho·∫∑c truy c·∫≠p tr·ª±c ti·∫øp"
            )
            logging.info("   c√°c ngu·ªìn d·ªØ li·ªáu ch√≠nh th·ª©c c·ªßa Vi·ªát Nam")

        else:
            logging.warning("‚ùå Kh√¥ng thu th·∫≠p ƒë∆∞·ª£c d·ªØ li·ªáu n√†o")

    except Exception as e:
        logging.error(f"üí• L·ªói h·ªá th·ªëng: {e}")


def run_continuously():
    """H√†m th∆∞·ªùng tr√∫ ƒë·ªÉ ch·∫°y main() l·∫∑p l·∫°i c·ª© 10 ph√∫t m·ªôt l·∫ßn"""
    while True:
        try:
            main()
            logging.info("‚è≥ ƒêang ch·ªù 10 ph√∫t ƒë·ªÉ ch·∫°y l·∫ßn ti·∫øp theo...")
            time.sleep(600)
        except Exception as e:
            logging.error(f"üí• L·ªói trong qu√° tr√¨nh ch·∫°y th∆∞·ªùng tr√∫: {e}")
            logging.info("üîÑ Th·ª≠ ch·∫°y l·∫°i sau 10 ph√∫t...")
            time.sleep(600)



if __name__ == "__main__":
    if CRAWL_MODE == "once":
        logging.info("üöÄ Ch·∫°y ch·∫ø ƒë·ªô 1 l·∫ßn (once)")
        main()
    else:
        logging.info("üîÅ Ch·∫°y ch·∫ø ƒë·ªô th∆∞·ªùng tr√∫ (continuous)")
        run_continuously()