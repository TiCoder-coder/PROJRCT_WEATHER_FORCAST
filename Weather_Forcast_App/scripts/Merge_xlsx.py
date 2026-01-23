import sys
import re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")


MERGE_DIR_NAME = "Merge_data"
OUTPUT_DIR_NAME = "output"

MERGE_FILENAME = "merged_vrain_data.xlsx"
MERGE_VIETNAM_FILENAME = "merged_vietnam_weather_data.xlsx"

LOG_FILENAME = "merged_files_log.txt"
LOG_VIETNAM_FILENAME = "merged_vietnam_files_log.txt"

SAVE_EVERY_N_ROWS = 30000

MASTER_COLUMNS = [
    "Mã trạm", "Tên trạm", "Tỉnh/Thành phố", "Huyện", "Vĩ độ", "Kinh độ",
    "Dấu thời gian", "Nguồn dữ liệu", "Chất lượng dữ liệu", "Thời gian cập nhật",
    "Nhiệt độ hiện tại", "Nhiệt độ tối đa", "Nhiệt độ tối thiểu", "Nhiệt độ trung bình",
    "Độ ẩm hiện tại", "Độ ẩm tối đa", "Độ ẩm tối thiểu", "Độ ẩm trung bình",
    "Áp suất hiện tại", "Áp suất tối đa", "Áp suất tối thiểu", "Áp suất trung bình",
    "Tốc độ gió hiện tại", "Tốc độ gió tối đa", "Tốc độ gió tối thiểu", "Tốc độ gió trung bình",
    "Hướng gió hiện tại", "Hướng gió trung bình",
    "Lượng mưa hiện tại", "Lượng mưa tối đa", "Lượng mưa tối thiểu", "Lượng mưa trung bình",
    "Tổng lượng mưa",
    "Độ che phủ mây hiện tại", "Độ che phủ mây tối đa", "Độ che phủ mây tổi thiểu", "Độ che phủ mây trung bình",
    "Tầm nhìn hiện tại", "Tầm nhìn đa", "Tầm nhìn tối thiểu", "Tầm nhìn trung bình",
    "Xác xuất sấm sét", "Tình trạng",
]

COLUMN_ALIASES = {
    "Độ che phủ mây tối thiểu": "Độ che phủ mây tổi thiểu",
    "Tầm nhìn tối đa": "Tầm nhìn đa",
    "Xác suất sấm sét": "Xác xuất sấm sét",
    "Xác suất sét": "Xác xuất sấm sét",
    "Xác suất sấm sét": "Xác xuất sấm sét",
}


def norm_col(x) -> str:
    s = re.sub(r"\s+", " ", str(x).strip())
    return COLUMN_ALIASES.get(s, s)


def load_processed_files(log_path: Path) -> set[str]:
    if not log_path.exists():
        return set()
    processed = set()
    try:
        with log_path.open("r", encoding="utf-8") as f:
            for line in f:
                name = line.strip()
                if name:
                    processed.add(name)
    except Exception as e:
        print(f"Loi khi doc log file {log_path.name}: {e}")
    return processed


def save_processed_files(log_path: Path, processed_files: set[str]) -> None:
    try:
        with log_path.open("w", encoding="utf-8") as f:
            for name in sorted(processed_files):
                f.write(name + "\n")
    except Exception as e:
        print(f"Loi khi ghi log file {log_path.name}: {e}")


def read_excel_file(file_path: Path) -> pd.DataFrame:
    try:
        df = pd.read_excel(file_path)
        if df is None or df.empty:
            return pd.DataFrame()
        return df
    except Exception as e:
        print(f"Loi khi doc file {file_path.name}: {e}")
        return pd.DataFrame()


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df.columns = [norm_col(c) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]

    drop_cols = [c for c in df.columns if str(c).lower().startswith("unnamed")]
    if drop_cols:
        df = df.drop(columns=drop_cols, errors="ignore")

    return df


def get_new_excel_files(output_dir: Path, processed_files: set[str]) -> tuple[list[Path], list[Path]]:
    if not output_dir.exists():
        print(f"Thu muc nguon khong ton tai: {output_dir}")
        return [], []

    all_excel_files = sorted(output_dir.glob("*.xlsx"))
    if not all_excel_files:
        print(f"Khong tim thay file .xlsx nao trong thu muc: {output_dir}")
        return [], []

    vietnam_files, other_files = [], []
    for file_path in all_excel_files:
        if file_path.name in processed_files:
            continue
        if file_path.name.startswith("vietnam_weather_"):
            vietnam_files.append(file_path)
        else:
            other_files.append(file_path)

    print(f"Tong so file .xlsx trong output: {len(all_excel_files)}")
    print(f"So file vietnam_weather_ moi: {len(vietnam_files)}")
    print(f"So file khac moi: {len(other_files)}")
    return vietnam_files, other_files


def _load_or_create_wb_ws(merge_path: Path) -> tuple[Workbook, Worksheet, list[str]]:
    if merge_path.exists():
        wb = load_workbook(merge_path)
        ws = wb.active

        header = []
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            v = ws.cell(row=1, column=col).value
            if v is None:
                continue
            header.append(norm_col(v))

        if not header:
            header = list(MASTER_COLUMNS)
            for i, col_name in enumerate(header, start=1):
                ws.cell(row=1, column=i).value = col_name

        return wb, ws, header

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    header = list(MASTER_COLUMNS)
    for i, col_name in enumerate(header, start=1):
        ws.cell(row=1, column=i).value = col_name

    wb.save(merge_path)
    return wb, ws, header


def _ensure_header_has_columns(ws: Worksheet, header: list[str], cols_to_ensure: list[str]) -> list[str]:
    changed = False
    for c in cols_to_ensure:
        c = norm_col(c)
        if c not in header:
            header.append(c)
            ws.cell(row=1, column=len(header)).value = c
            changed = True
    if changed:
        print(f"  + Da mo rong schema, tong so cot hien tai: {len(header)}")
    return header


def _to_excel_value(v):
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    return v


def append_df_incremental(
    wb: Workbook,
    ws: Worksheet,
    header: list[str],
    df: pd.DataFrame,
    merge_path: Path,
    save_every_n_rows: int = SAVE_EVERY_N_ROWS,
) -> int:
    for c in header:
        if c not in df.columns:
            df[c] = pd.NA

    df = df[header]

    appended = 0
    buffer_count = 0

    for row in df.itertuples(index=False, name=None):
        excel_row = [_to_excel_value(x) for x in row]
        ws.append(excel_row)
        appended += 1
        buffer_count += 1

        if save_every_n_rows > 0 and buffer_count >= save_every_n_rows:
            wb.save(merge_path)
            buffer_count = 0

    wb.save(merge_path)
    return appended


def merge_single_category_incremental(
    file_list: list[Path],
    merge_path: Path,
    log_path: Path,
    processed_files: set[str],
    category_name: str,
) -> None:
    if not file_list:
        print(f"Khong co file {category_name} moi de merge.")
        return

    print(f"\n=== MERGE INCREMENTAL: {category_name.upper()} ===")
    print(f"So luong file: {len(file_list)}")
    print(f"File merge: {merge_path}")

    wb, ws, header = _load_or_create_wb_ws(merge_path)

    header = _ensure_header_has_columns(ws, header, MASTER_COLUMNS)
    wb.save(merge_path)

    ok_count = 0
    for idx, file_path in enumerate(sorted(file_list), start=1):
        print(f"\n[{idx}/{len(file_list)}] Dang xu ly: {file_path.name}")

        df = read_excel_file(file_path)
        if df.empty:
            print("  - File rong/khong doc duoc, bo qua.")
            continue

        df = clean_dataframe(df)

        new_cols = [c for c in df.columns if c not in header]
        if new_cols:
            print(f"  + Phat hien {len(new_cols)} cot moi: {new_cols}")
            header = _ensure_header_has_columns(ws, header, new_cols)
            wb.save(merge_path)

        try:
            appended = append_df_incremental(
                wb=wb,
                ws=ws,
                header=header,
                df=df,
                merge_path=merge_path,
                save_every_n_rows=SAVE_EVERY_N_ROWS,
            )
            print(f"  ✓ Da append {appended} dong tu {file_path.name}")

            processed_files.add(file_path.name)
            save_processed_files(log_path, processed_files)
            ok_count += 1
        except Exception as e:
            print(f"  ✗ Loi khi append file {file_path.name}: {e}")
            print("  -> Khong danh dau processed (de lan sau chay lai).")

    wb.save(merge_path)
    print(f"\n=== XONG {category_name}: OK {ok_count}/{len(file_list)} file ===")


def merge_excel_files_once(base_dir: Path) -> None:
    output_dir = base_dir / OUTPUT_DIR_NAME
    merge_dir = base_dir / MERGE_DIR_NAME
    merge_dir.mkdir(parents=True, exist_ok=True)

    merge_vietnam_path = merge_dir / MERGE_VIETNAM_FILENAME
    log_vietnam_path = merge_dir / LOG_VIETNAM_FILENAME

    merge_other_path = merge_dir / MERGE_FILENAME
    log_other_path = merge_dir / LOG_FILENAME

    print("======== BAT DAU MERGE =========")
    print(f"Thu muc nguon (output):     {output_dir}")
    print(f"Thu muc merge (Merge_data): {merge_dir}")
    print("================================")

    processed_vietnam = load_processed_files(log_vietnam_path)
    processed_other = load_processed_files(log_other_path)

    processed_all = processed_vietnam.union(processed_other)
    vietnam_files, other_files = get_new_excel_files(output_dir, processed_all)

    merge_single_category_incremental(
        vietnam_files,
        merge_vietnam_path,
        log_vietnam_path,
        processed_vietnam,
        "vietnam_weather_"
    )

    merge_single_category_incremental(
        other_files,
        merge_other_path,
        log_other_path,
        processed_other,
        "khac"
    )

    print("\n======== KET THUC MERGE =========")


if __name__ == "__main__":
    SCRIPT_DIR = Path(__file__).parent
    BASE_DIR = SCRIPT_DIR.parent

    print(f"Script dir: {SCRIPT_DIR}")
    print(f"Base dir:   {BASE_DIR}")

    output_dir = BASE_DIR / OUTPUT_DIR_NAME
    if not output_dir.exists():
        print(f"ERROR: Khong tim thay thu muc output tai: {output_dir}")
        sys.exit(1)

    merge_excel_files_once(BASE_DIR)
